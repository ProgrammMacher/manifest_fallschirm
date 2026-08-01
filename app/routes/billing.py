from decimal import Decimal
from collections import defaultdict
from datetime import datetime, date, timedelta, time as dt_time
import re
import os
import base64
import io
import csv
import time
from xml.sax.saxutils import escape as xml_escape
from sqlalchemy import or_, and_, func
from sqlalchemy.exc import IntegrityError
from werkzeug.routing.exceptions import BuildError

from flask import (
    Blueprint, render_template, redirect, url_for,
    flash, request, session, current_app, get_flashed_messages, jsonify, make_response
)
from sqlalchemy.orm import selectinload, joinedload
import threading

from app import db, now_local, create_app
now_berlin = now_local  # Alias für Abwärtskompatibilität
from app.helpers.email_progress import set_progress, get_progress, clear_progress, mark_complete, get_active_job_ids
from app.models.person import Person
from app.models.invoice import Invoice
from app.models.invoice_item import InvoiceItem
from app.models.billing_config import BillingConfig, MANUAL_MAIL_BODY_TEMPLATE_DEFAULT
from app.models.load_entry import LoadEntry
from app.models.load import Load  # ✅ benötigt für Load.status Filter in invoice_list()
from app.models.sepa_config import SepaConfig
from app.models.sepa_export import SepaExport, SepaExportInvoice
from app.services.billing_service import BillingService, _image_to_data_uri, _invoice_payment_label
from app.services.payment_data_service import build_invoice_payment_purpose, build_payment_context
from app.services.sepa_export_service import build_pain_008_xml
from app.services.mailer_service import MailerService
from app.services.pdf_service import generate_pdf_from_html
from app.models.billing_config import BillingOrgaRule
from app.helpers.status_code import normalize_status_code
from app.security.credentials import get_runtime_home_dir
from app.constants import TANDEM_GUEST_STATUSES, TM_STATUSES, VIDEO_STATUSES
from app.constants import (
    INVOICE_PAYMENT_STATE_OPEN,
    INVOICE_PAYMENT_STATE_SEPA_PENDING,
    INVOICE_PAYMENT_STATE_SEPA_EXPORTED,
    INVOICE_PAYMENT_STATE_PAID,
    INVOICE_PAYMENT_STATE_SEPA_RETURNED,
    INVOICE_PAYMENT_STATES,
)

bp = Blueprint("billing", __name__, url_prefix="/billing")


def _set_no_store_headers(response) -> None:
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"


def _build_invoice_list_redirect_response(*, export_id: int | None = None):
    params = {}
    if export_id is not None:
        params["sepa_download_export_id"] = export_id

    try:
        location = url_for("billing.invoice_list", **params)
    except BuildError:
        location = "/billing/invoices"
        if params:
            sep = "?" if "?" not in location else "&"
            location = f"{location}{sep}{'&'.join(f'{k}={v}' for k, v in params.items())}"

    response = redirect(location)
    _set_no_store_headers(response)
    return response


PAYMENT_STATE_LABELS = {
    INVOICE_PAYMENT_STATE_OPEN: "Offen",
    INVOICE_PAYMENT_STATE_SEPA_PENDING: "SEPA vorgemerkt",
    INVOICE_PAYMENT_STATE_SEPA_EXPORTED: "SEPA exportiert",
    INVOICE_PAYMENT_STATE_PAID: "Bezahlt",
    INVOICE_PAYMENT_STATE_SEPA_RETURNED: "Rücklastschrift",
}

_TANDEM_GUEST_STATUS_CODES = {
    normalize_status_code(code) for code in TANDEM_GUEST_STATUSES
}


def _person_allows_sepa(person: Person | None) -> bool:
    if not person:
        return False
    if bool(getattr(person, "is_tandem_guest", False)):
        return False
    if not bool(getattr(person, "sepa_enabled", False)):
        return False
    if not (getattr(person, "iban", "") or "").strip():
        return False
    if not (getattr(person, "account_holder", "") or "").strip():
        return False
    if getattr(person, "sepa_mandate_date", None) is None:
        return False
    return True


def _invoice_has_tandem_guest_context(invoice: Invoice | None) -> bool:
    if not invoice:
        return False

    for item in getattr(invoice, "items", []) or []:
        load_entry = getattr(item, "load_entry", None)
        if not load_entry:
            continue
        status_code = normalize_status_code(getattr(load_entry, "status_code", ""))
        if status_code in _TANDEM_GUEST_STATUS_CODES:
            return True

    return False


def _invoice_allows_sepa(invoice: Invoice | None) -> bool:
    if not invoice:
        return False
    if not _person_allows_sepa(getattr(invoice, "person", None)):
        return False
    if _invoice_has_tandem_guest_context(invoice):
        return False
    return True


def _invoice_payment_state(invoice: Invoice | None) -> str:
    if not invoice:
        return INVOICE_PAYMENT_STATE_OPEN

    if bool(getattr(invoice, "is_paid", False)):
        return INVOICE_PAYMENT_STATE_PAID

    raw = (getattr(invoice, "payment_state", "") or "").strip().lower()
    if raw in INVOICE_PAYMENT_STATES:
        if raw == INVOICE_PAYMENT_STATE_PAID:
            return INVOICE_PAYMENT_STATE_OPEN
        if (
            raw in {
                INVOICE_PAYMENT_STATE_SEPA_PENDING,
                INVOICE_PAYMENT_STATE_SEPA_EXPORTED,
                INVOICE_PAYMENT_STATE_SEPA_RETURNED,
            }
            and not _invoice_allows_sepa(invoice)
        ):
            return INVOICE_PAYMENT_STATE_OPEN
        return raw

    if (getattr(invoice, "payment_method", "") or "").strip().lower() == "sepa" and _invoice_allows_sepa(invoice):
        return INVOICE_PAYMENT_STATE_SEPA_PENDING
    return INVOICE_PAYMENT_STATE_OPEN


def _invoice_payment_state_label(invoice: Invoice | None) -> str:
    return PAYMENT_STATE_LABELS.get(_invoice_payment_state(invoice), "Offen")


def _reset_invoice_after_sepa_rollback(invoice: Invoice) -> None:
    """Setzt eine Rechnung nach einem Dev-Rollback wieder auf einen neutralen offenen Zustand zurück."""
    _set_invoice_payment_state(invoice, INVOICE_PAYMENT_STATE_OPEN)
    invoice.payment_method = None
    invoice.is_paid = False
    invoice.paid_at = None


def _set_invoice_payment_state(invoice: Invoice, payment_state: str) -> str:
    state = (payment_state or "").strip().lower()
    if state not in INVOICE_PAYMENT_STATES:
        state = INVOICE_PAYMENT_STATE_OPEN

    if state in {
        INVOICE_PAYMENT_STATE_SEPA_PENDING,
        INVOICE_PAYMENT_STATE_SEPA_EXPORTED,
        INVOICE_PAYMENT_STATE_SEPA_RETURNED,
    } and not _invoice_allows_sepa(invoice):
        state = INVOICE_PAYMENT_STATE_OPEN

    invoice.payment_state = state
    if state == INVOICE_PAYMENT_STATE_PAID:
        invoice.is_paid = True
        if not getattr(invoice, "paid_at", None):
            invoice.paid_at = now_berlin().replace(tzinfo=None)
    else:
        invoice.is_paid = False
        invoice.paid_at = None

    if state in {INVOICE_PAYMENT_STATE_SEPA_PENDING, INVOICE_PAYMENT_STATE_SEPA_EXPORTED, INVOICE_PAYMENT_STATE_SEPA_RETURNED}:
        invoice.payment_method = "sepa"

    return state


def _is_allowed_payment_state_transition(current_state: str, next_state: str) -> bool:
    allowed = {
        INVOICE_PAYMENT_STATE_OPEN: {
            INVOICE_PAYMENT_STATE_OPEN,
            INVOICE_PAYMENT_STATE_SEPA_PENDING,
            INVOICE_PAYMENT_STATE_PAID,
        },
        INVOICE_PAYMENT_STATE_SEPA_PENDING: {
            INVOICE_PAYMENT_STATE_OPEN,
            INVOICE_PAYMENT_STATE_SEPA_PENDING,
            INVOICE_PAYMENT_STATE_SEPA_EXPORTED,
            INVOICE_PAYMENT_STATE_PAID,
        },
        INVOICE_PAYMENT_STATE_SEPA_EXPORTED: {
            INVOICE_PAYMENT_STATE_OPEN,
            INVOICE_PAYMENT_STATE_SEPA_EXPORTED,
            INVOICE_PAYMENT_STATE_SEPA_RETURNED,
            INVOICE_PAYMENT_STATE_PAID,
        },
        INVOICE_PAYMENT_STATE_SEPA_RETURNED: {
            INVOICE_PAYMENT_STATE_OPEN,
            INVOICE_PAYMENT_STATE_SEPA_PENDING,
            INVOICE_PAYMENT_STATE_SEPA_RETURNED,
            INVOICE_PAYMENT_STATE_PAID,
        },
        INVOICE_PAYMENT_STATE_PAID: {
            INVOICE_PAYMENT_STATE_OPEN,
            INVOICE_PAYMENT_STATE_SEPA_RETURNED,
            INVOICE_PAYMENT_STATE_PAID,
        },
    }
    return next_state in allowed.get(current_state, {INVOICE_PAYMENT_STATE_OPEN})


def _billing_person_main_status(person: Person) -> str:
    if bool(getattr(person, "is_tandem_guest", False)):
        return "Tandemgast"
    if bool(getattr(person, "is_partner_verein", False)):
        return "Partner-Verein"
    if bool(getattr(person, "is_member", False)):
        return "Verein"
    return "Gast"


def _billing_person_secondary_statuses(raw_statuses: set[str], main_status: str) -> list[str]:
    status_label_map = {
        "TD": "Tandemmaster",
        "TD-Vereins-Schirm": "Tandemmaster",
        "Videomann": "Video",
        "G-TD": "Tandemgast",
        "G-TD-Video": "Tandemgast",
    }
    excluded = {"Verein", "Gast", "Partner-Verein", "Tandemgast", (main_status or "").strip()}

    labels = []
    for status in sorted((s or "").strip() for s in (raw_statuses or set()) if (s or "").strip()):
        label = status_label_map.get(status, status)
        if label in excluded:
            continue
        if label not in labels:
            labels.append(label)

    return labels


def _full_admin_required(redirect_endpoint: str, **redirect_values):
    if session.get("is_admin"):
        return None

    if session.get("is_db_admin"):
        flash(
            "Voll-Admin-Rechte erforderlich. Aktuell sind Sie als Datenbank-Admin angemeldet.",
            "warning",
        )
    else:
        flash("Voll-Admin-Rechte erforderlich. Bitte mit dem Admin-Passwort anmelden.", "danger")

    return redirect(url_for(redirect_endpoint, **redirect_values))


def _admin_or_db_admin_required(redirect_endpoint: str, **redirect_values):
    if session.get("is_admin") or session.get("is_db_admin"):
        return None

    flash("Admin- oder Datenbank-Admin-Rechte erforderlich.", "warning")
    return redirect(url_for(redirect_endpoint, **redirect_values))


def _can_manage_sepa_exports() -> bool:
    return bool(session.get("is_admin") or session.get("is_db_admin"))


def _current_admin_actor_label() -> str:
    if session.get("is_admin"):
        return "admin"
    if session.get("is_db_admin"):
        return "db_admin"
    return "user"


def _sepa_export_storage_dir() -> str:
    root = get_runtime_home_dir()
    path = os.path.join(root, "data", "sepa_exports")
    os.makedirs(path, exist_ok=True)
    return path


def _invoice_load_snapshot(invoice: Invoice) -> tuple[date | None, date | None, str]:
    load_dates: list[date] = []

    for item in list(getattr(invoice, "items", []) or []):
        le = getattr(item, "load_entry", None)
        if not le:
            continue
        ld = getattr(le, "load", None)
        if not ld:
            continue

        op_date = getattr(ld, "operation_date", None)
        if op_date is None:
            dt = getattr(ld, "actual_time", None) or getattr(ld, "scheduled_time", None) or getattr(ld, "created_at", None)
            if dt is not None:
                op_date = dt.date()

        if isinstance(op_date, date):
            load_dates.append(op_date)

    if not load_dates:
        return None, None, "Manuelle Rechnung"

    unique_sorted = sorted(set(load_dates))
    from_d = unique_sorted[0]
    to_d = unique_sorted[-1]
    dates_txt = ", ".join(d.strftime("%d.%m.%Y") for d in unique_sorted)
    return from_d, to_d, dates_txt


def _next_sepa_export_code() -> tuple[str, int]:
    year = now_berlin().year
    prefix = f"{year}-"
    rows = (
        db.session.query(SepaExport.export_code)
        .filter(SepaExport.export_code.like(f"{prefix}%"))
        .all()
    )

    max_no = 0
    for (code,) in rows:
        try:
            n = int(str(code).split("-", 1)[1])
            if n > max_no:
                max_no = n
        except Exception:
            continue

    next_no = max_no + 1
    return f"{year}-{next_no:04d}", next_no


def _build_export_file_name(created_at: datetime, export_seq_no: int) -> str:
    return f"SEPA_{created_at.strftime('%Y-%m-%d_%H%M%S')}_Export{export_seq_no:04d}.xml"


def _build_sepa_export_placeholder_xml(export_code: str, created_at: datetime, rows: list[dict]) -> bytes:
    sepa_config = SepaConfig.query.order_by(SepaConfig.id.asc()).first()
    if not sepa_config:
        sepa_config = SepaConfig(
            creditor_id="",
            creditor_name="",
            creditor_iban="",
            creditor_bic="",
            creditor_country="DE",
            pain_version="pain.008.001.02",
        )
    collection_date = (created_at + timedelta(days=3)).date()
    billing_config = BillingConfig.query.first()
    return build_pain_008_xml(
        export_code=export_code,
        created_at=created_at,
        rows=rows,
        sepa_config=sepa_config,
        collection_date=collection_date,
        billing_config=billing_config,
    )


WAIVER_TEXT_SKYDIVER_DEFAULT = """Erklärung und Haftungsverzicht
Ich erkläre verbindlich, dass ich mir der besonderen Risiken des Fallschirmspringens bewusst bin. Mir ist bekannt, dass es trotz ordnungsgemäßer Durchführung, Einhaltung aller Sicherheitsvorschriften sowie sorgfältiger Organisation zu Unfällen, Verletzungen oder sonstigen Schäden kommen kann.
Für den Fall eines Unfalls oder sonstiger Schäden während meiner Teilnahme am gesamten Sprungbetrieb verzichte ich – soweit gesetzlich zulässig – auf sämtliche Schadensersatzansprüche materieller und immaterieller Art gegenüber dem Veranstalter bzw. Genehmigungsinhaber sowie dessen gesetzlichen Vertretern, Mitarbeitern, Beauftragten und Erfüllungsgehilfen, der Sprungbetriebsleitung, dem Luftfahrzeughalter sowie der eingesetzten Flugbesatzung.
Dieser Haftungsverzicht gilt auch für Ansprüche aus der Haltereigenschaft und dem Betrieb der eingesetzten Luftfahrzeuge sowie der im Sprungbetrieb verwendeten Ausrüstung.
Die Haftung bei Vorsatz und grober Fahrlässigkeit bleibt unberührt.
Diese Erklärung gilt ebenfalls für etwaige Ansprüche Dritter, insbesondere von unterhaltsberechtigten Personen oder solchen, auf die Ansprüche aus einem Schadensfall übergehen könnten.

Eigenerklärung zu Voraussetzungen und Eigenverantwortung
Ich erkläre weiterhin verbindlich:
- dass ich über eine gültige Fallschirmsprunglizenz sowie einen ausreichenden Versicherungsschutz gemäß den gesetzlichen Anforderungen verfüge,
- dass ich körperlich und geistig in der Lage bin, am Fallschirmsprungbetrieb teilzunehmen,
- dass ich nicht unter Alkohol-, Drogen- oder Medikamenteneinfluss stehe, der meine Leistungsfähigkeit beeinträchtigen könnte,
- dass ich die geltenden Platz-, Flug- und Sprungbetriebsregeln kenne und diese während des gesamten Aufenthalts einhalte,
- dass bei der Durchführung des Sprungbetriebs sowie bei der verwendeten Ausrüstung alle gesetzlichen, luftrechtlichen sowie allgemein anerkannten sicherheitsrelevanten Vorgaben eingehalten werden.

Die Sprungbetriebsleitung ist berechtigt, meine Teilnahme am Sprungbetrieb aus Sicherheitsgründen zu untersagen.

Schlussbestimmungen
Sollten einzelne Bestimmungen dieser Erklärung ganz oder teilweise unwirksam sein oder werden, bleibt die Wirksamkeit der übrigen Bestimmungen unberührt."""

WAIVER_TEXT_TANDEM_DEFAULT = """Erklärung zur Teilnahme am Tandem-Fallschirmsprung
Ich erkläre, dass ich freiwillig an einem Tandem-Fallschirmsprung teilnehme. Mir ist bewusst, dass Fallschirmspringen – auch bei größtmöglicher Sorgfalt – mit erheblichen Risiken verbunden ist und es zu Verletzungen bis hin zu schweren oder tödlichen Unfällen kommen kann.

Mir ist bekannt, dass ich beim Tandemsprung fest mit einem entsprechend ausgebildeten und zugelassenen Tandemmaster verbunden bin und die Durchführung des Sprunges vollständig nach dessen Anweisungen erfolgt.

Haftungsverzicht
Für den Fall von Unfällen, Verletzungen oder sonstigen Schäden verzichte ich – soweit gesetzlich zulässig – auf sämtliche Schadensersatzansprüche materieller und immaterieller Art gegenüber dem Veranstalter bzw. Genehmigungsinhaber sowie dessen gesetzlichen Vertretern, Mitarbeitern, Tandemmastern, Beauftragten und Erfüllungsgehilfen, der Sprungbetriebsleitung, dem Luftfahrzeughalter sowie der eingesetzten Flugbesatzung.

Dieser Haftungsverzicht gilt auch für Ansprüche aus der Haltereigenschaft und dem Betrieb der eingesetzten Luftfahrzeuge sowie der beim Tandemsprung verwendeten Ausrüstung.

Die Haftung bei Vorsatz und grober Fahrlässigkeit bleibt unberührt.

Diese Erklärung gilt ebenfalls für etwaige Ansprüche Dritter, insbesondere von unterhaltsberechtigten Personen oder solchen, auf die Ansprüche aus einem Schadensfall übergehen könnten.

Eigenverantwortung und Voraussetzungen
Ich erkläre weiterhin:
- dass ich körperlich und geistig gesund bin und keine mir bekannten gesundheitlichen Einschränkungen habe, die gegen eine Teilnahme sprechen,
- dass ich relevante gesundheitliche Besonderheiten wahrheitsgemäß angegeben habe,
- dass ich nicht unter Alkohol-, Drogen- oder beeinträchtigenden Medikamenteneinfluss stehe,
- dass ich die Anweisungen des Tandemmasters sowie der Sprungbetriebsleitung jederzeit befolge,
- dass bei der Durchführung des Tandemsprungs sowie bei der verwendeten Ausrüstung alle gesetzlichen, luftrechtlichen sowie allgemein anerkannten sicherheitsrelevanten Vorgaben eingehalten werden.

Der Veranstalter bzw. der Tandemmaster ist berechtigt, den Sprung aus Sicherheitsgründen auch kurzfristig abzusagen oder abzubrechen.

Schlussbestimmungen
Sollten einzelne Bestimmungen dieser Erklärung ganz oder teilweise unwirksam sein oder werden, bleibt die Wirksamkeit der übrigen Bestimmungen unberührt."""


# ---------------------------------------------------------
# EPC-QR (SEPA) dynamisch erzeugen (pro Rechnung)
# ---------------------------------------------------------
def _build_epc_payload(*, bic: str, name: str, iban: str, amount_eur: Decimal, remittance: str) -> str:
    """Erzeugt einen EPC/SEPA-QR Payload (UTF-8, Zeilenformat)."""
    amount_str = f"{amount_eur:.2f}"  # Punkt als Dezimaltrenner
    rem = (remittance or "").strip()
    # EPC-Remittance ist i.d.R. auf 140 Zeichen begrenzt
    if len(rem) > 140:
        rem = rem[:140]

    lines = [
        "BCD",
        "002",  # Version
        "1",    # UTF-8
        "SCT",
        (bic or "").strip(),
        (name or "").strip(),
        (iban or "").replace(" ", "").strip(),
        f"EUR{amount_str}",
        "",      # Purpose (optional)
        rem,     # Verwendungszweck / Remittance
        ""       # Info (optional)
    ]
    return "\n".join(lines)


def _make_qr_data_uri(payload: str) -> str:
    """Erzeugt einen QR als PNG und liefert ihn als data:image/png;base64,... URI.
    Keine Datei auf der Platte, kein statisches Asset.
    """
    import qrcode

    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=4,
        border=2,
    )
    qr.add_data(payload)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/png;base64,{b64}"


def _short_airfield_place(name: str) -> str:
    """Extrahiert aus Flugplatznamen moeglichst den Ortsanteil fuer den Verwendungszweck."""
    raw = (name or "").strip()
    if not raw:
        return ""

    lowered = raw.lower()
    prefixes = (
        "flugplatz ",
        "flugfeld ",
        "airfield ",
        "airport ",
        "dz ",
        "dropzone ",
    )
    for prefix in prefixes:
        if lowered.startswith(prefix):
            raw = raw[len(prefix):].strip(" -")
            break

    for sep in (" - ", " / ", ", ", "("):
        if sep in raw:
            raw = raw.split(sep, 1)[0].strip()
            break

    # Gewuenscht: nur der Ortsname bis zum ersten Leerzeichen.
    raw = raw.split()[0] if raw.split() else ""

    return raw


def _invoice_airfield_and_date_range(invoice: Invoice):
    """Liest Flugplatz und Datumsbereich aus den in der Rechnung enthaltenen Spruengen."""
    airfields = set()
    jump_dates = []

    for item in list(getattr(invoice, "items", []) or []):
        le = getattr(item, "load_entry", None)
        ld = getattr(le, "load", None) if le else None
        if not ld:
            continue

        airfield = getattr(ld, "airfield", None)
        airfield_name = getattr(airfield, "name", "") if airfield else ""
        short_name = _short_airfield_place(airfield_name)
        if short_name:
            airfields.add(short_name)

        dt_value = getattr(ld, "actual_time", None) or getattr(ld, "scheduled_time", None) or getattr(ld, "created_at", None)
        if dt_value:
            jump_dates.append(dt_value.date())

    airfield_text = ""
    if len(airfields) == 1:
        airfield_text = next(iter(airfields))
    elif len(airfields) > 1:
        airfield_text = "mehrere Orte"

    start_date = min(jump_dates) if jump_dates else None
    end_date = max(jump_dates) if jump_dates else None
    return airfield_text, start_date, end_date


def _format_purpose_date_range(start_date, end_date) -> str:
    if not start_date or not end_date:
        return ""
    if start_date == end_date:
        return f"vom {start_date.day}.{start_date.month}.{start_date.year}"
    if start_date.year == end_date.year:
        return (
            f"vom {start_date.day}.{start_date.month}. bis "
            f"{end_date.day}.{end_date.month}.{end_date.year}"
        )
    return (
        f"vom {start_date.day}.{start_date.month}.{start_date.year} bis "
        f"{end_date.day}.{end_date.month}.{end_date.year}"
    )


def _invoice_display_number(invoice: Invoice) -> int:
    """Einheitliche Rechnungsnummer fuer Anzeige/Zweck (seq_number bevorzugt)."""
    seq = getattr(invoice, "seq_number", None)
    try:
        if seq is not None:
            return int(seq)
    except Exception:
        pass
    return int(getattr(invoice, "id", 0) or 0)


def _next_free_invoice_seq(used_numbers: set[int]) -> int:
    candidate = 1
    while candidate in used_numbers:
        candidate += 1
    return candidate


def _draft_virtual_display_numbers() -> dict[int, int]:
    """
    Vergibt fuer Entwuerfe ohne seq_number stabile virtuelle Rechnungsnummern,
    indem Luecken in den vorhandenen Rechnungsnummern zuerst aufgefuellt werden.
    """
    used_numbers = {
        int(row[0])
        for row in db.session.query(Invoice.seq_number)
        .filter(Invoice.seq_number.isnot(None))
        .all()
        if row[0] is not None
    }

    drafts = (
        Invoice.query
        .with_entities(Invoice.id)
        .filter(
            Invoice.stage == "draft",
            Invoice.seq_number.is_(None),
        )
        # Neueste Entwürfe zuerst nummerieren, damit der aktuell erzeugte
        # Entwurf die kleinste freie Rechnungsnummer erhält.
        .order_by(Invoice.created_at.desc(), Invoice.id.desc())
        .all()
    )

    virtual_map: dict[int, int] = {}
    for (draft_id,) in drafts:
        next_number = _next_free_invoice_seq(used_numbers)
        virtual_map[int(draft_id)] = next_number
        used_numbers.add(next_number)

    return virtual_map


def _invoice_display_number_for_detail(invoice: Invoice) -> int:
    seq = getattr(invoice, "seq_number", None)
    if seq is not None:
        try:
            return int(seq)
        except Exception:
            pass

    if getattr(invoice, "stage", None) == "draft":
        virtual = _draft_virtual_display_numbers().get(int(invoice.id))
        if virtual is not None:
            return int(virtual)

    return int(getattr(invoice, "id", 0) or 0)


def _invoice_route_number(invoice: Invoice) -> int:
    """URL-Nummer immer aus der sichtbaren Rechnungsnummer ableiten."""
    return _invoice_display_number_for_detail(invoice)


def _get_invoice_by_display_number(invoice_number: int) -> Invoice | None:
    invoice = (
        Invoice.query
        .options(
            joinedload(Invoice.person),
            selectinload(Invoice.items).joinedload(InvoiceItem.load_entry),
        )
        .filter(Invoice.seq_number == invoice_number)
        .first()
    )
    if invoice is not None:
        return invoice

    draft_map = _draft_virtual_display_numbers()
    draft_id = None
    for _draft_id, _display_number in draft_map.items():
        if int(_display_number) == int(invoice_number):
            draft_id = int(_draft_id)
            break

    if draft_id is None:
        # Defensiver Fallback: Wenn URL genau auf die naechste freie Rechnungsnummer
        # zeigt und ein Entwurf offen ist, diesen Entwurf oeffnen.
        max_seq = db.session.query(db.func.max(Invoice.seq_number)).scalar() or 0
        if int(invoice_number) == int(max_seq) + 1:
            newest_draft = (
                Invoice.query
                .with_entities(Invoice.id)
                .filter(
                    Invoice.stage == "draft",
                    Invoice.seq_number.is_(None),
                )
                .order_by(Invoice.created_at.desc(), Invoice.id.desc())
                .first()
            )
            if newest_draft is not None:
                draft_id = int(newest_draft[0])

    if draft_id is None:
        # Rueckwaertskompatibel: alte Links mit DB-id weiterhin aufloesen.
        return (
            Invoice.query
            .options(
                joinedload(Invoice.person),
                selectinload(Invoice.items).joinedload(InvoiceItem.load_entry),
            )
            .get(invoice_number)
        )

    invoice = (
        Invoice.query
        .options(
            joinedload(Invoice.person),
            selectinload(Invoice.items).joinedload(InvoiceItem.load_entry),
        )
        .get(draft_id)
    )
    if invoice is not None:
        return invoice

    # Rueckwaertskompatibel: alte Links mit DB-id weiterhin aufloesen.
    return (
        Invoice.query
        .options(
            joinedload(Invoice.person),
            selectinload(Invoice.items).joinedload(InvoiceItem.load_entry),
        )
        .get(invoice_number)
    )


def _build_invoice_payment_purpose(
    invoice: Invoice,
    doc_label: str = "Rechnung",
    invoice_number: int | None = None,
) -> str:
    """Baut den Verwendungszweck fuer Anzeige und EPC-Remittance konsistent."""
    return build_invoice_payment_purpose(invoice, doc_label=doc_label, invoice_number=invoice_number)


@bp.context_processor
def inject_invoice_purpose():
    """Macht _build_invoice_payment_purpose im Template als generate_invoice_purpose verfuegbar."""
    return {
        "generate_invoice_purpose": _build_invoice_payment_purpose,
        "invoice_display_number": _invoice_display_number_for_detail,
    }


# ---------------------------------------------------------
# DEV / PROD Erkennung (zentral)
# ---------------------------------------------------------
def is_dev_mode() -> bool:
    """
    True = Entwicklungsmodus
    False = Produktivmodus

    Priorität:
    1) MANIFEST_ENV=dev\\development
    2) Flask Debug-Flag (Fallback)
    """
    env = os.environ.get("MANIFEST_ENV", "").lower()
    if env in {"dev", "development"}:
        return True
    try:
        return bool(current_app.debug)
    except Exception:
        return False


def _parse_invoice_list_filters(args) -> dict:
    allowed_invoice_source = {"all", "loads", "manual"}
    allowed_status = {
        "",
        INVOICE_PAYMENT_STATE_OPEN,
        INVOICE_PAYMENT_STATE_SEPA_PENDING,
        INVOICE_PAYMENT_STATE_SEPA_EXPORTED,
        INVOICE_PAYMENT_STATE_PAID,
        INVOICE_PAYMENT_STATE_SEPA_RETURNED,
        "sepa_pending",
        "sepa_exported",
        "non_sepa",
    }
    allowed_payment = {"", "cash", "card", "transfer", "wero", "sepa", "voucher"}
    allowed_email = {"", "not_sent", "error", "pending", "sent_unconfirmed", "sent_confirmed"}
    allowed_content_status = {
        "",
        "member",
        "guest",
        "partner",
        "manual",
        "tandem_guest",
        "student",
        "teacher",
        "aff_student",
        "aff_teacher",
        "video",
        "tandemmaster",
    }
    allowed_sort = {
        "date_desc", "date_asc",
        "amount_desc", "amount_asc",
        "status_asc", "status_desc",
        "pay_asc", "pay_desc",
        "inv_desc", "inv_asc",
        "person_asc", "person_desc",
        "email_desc", "email_asc",
        "content_status_asc", "content_status_desc",
        "sepa_pending_first", "sepa_exported_first", "sepa_last",
    }

    person_id = None
    person_id_raw = (args.get("person_id") or "").strip()
    if person_id_raw:
        try:
            person_id = int(person_id_raw)
        except Exception:
            person_id = None

    status = (args.get("status") or "").strip().lower()
    if status not in allowed_status:
        status = ""

    payment = (args.get("payment") or "").strip().lower()
    if payment not in allowed_payment:
        payment = ""

    email = (args.get("email") or "").strip().lower()
    if email not in allowed_email:
        email = ""

    content_status = (args.get("content_status") or "").strip().lower()
    if content_status not in allowed_content_status:
        content_status = ""

    sort = (args.get("sort") or "date_desc").strip()
    if sort not in allowed_sort:
        sort = "date_desc"

    invoice_source = (args.get("invoice_source") or "all").strip().lower()
    if invoice_source not in allowed_invoice_source:
        invoice_source = "all"

    return {
        "invoice_source": invoice_source,
        "person_id": person_id,
        "person": (args.get("person") or "").strip(),
        "text": (args.get("text") or "").strip(),
        "status": status,
        "payment": payment,
        "email": email,
        "content_status": content_status,
        "sort": sort,
    }


def _parse_invoice_delta_scope(args) -> str:
    scope = (args.get("delta_scope") or "").strip().lower()
    return "all" if scope == "all" else "visible"


def _resolve_named_date_range(
    from_str: str = "",
    to_str: str = "",
    *,
    fallback_label: str = "all",
):
    start_date = None
    end_date = None
    normalized_label = fallback_label

    try:
        if from_str:
            start_date = datetime.strptime(from_str, "%Y-%m-%d").date()
        if to_str:
            to_date = datetime.strptime(to_str, "%Y-%m-%d").date()
            end_date = to_date + timedelta(days=1)

        if start_date and end_date and end_date <= start_date:
            start_date = None
            end_date = None
            normalized_label = fallback_label
        elif start_date or end_date:
            normalized_label = "range"
    except Exception:
        start_date = None
        end_date = None
        normalized_label = fallback_label

    start_dt = datetime.combine(start_date, dt_time.min) if start_date else None
    end_dt = datetime.combine(end_date, dt_time.min) if end_date else None
    return normalized_label, start_dt, end_dt


def _resolve_load_date_range(period: str, from_str: str = "", to_str: str = ""):
    today = date.today()
    start_date = None
    end_date = None
    normalized_period = period

    if period == "today":
        start_date = today
        end_date = today + timedelta(days=1)

    elif period == "week":
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=7)

    elif period == "month":
        start_date = today.replace(day=1)
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1)

    elif period == "year":
        start_date = today.replace(month=1, day=1)
        end_date = start_date.replace(year=start_date.year + 1)

    elif period == "range":
        normalized_period, start_dt, end_dt = _resolve_named_date_range(
            from_str,
            to_str,
            fallback_label="all",
        )
        return normalized_period, start_dt, end_dt

    start_dt = datetime.combine(start_date, dt_time.min) if start_date else None
    end_dt = datetime.combine(end_date, dt_time.min) if end_date else None
    return normalized_period, start_dt, end_dt


PREPAID_ELIGIBLE_STATUSES = {normalize_status_code(s) for s in TANDEM_GUEST_STATUSES}


def _entries_allow_prepaid_voucher(entries: list[LoadEntry] | None) -> bool:
    for e in list(entries or []):
        status = normalize_status_code(getattr(e, "status_code", "") or "")
        if status in PREPAID_ELIGIBLE_STATUSES:
            return True
    return False


def _invoice_allows_prepaid_voucher(invoice: Invoice | None) -> bool:
    if not invoice:
        return False
    for item in list(getattr(invoice, "items", []) or []):
        if (getattr(item, "item_source", "") or "").strip().lower() == "manual":
            return True
    entries = [getattr(item, "load_entry", None) for item in list(getattr(invoice, "items", []) or [])]
    return _entries_allow_prepaid_voucher([e for e in entries if e is not None])


def _invoice_prepaid_amount(invoice: Invoice | None) -> Decimal:
    if not invoice:
        return Decimal("0.00")
    try:
        value = Decimal(str(getattr(invoice, "prepaid_voucher_amount", 0) or 0))
    except Exception:
        value = Decimal("0.00")
    if value < Decimal("0.00"):
        return Decimal("0.00")
    if value < Decimal("0.00"):
        value = Decimal("0.00")
    return value.quantize(Decimal("0.01"))


def _invoice_onsite_amount(invoice: Invoice | None) -> Decimal:
    total = Decimal(str(getattr(invoice, "total_amount", 0) or 0)) if invoice else Decimal("0.00")
    prepaid = _invoice_prepaid_amount(invoice)
    return (total - prepaid).quantize(Decimal("0.01"))


def _invoice_open_amount_for_kpi(invoice: Invoice | None) -> Decimal:
    """
    Offener Anteil fuer KPI-Summen.

    Regeln:
    - Bereits bezahlte Rechnung -> 0
    - Offene Rechnung mit positivem Betrag -> Rest (total - prepaid)
    - Offene Rechnung mit negativem Betrag (z.B. Gutschrift) -> voller negativer Betrag
    """
    if not invoice or _invoice_payment_state(invoice) == INVOICE_PAYMENT_STATE_PAID:
        return Decimal("0.00")

    total = Decimal(str(getattr(invoice, "total_amount", 0) or 0))
    if total <= Decimal("0.00"):
        return total.quantize(Decimal("0.01"))

    prepaid = _invoice_prepaid_amount(invoice)
    open_amount = total - prepaid
    return open_amount.quantize(Decimal("0.01"))


def _invoice_paid_amount_for_kpi(invoice: Invoice | None) -> Decimal:
    """
    Bereits bezahlter Anteil fuer KPI-Summen.

    Regeln:
    - Positive Rechnung: prepaid immer bezahlt; Rest nur wenn is_paid=True
    - Negative Rechnung: nur bei is_paid=True als bezahlt gezaehlt
    """
    if not invoice:
        return Decimal("0.00")

    total = Decimal(str(getattr(invoice, "total_amount", 0) or 0))
    if total <= Decimal("0.00"):
        return total.quantize(Decimal("0.01")) if _invoice_payment_state(invoice) == INVOICE_PAYMENT_STATE_PAID else Decimal("0.00")

    prepaid = _invoice_prepaid_amount(invoice)
    if _invoice_payment_state(invoice) == INVOICE_PAYMENT_STATE_PAID:
        return total.quantize(Decimal("0.01"))
    return prepaid.quantize(Decimal("0.01"))


def _invoice_split_payment_label(invoice: Invoice | None) -> str:
    if not invoice:
        return ""
    onsite_label = _invoice_payment_label(getattr(invoice, "payment_method", None))
    prepaid = _invoice_prepaid_amount(invoice)
    if prepaid > Decimal("0.00"):
        if onsite_label:
            return f"{onsite_label} + Vorkasse / Gutschein"
        return "Vorkasse / Gutschein"
    return onsite_label


def _parse_form_bool(value, *, default: bool = False) -> bool:
    if value is None:
        return bool(default)
    return str(value).strip().lower() in {"1", "true", "yes", "on", "ja"}


def _person_tandem_ku_default(person: Person | None) -> bool:
    if not person:
        return False
    return _parse_form_bool(getattr(person, "is_tandem_kleinunternehmer", None), default=False)


def _person_video_ku_default(person: Person | None) -> bool:
    if not person:
        return False
    return _parse_form_bool(getattr(person, "is_video_kleinunternehmer", None), default=False)


def _is_tandemmaster_entry(entry: LoadEntry | None) -> bool:
    if not entry:
        return False
    return normalize_status_code(getattr(entry, "status_code", "") or "") in TM_STATUSES


def _is_video_entry(entry: LoadEntry | None) -> bool:
    if not entry:
        return False
    return normalize_status_code(getattr(entry, "status_code", "") or "") in VIDEO_STATUSES


def _is_ku_eligible_entry(entry: LoadEntry | None) -> bool:
    return _is_tandemmaster_entry(entry) or _is_video_entry(entry)


def _is_jump_item(item: InvoiceItem | None) -> bool:
    if not item:
        return False
    desc = (getattr(item, "description", "") or "").strip()
    if not desc.startswith("Sprung"):
        return False
    return bool(getattr(item, "load_entry", None))


def _is_tandemmaster_jump_item(item: InvoiceItem | None) -> bool:
    if not _is_jump_item(item):
        return False
    return _is_tandemmaster_entry(getattr(item, "load_entry", None))


def _is_video_jump_item(item: InvoiceItem | None) -> bool:
    if not _is_jump_item(item):
        return False
    return _is_video_entry(getattr(item, "load_entry", None))


def _is_ku_eligible_jump_item(item: InvoiceItem | None) -> bool:
    return _is_tandemmaster_jump_item(item) or _is_video_jump_item(item)


def _invoice_has_tandem_jump_positions(invoice: Invoice | None) -> bool:
    if not invoice:
        return False
    for item in list(getattr(invoice, "items", []) or []):
        if _is_tandemmaster_jump_item(item):
            return True
    return False


def _invoice_has_video_jump_positions(invoice: Invoice | None) -> bool:
    if not invoice:
        return False
    for item in list(getattr(invoice, "items", []) or []):
        if _is_video_jump_item(item):
            return True
    return False


def _invoice_has_ku_jump_positions(invoice: Invoice | None) -> bool:
    return _invoice_has_tandem_jump_positions(invoice) or _invoice_has_video_jump_positions(invoice)


def _invoice_totals(invoice: Invoice | None) -> tuple[Decimal, Decimal, Decimal]:
    if not invoice:
        return Decimal("0.00"), Decimal("0.00"), Decimal("0.00")

    net_total = Decimal("0.00")
    vat_total = Decimal("0.00")
    gross_total = Decimal("0.00")
    for item in list(getattr(invoice, "items", []) or []):
        net_total += Decimal(str(getattr(item, "net_amount", 0) or 0))
        vat_total += Decimal(str(getattr(item, "vat_amount", 0) or 0))
        gross_total += Decimal(str(getattr(item, "amount", 0) or 0))

    q = Decimal("0.01")
    return (
        net_total.quantize(q),
        vat_total.quantize(q),
        gross_total.quantize(q),
    )


def _invoice_totals_net_vat(invoice: Invoice | None) -> tuple[Decimal, Decimal]:
    net_total, vat_total, _ = _invoice_totals(invoice)
    return net_total, vat_total


def _paid_method_breakdown(invoices: list[Invoice] | None) -> dict[str, Decimal]:
    sums = {
        "cash": Decimal("0.00"),
        "card": Decimal("0.00"),
        "transfer": Decimal("0.00"),
        "wero": Decimal("0.00"),
        "sepa": Decimal("0.00"),
        "voucher": Decimal("0.00"),
    }
    for inv in list(invoices or []):
        prepaid = _invoice_prepaid_amount(inv)
        if prepaid > Decimal("0.00"):
            sums["voucher"] += prepaid

        if _invoice_payment_state(inv) != INVOICE_PAYMENT_STATE_PAID:
            continue

        pm = (getattr(inv, "payment_method", "") or "").strip().lower()
        onsite = _invoice_onsite_amount(inv)
        if pm in sums:
            sums[pm] += onsite
    return sums


def _parse_prepaid_amount(raw_value: str | None, *, total_amount: Decimal, allow_prepaid: bool) -> tuple[Decimal, str | None]:
    raw = (raw_value or "").strip().replace("€", "").replace(" ", "")
    if not raw:
        return Decimal("0.00"), None

    normalized = raw.replace(".", "").replace(",", ".") if "," in raw else raw
    try:
        value = Decimal(normalized)
    except Exception:
        return Decimal("0.00"), "Ungültiger Betrag für Vorkasse / Gutschein."

    value = value.quantize(Decimal("0.01"))
    if value < Decimal("0.00"):
        return Decimal("0.00"), "Vorkasse / Gutschein darf nicht negativ sein."
    if value == Decimal("0.00"):
        return Decimal("0.00"), None

    if not allow_prepaid:
        return Decimal("0.00"), "Vorkasse / Gutschein ist aktuell nur für Tandemgäste (inkl. Mitflieger/Video) erlaubt."

    return value, None


def _parse_decimal_de(raw_value: str | None, *, allow_negative: bool = False) -> Decimal:
    """
    Parst deutsche Zahlformate, z.B.:
    - 1.234,56
    - 1234,56
    - 1234.56
    """
    raw = (raw_value or "").strip().replace("€", "").replace(" ", "")
    if not raw:
        return Decimal("0.00")

    if "," in raw:
        normalized = raw.replace(".", "").replace(",", ".")
    else:
        normalized = raw

    value = Decimal(normalized)
    if not allow_negative and value < Decimal("0.00"):
        raise ValueError("negative-not-allowed")
    return value


def _format_decimal_de_for_input(value: Decimal | int | float | None, fallback: str = "0,00") -> str:
    try:
        return "{:,.2f}".format(Decimal(str(value or 0))).replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return fallback


def _is_manual_draft_invoice(invoice: Invoice | None) -> bool:
    if not invoice or getattr(invoice, "stage", None) != "draft":
        return False
    for item in list(getattr(invoice, "items", []) or []):
        if (getattr(item, "item_source", "") or "").strip().lower() == "manual":
            return True
    return False


def _is_manual_invoice(invoice: Invoice | None) -> bool:
    if not invoice:
        return False

    has_manual_items = False
    has_load_items = False
    for item in list(getattr(invoice, "items", []) or []):
        if (getattr(item, "item_source", "") or "").strip().lower() == "manual":
            has_manual_items = True
        if getattr(item, "load_entry", None):
            has_load_items = True

    return has_manual_items and not has_load_items


def _invoice_email_status(invoice: Invoice) -> str:
    err_newer = (
        getattr(invoice, "email_last_error", None)
        and getattr(invoice, "email_last_attempt_at", None)
        and (
            not getattr(invoice, "email_sent_at", None)
            or invoice.email_last_attempt_at >= invoice.email_sent_at
        )
    )
    pending = (
        getattr(invoice, "email_last_attempt_at", None)
        and not getattr(invoice, "email_sent_ok", False)
        and not getattr(invoice, "email_last_error", None)
    )

    if err_newer:
        return "error"
    if pending:
        return "pending"
    if getattr(invoice, "email_sent_ok", False) and getattr(invoice, "email_delivery_confirmed_at", None):
        return "sent_confirmed"
    if getattr(invoice, "email_sent_ok", False):
        return "sent_unconfirmed"
    return "not_sent"


def _invoice_content_status_codes(invoice: Invoice) -> set[str]:
    codes: set[str] = set()

    for item in list(getattr(invoice, "items", []) or []):
        if (getattr(item, "item_source", "") or "").strip().lower() == "manual":
            codes.add("manual")
            break

    for item in list(getattr(invoice, "items", []) or []):
        entry = getattr(item, "load_entry", None)
        if not entry:
            continue
        status = normalize_status_code(getattr(entry, "status_code", "") or "")
        if status == "Verein" or status == "Auffüller Verein":
            codes.add("member")
        elif status == "Partner-Verein" or status == "Auffüller Partner-Verein":
            codes.add("partner")
        elif status == "Gast" or status == "Auffüller Gast":
            codes.add("guest")
        elif status in {"G-TD", "G-TD-Video"}:
            codes.add("tandem_guest")
        elif status in {"Schüler", "Schüler Ek 1", "Schüler Ek 2", "Schüler GK 6"}:
            codes.add("student")
        elif status in {"Schueler-Aff-1", "Schueler-Aff-2"}:
            codes.add("aff_student")
            codes.add("student")
        elif status == "Lehrer":
            codes.add("teacher")
        elif status == "Aff-Lehrer":
            codes.add("aff_teacher")
            codes.add("teacher")
        elif status == "Video":
            codes.add("video")
        elif status in {"TD", "TD-Vereins-Schirm"}:
            codes.add("tandemmaster")

    return codes


def _invoice_content_status_label(code: str) -> str:
    labels = {
        "member": "Verein",
        "guest": "Gast",
        "partner": "Partner-Verein",
        "manual": "Manuell",
        "tandem_guest": "Tandemgast",
        "student": "Schüler",
        "teacher": "Lehrer",
        "aff_student": "AFF-Schüler",
        "aff_teacher": "AFF-Lehrer",
        "video": "Video",
        "tandemmaster": "Tandemmaster",
    }
    return labels.get(code, code)


def _invoice_content_status_sort_key(invoice: Invoice) -> str:
    labels = [_invoice_content_status_label(code) for code in sorted(_invoice_content_status_codes(invoice))]
    return " | ".join(labels).casefold()


def _invoice_content_status_codes_csv(invoice: Invoice) -> str:
    return ",".join(sorted(_invoice_content_status_codes(invoice)))


def _billable_person_content_status_codes(person: Person | None, entries: list[LoadEntry] | None) -> set[str]:
    codes: set[str] = set()

    for e in list(entries or []):
        status = normalize_status_code(getattr(e, "status_code", "") or "")
        if status == "Verein" or status == "Auffüller Verein":
            codes.add("member")
        elif status == "Partner-Verein" or status == "Auffüller Partner-Verein":
            codes.add("partner")
        elif status == "Gast" or status == "Auffüller Gast":
            codes.add("guest")
        elif status in {"G-TD", "G-TD-Video"}:
            codes.add("tandem_guest")
        elif status in {"Schüler", "Schüler Ek 1", "Schüler Ek 2", "Schüler GK 6"}:
            codes.add("student")
        elif status in {"Schueler-Aff-1", "Schueler-Aff-2"}:
            codes.add("aff_student")
            codes.add("student")
        elif status == "Lehrer":
            codes.add("teacher")
        elif status == "Aff-Lehrer":
            codes.add("aff_teacher")
            codes.add("teacher")
        elif status == "Video":
            codes.add("video")
        elif status in {"TD", "TD-Vereins-Schirm"}:
            codes.add("tandemmaster")

    return codes

def _invoice_matches_filters(invoice: Invoice, filters: dict | None) -> bool:
    if not filters:
        return True

    invoice_source = (filters.get("invoice_source") or "all").strip().lower()
    person_id = filters.get("person_id")
    person_query = (filters.get("person") or "").strip().casefold()
    text_query = (filters.get("text") or "").strip().casefold()
    status = filters.get("status") or ""
    payment = filters.get("payment") or ""
    email = filters.get("email") or ""
    content_status = filters.get("content_status") or ""

    person = getattr(invoice, "person", None)
    person_name = (person.full_name if person else "").casefold()

    has_manual_items = False
    has_load_items = False
    for item in list(getattr(invoice, "items", []) or []):
        if (getattr(item, "item_source", "") or "").strip().lower() == "manual":
            has_manual_items = True
        if getattr(item, "load_entry_id", None) or getattr(item, "load_entry", None):
            has_load_items = True

    if invoice_source == "manual" and not (has_manual_items and not has_load_items):
        return False
    if invoice_source == "loads" and not has_load_items:
        return False

    if person_id is not None:
        if getattr(invoice, "person_id", None) != person_id:
            return False
    elif person_query and person_query not in person_name:
        return False

    if status == "non_sepa":
        if _invoice_payment_state(invoice) in {
            INVOICE_PAYMENT_STATE_SEPA_PENDING,
            INVOICE_PAYMENT_STATE_SEPA_EXPORTED,
        }:
            return False
    elif status and _invoice_payment_state(invoice) != status:
        return False

    if payment:
        if payment == "voucher":
            if _invoice_prepaid_amount(invoice) <= Decimal("0.00"):
                return False
        elif (invoice.payment_method or "") != payment:
            return False

    if email and _invoice_email_status(invoice) != email:
        return False

    if content_status and content_status not in _invoice_content_status_codes(invoice):
        return False

    if text_query:
        invoice_label = ""
        if getattr(invoice, "created_at", None):
            invoice_label = f"{invoice.created_at.strftime('%Y')}-Sprünge #{_invoice_display_number(invoice)}"
        search_text = " ".join([
            invoice_label,
            person.full_name if person else "",
            _invoice_payment_label(getattr(invoice, "payment_method", None)),
        ]).casefold()
        if text_query not in search_text:
            return False

    return True


def _billable_row_matches_filters(row: dict, filters: dict | None) -> bool:
    if not filters:
        return True

    status = filters.get("status") or ""
    payment = filters.get("payment") or ""
    email = filters.get("email") or ""
    if status == "paid" or payment or email:
        return False

    person = row.get("person")
    if not person:
        return False

    person_id = filters.get("person_id")
    if person_id is not None and getattr(person, "id", None) != person_id:
        return False

    person_query = (filters.get("person") or "").strip().casefold()
    person_name = (person.full_name or "").casefold()
    if person_id is None and person_query and person_query not in person_name:
        return False

    text_query = (filters.get("text") or "").strip().casefold()
    if text_query:
        amount_text = f"{Decimal(str(row.get('amount') or '0.00')):.2f}".casefold()
        if text_query not in person_name and text_query not in amount_text:
            return False

    return True


def _sort_invoices_for_list(invoices: list[Invoice], sort_mode: str) -> list[Invoice]:
    sort_mode = (sort_mode or "date_desc").strip()

    def _invoice_sort_state(inv: Invoice) -> str:
        return (getattr(inv, "payment_state", "") or "").strip().lower()

    def _sepa_sort_priority(inv: Invoice) -> tuple[int, int, datetime]:
        raw_state = _invoice_sort_state(inv)
        if raw_state == INVOICE_PAYMENT_STATE_SEPA_PENDING:
            return (0, 0, getattr(inv, "created_at", datetime.min))
        if raw_state == INVOICE_PAYMENT_STATE_SEPA_EXPORTED:
            return (1, 0, getattr(inv, "created_at", datetime.min))
        return (2, 0, getattr(inv, "created_at", datetime.min))

    def _sepa_sort_priority_for_exported_first(inv: Invoice) -> tuple[int, int, datetime]:
        raw_state = _invoice_sort_state(inv)
        if raw_state == INVOICE_PAYMENT_STATE_SEPA_EXPORTED:
            return (0, 0, getattr(inv, "created_at", datetime.min))
        if raw_state == INVOICE_PAYMENT_STATE_SEPA_PENDING:
            return (1, 0, getattr(inv, "created_at", datetime.min))
        return (2, 0, getattr(inv, "created_at", datetime.min))

    if sort_mode == "date_asc":
        return sorted(invoices, key=lambda inv: getattr(inv, "created_at", datetime.min))
    if sort_mode == "amount_desc":
        return sorted(invoices, key=lambda inv: Decimal(str(inv.total_amount or "0.00")), reverse=True)
    if sort_mode == "amount_asc":
        return sorted(invoices, key=lambda inv: Decimal(str(inv.total_amount or "0.00")))
    if sort_mode == "status_asc":
        return sorted(invoices, key=lambda inv: _invoice_payment_state(inv))
    if sort_mode == "status_desc":
        return sorted(invoices, key=lambda inv: _invoice_payment_state(inv), reverse=True)
    if sort_mode == "sepa_pending_first":
        return sorted(
            invoices,
            key=lambda inv: _sepa_sort_priority(inv),
            reverse=False,
        )
    if sort_mode == "sepa_exported_first":
        return sorted(
            invoices,
            key=lambda inv: _sepa_sort_priority_for_exported_first(inv),
            reverse=False,
        )
    if sort_mode == "sepa_last":
        return sorted(
            invoices,
            key=lambda inv: (
                0 if _invoice_sort_state(inv) == INVOICE_PAYMENT_STATE_OPEN else 1,
                0 if _invoice_sort_state(inv) == INVOICE_PAYMENT_STATE_SEPA_PENDING else 1,
                getattr(inv, "created_at", datetime.min),
            ),
            reverse=False,
        )
    if sort_mode == "pay_asc":
        return sorted(invoices, key=lambda inv: (inv.payment_method or ""))
    if sort_mode == "pay_desc":
        return sorted(invoices, key=lambda inv: (inv.payment_method or ""), reverse=True)
    if sort_mode == "inv_asc":
        return sorted(invoices, key=lambda inv: getattr(inv, "id", 0))
    if sort_mode == "inv_desc":
        return sorted(invoices, key=lambda inv: getattr(inv, "id", 0), reverse=True)
    if sort_mode == "person_asc":
        return sorted(
            invoices,
            key=lambda inv: (
                ((inv.person.last_name or "") + " " + (inv.person.first_name or "")).casefold()
                if inv.person else ""
            ),
        )
    if sort_mode == "person_desc":
        return sorted(
            invoices,
            key=lambda inv: (
                ((inv.person.last_name or "") + " " + (inv.person.first_name or "")).casefold()
                if inv.person else ""
            ),
            reverse=True,
        )
    if sort_mode == "email_asc":
        return sorted(
            invoices,
            key=lambda inv: inv.email_sent_at.timestamp() if getattr(inv, "email_sent_ok", False) and getattr(inv, "email_sent_at", None) else 0,
        )
    if sort_mode == "email_desc":
        return sorted(
            invoices,
            key=lambda inv: inv.email_sent_at.timestamp() if getattr(inv, "email_sent_ok", False) and getattr(inv, "email_sent_at", None) else 0,
            reverse=True,
        )
    if sort_mode == "content_status_asc":
        return sorted(invoices, key=lambda inv: _invoice_content_status_sort_key(inv))
    if sort_mode == "content_status_desc":
        return sorted(invoices, key=lambda inv: _invoice_content_status_sort_key(inv), reverse=True)

    return sorted(invoices, key=lambda inv: getattr(inv, "created_at", datetime.min), reverse=True)


def _build_invoice_filter_labels(filters: dict | None) -> list[str]:
    if not filters:
        return []

    labels = []
    invoice_source = (filters.get("invoice_source") or "all").strip().lower()
    source_map = {
        "all": "Alle",
        "loads": "Loads",
        "manual": "Manuelle Rechnungen",
    }
    if invoice_source != "all":
        labels.append(f"Quelle: {source_map.get(invoice_source, invoice_source)}")
    if filters.get("person_id") is not None or filters.get("person"):
        labels.append(f"Person: {filters.get('person') or filters.get('person_id')}")
    if filters.get("text"):
        labels.append(f"Freitext: {filters['text']}")
    if filters.get("status"):
        labels.append(f"Status: {PAYMENT_STATE_LABELS.get(filters['status'], filters['status'])}")
    if filters.get("payment"):
        labels.append(f"Bezahlart: {_invoice_payment_label(filters['payment'])}")
    email_map = {
        "not_sent": "Nicht versendet",
        "error": "Fehler",
        "pending": "Versuch ohne Ergebnis",
        "sent_unconfirmed": "SMTP ok, unbestätigt",
        "sent_confirmed": "Eingang bestätigt",
    }
    if filters.get("email"):
        labels.append(f"E-Mail: {email_map.get(filters['email'], filters['email'])}")
    if filters.get("content_status"):
        labels.append(f"Inhaltsstatus: {_invoice_content_status_label(filters['content_status'])}")
    if (filters.get("sort") or "date_desc") != "date_desc":
        labels.append(f"Sortierung: {filters['sort']}")
    return labels


@bp.context_processor
def inject_invoice_content_status_helpers():
    return {
        "invoice_content_status_sort_key": _invoice_content_status_sort_key,
        "invoice_content_status_codes_csv": _invoice_content_status_codes_csv,
        "invoice_prepaid_amount": _invoice_prepaid_amount,
        "invoice_onsite_amount": _invoice_onsite_amount,
        "invoice_split_payment_label": _invoice_split_payment_label,
        "invoice_payment_state_code": _invoice_payment_state,
        "invoice_payment_state_label": _invoice_payment_state_label,
    }


# ---------------------------------------------------------
# Übersicht (Startseite Abrechnung)
# ---------------------------------------------------------
@bp.route("/overview")
def overview():
    return render_template("billing/overview.html")


@bp.route("/manual/new", methods=["GET", "POST"])
def manual_invoice_new():
    def _parse_int_loose(raw: str | None) -> int | None:
        text = (raw or "").strip()
        if not text:
            return None
        m = re.search(r"\d+", text)
        if not m:
            return None
        try:
            return int(m.group(0))
        except Exception:
            return None

    def _manual_form_redirect(*, person_id: int | None = None, draft_invoice_id: int | None = None):
        params = {}
        if person_id is not None:
            params["person_id"] = person_id
        if draft_invoice_id is not None:
            params["draft_invoice_id"] = draft_invoice_id
        return redirect(url_for("billing.manual_invoice_new", **params))

    def _blank_manual_line() -> dict:
        return {
            "description": "",
            "quantity": "1,00",
            "unit": "",
            "unit_price_gross": "0,00",
            "vat_rate": "19,00",
        }

    def _build_manual_form_data_from_invoice(invoice: Invoice | None) -> dict:
        form_data = {
            "draft_invoice_id": invoice.id if invoice else None,
            "service_date": (invoice.service_date if invoice and invoice.service_date else date.today()).strftime("%Y-%m-%d"),
            "prepaid_voucher_amount": _format_decimal_de_for_input(invoice.prepaid_voucher_amount, "") if invoice and Decimal(str(invoice.prepaid_voucher_amount or 0)) > Decimal("0.00") else "",
            "manual_title": (invoice.manual_title if invoice else "") or "Manuelle Positionen",
            "billing_address_name": (invoice.billing_address_name if invoice else "") or "",
            "billing_address_email": (invoice.billing_address_email if invoice else "") or "",
            "billing_address_street": (invoice.billing_address_street if invoice else "") or "",
            "billing_address_zip": (invoice.billing_address_zip if invoice else "") or "",
            "billing_address_city": (invoice.billing_address_city if invoice else "") or "",
            "lines": [],
        }
        if invoice:
            for item in list(invoice.items or []):
                if (getattr(item, "item_source", "") or "").strip().lower() != "manual":
                    continue
                form_data["lines"].append(
                    {
                        "description": item.description or "",
                        "quantity": _format_decimal_de_for_input(item.quantity, "1,00"),
                        "unit": item.manual_unit or "",
                        "unit_price_gross": _format_decimal_de_for_input(item.unit_price_gross, "0,00"),
                        "vat_rate": _format_decimal_de_for_input(item.vat_rate, "19,00"),
                    }
                )
        if not form_data["lines"]:
            form_data["lines"].append(_blank_manual_line())
        return form_data

    def _build_manual_form_data_from_request(form, *, draft_invoice_id: int | None = None) -> dict:
        descriptions = form.getlist("line_description[]")
        quantities = form.getlist("line_quantity[]")
        units = form.getlist("line_unit[]")
        unit_prices = form.getlist("line_unit_price_gross[]")
        vat_rates = form.getlist("line_vat_rate[]")
        line_count = max(
            len(descriptions),
            len(quantities),
            len(units),
            len(unit_prices),
            len(vat_rates),
            1,
        )

        lines = []
        for i in range(line_count):
            lines.append(
                {
                    "description": descriptions[i] if i < len(descriptions) else "",
                    "quantity": quantities[i] if i < len(quantities) else "",
                    "unit": units[i] if i < len(units) else "",
                    "unit_price_gross": unit_prices[i] if i < len(unit_prices) else "",
                    "vat_rate": vat_rates[i] if i < len(vat_rates) else "",
                }
            )

        if not lines:
            lines.append(_blank_manual_line())

        return {
            "draft_invoice_id": draft_invoice_id,
            "service_date": (form.get("service_date") or "").strip(),
            "prepaid_voucher_amount": (form.get("prepaid_voucher_amount") or "").strip(),
            "manual_title": (form.get("manual_title") or "").strip() or "Manuelle Positionen",
            "billing_address_name": (form.get("billing_address_name") or "").strip(),
            "billing_address_email": (form.get("billing_address_email") or "").strip(),
            "billing_address_street": (form.get("billing_address_street") or "").strip(),
            "billing_address_zip": (form.get("billing_address_zip") or "").strip(),
            "billing_address_city": (form.get("billing_address_city") or "").strip(),
            "lines": lines,
        }

    def _render_manual_form(*, form_data: dict, selected_person: Person | None, draft_invoice: Invoice | None):
        return render_template(
            "billing/manual_invoice_new.html",
            selected_person=selected_person,
            today=date.today(),
            form_data=form_data,
            draft_invoice=draft_invoice,
        )

    def _is_manual_recipient_person(person: Person | None) -> bool:
        if not person:
            return False
        return ((getattr(person, "phone", "") or "").strip() == "__MANUAL_INVOICE__")

    def _create_manual_recipient_person(
        *,
        name: str,
        email: str | None,
        street: str | None,
        zip_code: str | None,
        city: str | None,
        existing_person: Person | None = None,
    ) -> Person:
        raw_name = (name or "").strip()
        collapsed_name = re.sub(r"\s+", " ", raw_name)
        if not collapsed_name:
            raise ValueError("missing-manual-recipient-name")

        if len(collapsed_name) <= 50:
            first_name = collapsed_name
            last_name = ""
        else:
            split_at = collapsed_name.rfind(" ", 0, 50)
            if split_at <= 0:
                split_at = 50
            first_name = collapsed_name[:split_at].strip() or collapsed_name[:50]
            last_name = collapsed_name[split_at:].strip()[:50]

        person = existing_person if _is_manual_recipient_person(existing_person) else Person()
        person.first_name = first_name[:50]
        person.last_name = last_name[:50]
        person.phone = "__MANUAL_INVOICE__"
        person.email = (email or "").strip() or None
        person.weight_kg = 0
        person.street_and_number = (street or "").strip() or None
        person.zip_code = (zip_code or "").strip() or None
        person.city = (city or "").strip() or None
        person.deleted_at = datetime.utcnow()
        person.deleted_reason = "System-Platzhalter für manuelle Rechnungen"
        person.comment = "Automatisch erzeugter Rechnungsempfänger für manuelle Rechnung"
        if person.full_name.strip() != collapsed_name:
            person.original_name = collapsed_name[:120]
        else:
            person.original_name = None
        if person.id is None:
            db.session.add(person)
            db.session.flush()
        return person

    def _cleanup_unsaved_drafts(*, keep_invoice_id: int | None = None) -> None:
        """
        Entfernt alle ungespeicherten Entwürfe (stage=draft), außer optional einem
        aktuell bearbeiteten Entwurf.
        """
        q = Invoice.query.filter(Invoice.stage == "draft")
        if keep_invoice_id is not None:
            q = q.filter(Invoice.id != keep_invoice_id)

        for old in q.all():
            for item in list(old.items or []):
                db.session.delete(item)
            db.session.delete(old)
        db.session.flush()

    draft_invoice_id_raw = (request.values.get("draft_invoice_id") or request.values.get("invoice_id") or "").strip()
    if not draft_invoice_id_raw:
        draft_invoice_id_raw = str(session.get("manual_edit_draft_invoice_id") or "").strip()
    draft_invoice = None
    draft_invoice_id = None
    if draft_invoice_id_raw:
        draft_invoice_id = _parse_int_loose(draft_invoice_id_raw)
        try:
            if draft_invoice_id is None:
                raise ValueError("invalid-draft-id")
            draft_invoice = (
                Invoice.query
                .options(selectinload(Invoice.items), joinedload(Invoice.person))
                .get(draft_invoice_id)
            )
        except Exception:
            draft_invoice = None
            draft_invoice_id = None

        if draft_invoice_id and not _is_manual_draft_invoice(draft_invoice):
            flash("Dieser Entwurf kann nicht über das manuelle Eingabeformular bearbeitet werden.", "warning")
            if draft_invoice:
                return redirect(url_for("billing.invoice_detail", invoice_id=draft_invoice.id))
            return redirect(url_for("billing.manual_invoice_new"))

    if request.method == "GET" and draft_invoice:
        session.pop("manual_edit_draft_invoice_id", None)
        session.pop("manual_edit_person_id", None)

    selected_person_id = (request.values.get("person_id") or "").strip()
    if not selected_person_id:
        selected_person_id = str(session.get("manual_edit_person_id") or "").strip()
    if not selected_person_id and draft_invoice:
        selected_person_id = str(draft_invoice.person_id)

    selected_person = None
    if selected_person_id:
        try:
            parsed_selected_person_id = _parse_int_loose(selected_person_id)
            selected_person = Person.query.get(parsed_selected_person_id) if parsed_selected_person_id is not None else None
        except Exception:
            selected_person = None
    if _is_manual_recipient_person(selected_person):
        selected_person = None

    # Fallback: Wenn kein expliziter draft_invoice_id gesetzt ist, aber eine Person
    # vorhanden ist, den aktuellen manuellen Entwurf der Person vorbefuellen.
    if not draft_invoice and selected_person:
        candidate_drafts = (
            Invoice.query
            .options(selectinload(Invoice.items), joinedload(Invoice.person))
            .filter(
                Invoice.person_id == selected_person.id,
                Invoice.stage == "draft",
            )
            .order_by(Invoice.created_at.desc(), Invoice.id.desc())
            .all()
        )
        for candidate in candidate_drafts:
            if _is_manual_draft_invoice(candidate):
                draft_invoice = candidate
                draft_invoice_id = candidate.id
                break

    if request.method == "GET":
        form_data = _build_manual_form_data_from_invoice(draft_invoice)
        return _render_manual_form(
            form_data=form_data,
            selected_person=selected_person,
            draft_invoice=draft_invoice,
        )

    draft_invoice_id_raw = (request.form.get("draft_invoice_id") or "").strip()
    draft_invoice = None
    draft_invoice_id = None
    if draft_invoice_id_raw:
        draft_invoice_id = _parse_int_loose(draft_invoice_id_raw)
        try:
            if draft_invoice_id is None:
                raise ValueError("invalid-draft-id")
            draft_invoice = (
                Invoice.query
                .options(selectinload(Invoice.items), joinedload(Invoice.person))
                .get(draft_invoice_id)
            )
        except Exception:
            draft_invoice = None
            draft_invoice_id = None

        if draft_invoice_id and not _is_manual_draft_invoice(draft_invoice):
            flash("Dieser Entwurf kann nicht bearbeitet werden.", "warning")
            if draft_invoice:
                return redirect(url_for("billing.invoice_detail", invoice_id=draft_invoice.id))
            return redirect(url_for("billing.manual_invoice_new"))

    form_data = _build_manual_form_data_from_request(
        request.form,
        draft_invoice_id=draft_invoice_id,
    )

    billing_address_name = form_data["billing_address_name"] or None
    billing_address_email = form_data["billing_address_email"] or None
    billing_address_street = form_data["billing_address_street"] or None
    billing_address_zip = form_data["billing_address_zip"] or None
    billing_address_city = form_data["billing_address_city"] or None

    person_id_raw = (request.form.get("person_id") or "").strip()
    person_id = _parse_int_loose(person_id_raw)
    person = Person.query.get(person_id) if person_id is not None else None
    if person_id is not None and not person:
        flash("Die ausgewählte Person wurde nicht gefunden.", "warning")
        return _render_manual_form(
            form_data=form_data,
            selected_person=None,
            draft_invoice=draft_invoice,
        )

    service_date_raw = (request.form.get("service_date") or "").strip()
    if not service_date_raw:
        flash("Bitte ein Leistungsdatum angeben.", "warning")
        return _render_manual_form(
            form_data=form_data,
            selected_person=person,
            draft_invoice=draft_invoice,
        )
    try:
        service_date = datetime.strptime(service_date_raw, "%Y-%m-%d").date()
    except Exception:
        flash("Ungültiges Leistungsdatum.", "warning")
        return _render_manual_form(
            form_data=form_data,
            selected_person=person,
            draft_invoice=draft_invoice,
        )

    manual_title = (request.form.get("manual_title") or "").strip() or "Manuelle Positionen"
    manual_title = manual_title[:120]

    descriptions = request.form.getlist("line_description[]")
    quantities = request.form.getlist("line_quantity[]")
    units = request.form.getlist("line_unit[]")
    unit_prices = request.form.getlist("line_unit_price_gross[]")
    vat_rates = request.form.getlist("line_vat_rate[]")

    manual_lines = []
    for i, desc in enumerate(descriptions):
        desc = (desc or "").strip()
        qty_raw = (quantities[i] if i < len(quantities) else "").strip()
        unit_label = (units[i] if i < len(units) else "").strip()
        unit_raw = (unit_prices[i] if i < len(unit_prices) else "").strip()
        vat_raw = (vat_rates[i] if i < len(vat_rates) else "").strip()

        if not desc and not qty_raw and not unit_label and not unit_raw and not vat_raw:
            continue

        try:
            qty = _parse_decimal_de(qty_raw, allow_negative=True) if qty_raw else Decimal("0")
            unit = _parse_decimal_de(unit_raw, allow_negative=True) if unit_raw else Decimal("0")
            vat_rate = _parse_decimal_de(vat_raw) if vat_raw else Decimal("0")
        except Exception:
            flash(f"Ungültige Zahlenwerte in Position {i + 1}.", "warning")
            return _render_manual_form(
                form_data=form_data,
                selected_person=person,
                draft_invoice=draft_invoice,
            )

        manual_lines.append(
            {
                "description": desc,
                "quantity": qty,
                "manual_unit": unit_label,
                "unit_price_gross": unit,
                "vat_rate": vat_rate,
                "manual_position_code": "manual",
            }
        )

    if not manual_lines:
        flash("Bitte mindestens eine gültige Position eingeben.", "warning")
        return _render_manual_form(
            form_data=form_data,
            selected_person=person,
            draft_invoice=draft_invoice,
        )

    preview_total = Decimal("0.00")
    for line in manual_lines:
        try:
            preview_total += Decimal(str(line.get("quantity") or "0")) * Decimal(str(line.get("unit_price_gross") or "0"))
        except Exception:
            pass

    prepaid_voucher_raw = request.form.get("prepaid_voucher_amount") or ""
    parsed_prepaid, prepaid_error = _parse_prepaid_amount(
        prepaid_voucher_raw,
        total_amount=preview_total,
        allow_prepaid=True,
    )
    if prepaid_error:
        flash(prepaid_error, "warning")
        return _render_manual_form(
            form_data=form_data,
            selected_person=person,
            draft_invoice=draft_invoice,
        )

    _cleanup_unsaved_drafts(
        keep_invoice_id=(draft_invoice.id if draft_invoice else None),
    )

    if person is None:
        if not billing_address_name:
            flash("Bitte eine Person auswählen oder einen Rechnungsempfänger eingeben.", "warning")
            return _render_manual_form(
                form_data=form_data,
                selected_person=None,
                draft_invoice=draft_invoice,
            )

        person = _create_manual_recipient_person(
            name=billing_address_name,
            email=billing_address_email,
            street=billing_address_street,
            zip_code=billing_address_zip,
            city=billing_address_city,
            existing_person=draft_invoice.person if draft_invoice else None,
        )

    if draft_invoice:
        draft_invoice.person_id = person.id
        draft_invoice.service_date = service_date
        draft_invoice.manual_title = manual_title
        draft_invoice.billing_address_name = billing_address_name
        draft_invoice.billing_address_street = billing_address_street
        draft_invoice.billing_address_zip = billing_address_zip
        draft_invoice.billing_address_city = billing_address_city
        draft_invoice.billing_address_email = billing_address_email

        for item in list(draft_invoice.items or []):
            db.session.delete(item)
        db.session.flush()

        for line in manual_lines:
            gross = (Decimal(str(line.get("quantity") or "0")) * Decimal(str(line.get("unit_price_gross") or "0"))).quantize(Decimal("0.01"))
            net, vat = BillingService.split_gross_into_net_and_vat(
                gross=gross,
                vat_rate=Decimal(str(line.get("vat_rate") or "0")),
            )
            db.session.add(
                InvoiceItem(
                    invoice_id=draft_invoice.id,
                    load_entry_id=None,
                    amount=gross,
                    vat_rate=Decimal(str(line.get("vat_rate") or "0")),
                    net_amount=net,
                    vat_amount=vat,
                    description=line.get("description") or "",
                    item_source="manual",
                    quantity=Decimal(str(line.get("quantity") or "0")),
                    manual_unit=(line.get("manual_unit") or "").strip() or None,
                    unit_price_gross=Decimal(str(line.get("unit_price_gross") or "0")),
                    manual_position_code=(line.get("manual_position_code") or "manual"),
                )
            )

        draft_invoice.calculate_total()
        draft_invoice.prepaid_voucher_amount = parsed_prepaid
        db.session.commit()
        invoice = draft_invoice
    else:
        invoice = BillingService.create_manual_invoice(
            person_id=person.id,
            service_date=service_date,
            manual_title=manual_title,
            manual_lines=manual_lines,
            billing_address_name=billing_address_name,
            billing_address_street=billing_address_street,
            billing_address_zip=billing_address_zip,
            billing_address_city=billing_address_city,
            billing_address_email=billing_address_email,
            prepaid_voucher_amount=parsed_prepaid,
        )
        if invoice:
            db.session.commit()

    if not invoice:
        flash("Manuelle Rechnung konnte nicht erstellt werden. Prüfen Sie die Positionen.", "warning")
        return _render_manual_form(
            form_data=form_data,
            selected_person=person,
            draft_invoice=draft_invoice,
        )

    if draft_invoice:
        flash("Entwurf wurde aktualisiert.", "success")
    else:
        flash("Manuelle Rechnung wurde als Entwurf erstellt.", "success")
    return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number_for_detail(invoice)))


@bp.route("/invoice/<int:invoice_id>/edit", methods=["GET"])
def invoice_edit(invoice_id):
    invoice = (
        Invoice.query
        .options(selectinload(Invoice.items), joinedload(Invoice.person))
        .get_or_404(invoice_id)
    )

    if invoice.stage != "draft":
        flash("Nur Entwürfe können bearbeitet werden.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice)))

    if not _is_manual_draft_invoice(invoice):
        flash("Dieser Entwurf ist kein manueller Rechnungsentwurf.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice)))

    session["manual_edit_draft_invoice_id"] = invoice.id
    session["manual_edit_person_id"] = invoice.person_id

    return redirect(
        url_for(
            "billing.manual_invoice_new",
            draft_invoice_id=invoice.id,
            person_id=invoice.person_id,
        )
    )


# ---------------------------------------------------------
# Personenübersicht mit KORREKTEN Summen (Sprünge + Schirmmiete + Orga)
# ---------------------------------------------------------
@bp.route("/persons")
def persons_overview():
    search = request.args.get("search", "").strip()
    filters_str = request.args.get("filters", "").strip()
    filters_list = [f.strip() for f in filters_str.split(",") if f.strip()] if filters_str else []
    sort = request.args.get("sort", "last_name")
    direction = request.args.get("direction", "asc")
    if direction not in {"asc", "desc"}:
        direction = "asc"

    query = (
        db.session.query(Person)
        .join(LoadEntry, LoadEntry.person_id == Person.id)
        .filter(LoadEntry.billed.is_(False))
    )

    if "archived" in filters_list:
        query = query.filter(Person.deleted_at.isnot(None))
    else:
        query = query.filter(Person.deleted_at.is_(None))

    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                Person.first_name.ilike(like),
                Person.last_name.ilike(like),
                Person.phone.ilike(like),
                Person.email.ilike(like),
            )
        )
        if search.lower() == "lehrer":
            query = query.filter(Person.is_teacher.is_(True))
        if search.lower() == "gast":
            query = query.filter(
                Person.is_member.is_(False),
                Person.is_tandem_guest.is_(False),
                Person.is_partner_verein.is_(False),
            )

    for f in filters_list:
        if f == "members":
            query = query.filter(Person.is_member.is_(True))
        elif f == "partner":
            query = query.filter(Person.is_partner_verein.is_(True))
        elif f == "tandem":
            query = query.filter(Person.is_tandem_guest.is_(True))
        elif f == "teacher":
            query = query.filter(Person.is_teacher.is_(True))
        elif f == "aff_teacher":
            query = query.filter(Person.is_aff_teacher.is_(True))
        elif f == "student":
            query = query.filter(Person.is_student.is_(True))
        elif f == "video":
            query = query.filter(Person.is_video.is_(True))
        elif f == "aff_student":
            query = query.filter(Person.is_aff_student.is_(True))
        elif f == "tandemmaster":
            query = query.filter(Person.is_tandemmaster.is_(True))
        elif f == "guest":
            query = query.filter(
                Person.is_member.is_(False),
                Person.is_tandem_guest.is_(False),
                Person.is_partner_verein.is_(False),
            )
        elif f == "liability_ok":
            query = query.filter(Person.liability_waiver_date.isnot(None))
        elif f == "liability_bad":
            query = query.filter(
                or_(
                    Person.liability_waiver_date.is_(None),
                    Person.liability_waiver_date < date(date.today().year, 1, 1),
                )
            )
        elif f == "weight_bad":
            query = query.filter(
                or_(
                    Person.weight_kg.is_(None),
                    Person.weight_kg < 40,
                    Person.weight_kg > 120,
                )
            )

    persons = (
        query
        .group_by(Person.id)
        .order_by(Person.last_name.asc(), Person.first_name.asc())
        .all()
    )

    rows = []
    billing_config = BillingConfig.query.first()

    for person in persons:
        open_entries = BillingService.get_open_entries_for_person(person.id)
        if not open_entries:
            continue

        current_statuses = {
            (getattr(entry, "status_code", "") or "").strip()
            for entry in open_entries
            if (getattr(entry, "status_code", "") or "").strip()
        }

        # 1) Sprungpreise
        total_jump = Decimal("0.00")
        for e in open_entries:
            try:
                total_jump += Decimal(
                    str(BillingService.calculate_price_for_entry(e) or "0.00")
                )
            except Exception:
                pass

        # 2) Schirmmiete (Preview)
        rent_total = Decimal("0.00")
        if billing_config:
            rent_counts = defaultdict(int)
            for e in open_entries:
                try:
                    if not BillingService._is_rent_eligible(e):
                        continue
                    day = BillingService._entry_day(e)
                    cat = BillingService._rent_category_from_status(e.status_code)
                    if cat:
                        rent_counts[(day, cat)] += 1
                except Exception:
                    continue

            for (_day, cat), count in rent_counts.items():
                try:
                    price_per_jump, max_count, _ = BillingService._rent_params(cat, billing_config)
                    price_per_jump = Decimal(str(price_per_jump or "0.00"))
                    max_count = int(max_count or 0)
                    charged = min(count, max_count) if max_count > 0 else count
                    rent_total += price_per_jump * Decimal(charged)
                except Exception:
                    continue


        # 3) Orga (Preview, jetzt korrekt: nur wenn noch nicht abgerechnet)
        orga_total = Decimal("0.00")
        try:
            ctx_map = defaultdict(list)
            for e in open_entries:
                ld = e.load
                if not ld:
                    continue
                model_id = getattr(ld, "pricing_model_id", None)
                if model_id:
                    pid = int(model_id)
                else:
                    d0 = BillingService._entry_day(e)
                    p0 = BillingService.get_active_price_period(
                        day=d0
                    )
                    pid = p0.id if p0 else None
                if pid is None:
                    continue
                ctx_map[pid].append(e)

            _orga_persons_merged: dict = {}
            for pid, entries_ctx in ctx_map.items():
                orga_amount, orga_mode, _ = BillingService._get_orga_config(
                    period_id=pid,
                )
                orga_amount = Decimal(str(orga_amount or "0.00"))
                _pm_key = (orga_amount, orga_mode or "period")
                if _pm_key not in _orga_persons_merged:
                    _orga_persons_merged[_pm_key] = [pid, list(entries_ctx)]
                else:
                    _orga_persons_merged[_pm_key][1].extend(entries_ctx)
                    if pid > _orga_persons_merged[_pm_key][0]:
                        _orga_persons_merged[_pm_key][0] = pid

            for (orga_amount, orga_mode), pm_orga_data in _orga_persons_merged.items():
                pid = pm_orga_data[0]
                entries_ctx = pm_orga_data[1]

                rules = (
                    BillingOrgaRule.query
                    .filter_by(period_id=pid)
                    .all()
                )
                rules_map = {r.status_code: bool(r.apply_orga) for r in rules}

                def _is_relevant(x, _rules_map=rules_map):
                    code = normalize_status_code(getattr(x, "status_code", "") or "")
                    return _rules_map.get(code, True)

                relevant = [x for x in entries_ctx if _is_relevant(x)]
                if not relevant:
                    continue

                days = sorted({BillingService._entry_day(x) for x in relevant})
                person_id = relevant[0].person_id if relevant else None
                if not person_id:
                    continue
                invoice_nr, abgerechnete_tage = BillingService._find_existing_orga_invoice(person_id, pid, orga_mode, days)
                if orga_mode == "day":
                    # Nur Tage berechnen, die noch nicht abgerechnet wurden
                    tage_offen = [d for d in days if d.strftime('%d.%m.%Y') not in abgerechnete_tage]
                    orga_total += orga_amount * Decimal(len(tage_offen))
                else:
                    # Nur berechnen, wenn noch keine Orga-Rechnung existiert
                    if not invoice_nr:
                        orga_total += orga_amount
        except Exception:
            pass

        main_status = _billing_person_main_status(person)
        secondary_statuses = _billing_person_secondary_statuses(
            current_statuses,
            main_status,
        )

        rows.append({
            "person": person,
            "count": len(open_entries),
            "amount": total_jump + rent_total + orga_total,
            "main_status": main_status,
            "secondary_statuses": secondary_statuses,
        })

    reverse = direction == "desc"
    if sort == "count":
        rows = sorted(
            rows,
            key=lambda r: (int(r.get("count", 0)), (r["person"].last_name or "").lower(), (r["person"].first_name or "").lower()),
            reverse=reverse,
        )
    elif sort == "amount":
        rows = sorted(
            rows,
            key=lambda r: (Decimal(str(r.get("amount") or "0.00")), (r["person"].last_name or "").lower(), (r["person"].first_name or "").lower()),
            reverse=reverse,
        )
    elif sort == "main_status":
        rows = sorted(
            rows,
            key=lambda r: ((r.get("main_status") or "").lower(), (r["person"].last_name or "").lower(), (r["person"].first_name or "").lower()),
            reverse=reverse,
        )
    else:
        rows = sorted(
            rows,
            key=lambda r: ((r["person"].last_name or "").lower(), (r["person"].first_name or "").lower()),
            reverse=reverse,
        )

    return render_template(
        "billing/persons_overview.html",
        rows=rows,
        search=search,
        filters_str=filters_str,
        sort=sort,
        direction=direction,
    )


# ---------------------------------------------------------
# Detailansicht einer Person (Vorschau ≈ Rechnung)
# ---------------------------------------------------------
@bp.route("/person/<int:person_id>", endpoint="person_billing")
def person_billing(person_id):

    from app import db
    db.session.expire_all()  # Verhindert doppelte Anzeige durch ORM-Cache nach Rechnungserstellung
    person = Person.query.get_or_404(person_id)
    open_entries = BillingService.get_open_entries_for_person(person_id)
    default_tandem_kleinunternehmer = _person_tandem_ku_default(person)
    default_video_kleinunternehmer = _person_video_ku_default(person)
    preview_tandem_ku_enabled = bool(getattr(person, "is_tandemmaster", False)) and default_tandem_kleinunternehmer
    preview_video_ku_enabled = bool(getattr(person, "is_video", False)) and default_video_kleinunternehmer

    entry_rows = []
    total_jump = Decimal("0.00")

    billing_config = BillingConfig.query.first()
    rent_counts = defaultdict(int)
    days_set = set()

    for e in open_entries:
        load = e.load
        dt = (
            load.actual_time
            or load.scheduled_time
            or load.created_at
            if load else e.created_at
        )
        day = dt.date() if dt else None
        if day:
            days_set.add(day)

        price = Decimal(str(BillingService.calculate_price_for_entry(e) or "0.00"))
        total_jump += price

        base_vat_rate = Decimal(str(BillingService.get_entry_vat_rate(e) or "0.00"))
        is_tandemmaster_jump = BillingService._is_tandemmaster_jump_entry(e)
        is_video_jump = BillingService._is_video_jump_entry(e)
        ku_active_for_entry = (
            (preview_tandem_ku_enabled and is_tandemmaster_jump)
            or (preview_video_ku_enabled and is_video_jump)
        )
        vat_rate = Decimal("0.00") if ku_active_for_entry else base_vat_rate
        net, vat = BillingService.split_gross_into_net_and_vat(gross=price, vat_rate=vat_rate)

        entry_rows.append({
            "entry": e,
            "date": dt,
            "load_number": load.load_number if load else None,
            "status_code": e.status_code,
            "height_m": e.height_m,
            "price": price,
            "net": net,
            "vat": vat,
            "vat_rate": vat_rate,
            "base_vat_rate": base_vat_rate,
            "is_tandemmaster_jump": is_tandemmaster_jump,
            "is_video_jump": is_video_jump,
            "ku_eligible": bool(is_tandemmaster_jump or is_video_jump),
            "ku_notice": bool(ku_active_for_entry),
        })

        if billing_config and BillingService._is_rent_eligible(e) and day:
            cat = BillingService._rent_category_from_status(e.status_code)
            if cat:
                rent_counts[(day, cat)] += 1

    days_sorted = sorted(days_set)

    rent_lines = []
    rent_total = Decimal("0.00")
    if billing_config:
        for (day, cat), count in rent_counts.items():
            price_per_jump, max_count, vat_rate = BillingService._rent_params(cat, billing_config)
            price_per_jump = Decimal(str(price_per_jump or "0.00"))
            vat_rate = Decimal(str(vat_rate or "0.00"))
            max_count = int(max_count or 0)

            charged = min(count, max_count) if max_count > 0 else 0
            gross = price_per_jump * Decimal(charged)
            rent_total += gross

            net, vat = BillingService.split_gross_into_net_and_vat(gross=gross, vat_rate=vat_rate)
            rent_lines.append({
                "day": day,
                "category": cat,
                "charged": charged,
                "price": price_per_jump,
                "subtotal": gross,
                "net": net,
                "vat": vat,
                "vat_rate": vat_rate
            })

    # Orga-Vorschau (korrigiert: mehrere Kontexte möglich)
    #
    # Für die UI behalten wir:
    # - orga_total = Gesamt-Orga
    # - orga_amount/org_mode = nur "Hauptanzeige" (kompatibel),
    #   zusätzlich orga_lines für detaillierte Anzeige
    orga_amount = Decimal("0.00")
    orga_mode = None
    orga_total = Decimal("0.00")
    orga_lines = []
    days_sorted = sorted(days_set)

    try:
        ctx_map = defaultdict(list)

        for e in open_entries:
            ld = e.load
            if not ld:
                continue

            model_id = getattr(ld, "pricing_model_id", None)
            if model_id:
                pid = int(model_id)
            else:
                d0 = BillingService._entry_day(e)
                p0 = BillingService.get_active_price_period(
                    day=d0
                )
                pid = p0.id if p0 else None

            if pid is None:
                continue

            ctx_map[pid].append(e)

        # Perioden-Gruppen mit identischer Orga-Config (gleicher Betrag + Modus) zusammenführen,
        # damit die Pauschale nicht mehrfach berechnet wird, wenn Entries Loads mit
        # unterschiedlichen pricing_model_id-Werten umfassen, die aber dieselbe Orga haben.
        orga_merged_preview: dict = {}
        for pid, entries_ctx in ctx_map.items():
            _preview_amt, _preview_mode, _ = BillingService._get_orga_config(period_id=pid)
            _preview_key = (Decimal(str(_preview_amt or "0.00")), _preview_mode or "period")
            if _preview_key not in orga_merged_preview:
                orga_merged_preview[_preview_key] = [pid, list(entries_ctx)]
            else:
                orga_merged_preview[_preview_key][1].extend(entries_ctx)
                if pid > orga_merged_preview[_preview_key][0]:
                    orga_merged_preview[_preview_key][0] = pid

        for orga_preview_data in orga_merged_preview.values():
            pid = orga_preview_data[0]
            entries_ctx = orga_preview_data[1]

            amt, mode, _ = BillingService._get_orga_config(period_id=pid)
            amt = Decimal(str(amt or "0.00"))

            rules = (
                BillingOrgaRule.query
                .filter_by(period_id=pid)
                .all()
            )
            rules_map = {r.status_code: bool(r.apply_orga) for r in rules}

            def _is_relevant(x, _rules_map=rules_map):
                code = normalize_status_code(getattr(x, "status_code", "") or "")
                return _rules_map.get(code, True)

            relevant = [x for x in entries_ctx if _is_relevant(x)]
            if not relevant:
                continue

            vat_rates = [BillingService.get_entry_vat_rate(x) for x in relevant]
            vat_rate = max(vat_rates) if vat_rates else Decimal("0.00")

            days = sorted({BillingService._entry_day(x) for x in relevant})
            person_id = relevant[0].person_id if relevant else None
            if not person_id:
                continue
            invoice_nr, abgerechnete_tage = BillingService._find_existing_orga_invoice(person_id, pid, mode, days)
            if mode == "day":
                for d in days:
                    tag_str = d.strftime('%d.%m.%Y')
                    if tag_str in abgerechnete_tage:
                        # Bereits abgerechnet: Hinweiszeile mit 0 € und Rechnungsnummer
                        orga_lines.append({
                            "period_id": pid,
                            "mode": "day",
                            "amount": amt,
                            "days": [d],
                            "gross": Decimal("0.00"),
                            "net": Decimal("0.00"),
                            "vat": Decimal("0.00"),
                            "vat_rate": vat_rate,
                            "description": "Organisationspauschale",
                            "invoice_nr": invoice_nr,
                        })
                    else:
                        _net, _vat = BillingService.split_gross_into_net_and_vat(gross=amt, vat_rate=vat_rate)
                        orga_total += amt
                        orga_lines.append({
                            "period_id": pid,
                            "mode": "day",
                            "amount": amt,
                            "days": [d],
                            "gross": amt,
                            "net": _net,
                            "vat": _vat,
                            "vat_rate": vat_rate,
                            "description": "Organisationspauschale",
                            "invoice_nr": None,
                        })
            else:
                if invoice_nr:
                    # Bereits abgerechnet: Hinweiszeile mit 0 € und Rechnungsnummer
                    orga_lines.append({
                        "period_id": pid,
                        "mode": "period",
                        "amount": amt,
                        "days": days,
                        "gross": Decimal("0.00"),
                        "net": Decimal("0.00"),
                        "vat": Decimal("0.00"),
                        "vat_rate": vat_rate,
                        "description": "Organisationspauschale",
                        "invoice_nr": invoice_nr,
                    })
                else:
                    _net, _vat = BillingService.split_gross_into_net_and_vat(gross=amt, vat_rate=vat_rate)
                    orga_total += amt
                    orga_lines.append({
                        "period_id": pid,
                        "mode": "period",
                        "amount": amt,
                        "days": days,
                        "gross": amt,
                        "net": _net,
                        "vat": _vat,
                        "vat_rate": vat_rate,
                        "description": "Organisationspauschale",
                        "invoice_nr": None,
                    })

        # Kompatibilität: falls genau eine Orga-Line existiert
        if len(orga_lines) == 1:
            orga_amount = orga_lines[0]["amount"]
            orga_mode = orga_lines[0]["mode"]
        elif len(orga_lines) > 1:
            # mehrere Kontexte: in der alten Summary zeigen wir nur "Gesamt"
            orga_amount = Decimal("0.00")
            orga_mode = "multi"

    except Exception:
        pass


    total_preview = total_jump + rent_total + orga_total
    prepaid_allowed = _entries_allow_prepaid_voucher(open_entries)

    # Netto/MwSt Summen (optional für Anzeige)
    fixed_net_preview = (
        sum(rl["net"] for rl in rent_lines)
        + sum(ol.get("net", Decimal("0.00")) for ol in orga_lines)
    )
    fixed_vat_preview = (
        sum(rl["vat"] for rl in rent_lines)
        + sum(ol.get("vat", Decimal("0.00")) for ol in orga_lines)
    )
    total_net_preview = (
        sum(r["net"] for r in entry_rows)
        + fixed_net_preview
    )
    total_vat_preview = (
        sum(r["vat"] for r in entry_rows)
        + fixed_vat_preview
    )

    return render_template(
        "billing/person_billing.html",
        person=person,
        entry_rows=entry_rows,
        total_jump=total_jump,
        rent_lines=rent_lines,
        rent_total=rent_total,
        orga_amount=orga_amount,
        orga_mode=orga_mode,
        orga_total=orga_total,
        orga_lines=orga_lines,
        days_sorted=days_sorted,
        total_preview=total_preview,
        prepaid_allowed=prepaid_allowed,
        prepaid_voucher_amount=Decimal("0.00"),
        default_tandem_kleinunternehmer=default_tandem_kleinunternehmer,
        default_video_kleinunternehmer=default_video_kleinunternehmer,
        preview_tandem_ku_enabled=preview_tandem_ku_enabled,
        preview_video_ku_enabled=preview_video_ku_enabled,
        fixed_net_preview=fixed_net_preview,
        fixed_vat_preview=fixed_vat_preview,
        total_net_preview=total_net_preview,
        total_vat_preview=total_vat_preview,
    )


# ---------------------------------------------------------
# Rechnung für eine Person erstellen
# ---------------------------------------------------------
@bp.route("/person/<int:person_id>/create_invoice", methods=["POST"])
def create_invoice_for_person(person_id):
    person = Person.query.get_or_404(person_id)
    # Abweichende Rechnungsanschrift/E-Mail aus dem Formular übernehmen
    billing_address_name = request.form.get("billing_address_name") or None
    billing_address_street = request.form.get("billing_address_street") or None
    billing_address_zip = request.form.get("billing_address_zip") or None
    billing_address_city = request.form.get("billing_address_city") or None
    billing_address_email = request.form.get("billing_address_email") or None
    prepaid_voucher_raw = request.form.get("prepaid_voucher_amount") or ""
    person_ku_default = _person_tandem_ku_default(person)
    person_video_ku_default = _person_video_ku_default(person)
    ku_form_value = request.form.get("invoice_is_tandem_kleinunternehmer")
    video_ku_form_value = request.form.get("invoice_is_video_kleinunternehmer")
    if ku_form_value is None or str(ku_form_value).strip() == "":
        is_tandem_kleinunternehmer = person_ku_default
    else:
        is_tandem_kleinunternehmer = _parse_form_bool(
            ku_form_value,
            default=person_ku_default,
        )
    if video_ku_form_value is None or str(video_ku_form_value).strip() == "":
        is_video_kleinunternehmer = person_video_ku_default
    else:
        is_video_kleinunternehmer = _parse_form_bool(
            video_ku_form_value,
            default=person_video_ku_default,
        )
    if not bool(getattr(person, "is_tandemmaster", False)):
        is_tandem_kleinunternehmer = False
    if not bool(getattr(person, "is_video", False)):
        is_video_kleinunternehmer = False

    # Vorschau-Eingabe validieren: nur Tandem-/Mitflieger-Einträge dürfen Vorkasse/Gutschein nutzen.
    open_entries_for_precheck = BillingService.get_open_entries_for_person(person_id)
    allow_prepaid = _entries_allow_prepaid_voucher(open_entries_for_precheck)
    parsed_prepaid, prepaid_error = _parse_prepaid_amount(
        prepaid_voucher_raw,
        total_amount=Decimal("999999.99"),
        allow_prepaid=allow_prepaid,
    )
    if prepaid_error:
        flash(prepaid_error, "warning")
        return redirect(url_for("billing.person_billing", person_id=person_id))

    invoice = BillingService.create_invoice_for_person(
        person_id,
        billing_address_name=billing_address_name,
        billing_address_street=billing_address_street,
        billing_address_zip=billing_address_zip,
        billing_address_city=billing_address_city,
        billing_address_email=billing_address_email,
        prepaid_voucher_amount=parsed_prepaid,
        is_tandem_kleinunternehmer=is_tandem_kleinunternehmer,
        is_video_kleinunternehmer=is_video_kleinunternehmer,
    )

    if not invoice:
        flash("Keine offenen Sprünge für diese Person.", "warning")
        return redirect(url_for("billing.person_billing", person_id=person_id))

    if not (getattr(invoice, "payment_method", "") or "").strip() and _invoice_allows_sepa(invoice):
        invoice.payment_method = "sepa"

    flash("Rechnung wurde erstellt.", "success")
    return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice)))


# ---------------------------------------------------------
# Rechnungsdetails anzeigen
# ---------------------------------------------------------
@bp.route("/invoice/<int:invoice_id>")
def invoice_detail(invoice_id):
    try:
        # URL-Nummer: seq_number bevorzugt, dann id-Fallback,
        # plus virtuelle Entwurfsnummern fuer konsistente Vorschauanzeige.
        invoice = _get_invoice_by_display_number(invoice_id)

        if not invoice:
            return "Invoice NOT FOUND", 404

        if invoice.is_deleted and not session.get("is_admin"):
            flash("Stornierte Rechnungen sind nur im Admin-Modus sichtbar.", "warning")
            return redirect(url_for("billing.invoice_list"))

        billing_config = BillingConfig.query.first()
        invoice_display_number = _invoice_display_number_for_detail(invoice)

        invoice_purpose = _build_invoice_payment_purpose(
            invoice,
            invoice_number=invoice_display_number,
        )

        # -----------------------------------------------------
        # EPC-Zahlungs-QR (dynamisch pro Rechnung)
        # Nur wenn ein positiver Zahlbetrag vorliegt (keine Gutschrift).
        # Empfänger (Kontoinhaber) = Rechnungssteller (BillingConfig)
        # Springer (invoice.person) steht zusätzlich im Verwendungszweck.
        # -----------------------------------------------------
        epc_qr_data_uri = None
        prepaid_voucher_amount = _invoice_prepaid_amount(invoice)
        onsite_amount = _invoice_onsite_amount(invoice)
        prepaid_allowed = _invoice_allows_prepaid_voucher(invoice)
        try:
            amount = onsite_amount
            if amount > 0 and billing_config and getattr(billing_config, "iban", None):
                remittance = invoice_purpose

                payment_context = build_payment_context(
                    invoice=invoice,
                    billing_config=billing_config,
                    invoice_number=invoice_display_number,
                    amount_eur=amount,
                )
                payload = payment_context["epc_payload"]
                epc_qr_data_uri = _make_qr_data_uri(payload)
        except Exception:
            epc_qr_data_uri = None

        # --------------------------------------------------
        # Bilder als Data-URIs laden (für PDF-Kompatibilität)
        # --------------------------------------------------
        static_dir = os.path.join(current_app.root_path, "static")
        logo_data_uri = None
        if billing_config and billing_config.logo_filename:
            logo_path = os.path.join(static_dir, "img", billing_config.logo_filename)
            logo_data_uri = _image_to_data_uri(logo_path)

        qr_instagram_data_uri = None
        if billing_config and billing_config.qr_instagram_filename:
            qr_path = os.path.join(static_dir, "img", "qr", billing_config.qr_instagram_filename)
            qr_instagram_data_uri = _image_to_data_uri(qr_path)

        qr_facebook_data_uri = None
        if billing_config and billing_config.qr_facebook_filename:
            qr_path = os.path.join(static_dir, "img", "qr", billing_config.qr_facebook_filename)
            qr_facebook_data_uri = _image_to_data_uri(qr_path)

        qr_website_data_uri = None
        if billing_config and billing_config.qr_website_filename:
            qr_path = os.path.join(static_dir, "img", "qr", billing_config.qr_website_filename)
            qr_website_data_uri = _image_to_data_uri(qr_path)

        # Ist gerade ein Versand-Job aktiv? (In-Memory Progress Store)
        _job_progress = get_progress(invoice.id)
        email_job_is_live = (
            _job_progress is not None
            and _job_progress.get("status") == "processing"
        )

        # Letzter Fehler (nur wenn neuer als letzter Erfolg)
        _last_err = getattr(invoice, "email_last_error", None)
        _last_attempt = getattr(invoice, "email_last_attempt_at", None)
        _last_success = getattr(invoice, "email_sent_at", None)
        email_last_error_active = bool(
            _last_err
            and _last_attempt
            and (not _last_success or _last_attempt >= _last_success)
        )

        # Dynamische KU-Umschaltung in der Rechnungsansicht:
        # Fixe Summen (alle nicht betroffenen Positionen) + Basis-MwSt je KU-relevanter Sprungzeile.
        invoice_ku_regular_vat_rates: dict[int, Decimal] = {}
        invoice_dynamic_fixed_net = Decimal("0.00")
        invoice_dynamic_fixed_vat = Decimal("0.00")
        for _item in list(getattr(invoice, "items", []) or []):
            if _is_jump_item(_item):
                _entry = getattr(_item, "load_entry", None)
                _base_rate = BillingService.get_entry_vat_rate(_entry) if _entry else Decimal("0.00")
                if getattr(_item, "id", None) is not None:
                    invoice_ku_regular_vat_rates[int(_item.id)] = Decimal(str(_base_rate or "0.00"))
                continue
            invoice_dynamic_fixed_net += Decimal(str(getattr(_item, "net_amount", 0) or 0))
            invoice_dynamic_fixed_vat += Decimal(str(getattr(_item, "vat_amount", 0) or 0))

        invoice_dynamic_fixed_net = invoice_dynamic_fixed_net.quantize(Decimal("0.01"))
        invoice_dynamic_fixed_vat = invoice_dynamic_fixed_vat.quantize(Decimal("0.01"))

        _tpl_kwargs = dict(
            invoice=invoice,
            invoice_display_number=invoice_display_number,
            invoice_purpose=invoice_purpose,
            billing_config=billing_config,
            epc_qr_data_uri=epc_qr_data_uri,
            logo_data_uri=logo_data_uri,
            qr_instagram_data_uri=qr_instagram_data_uri,
            qr_facebook_data_uri=qr_facebook_data_uri,
            qr_website_data_uri=qr_website_data_uri,
            is_dev_mode=is_dev_mode(),
            is_admin=session.get("is_admin", False),
            email_job_is_live=email_job_is_live,
            email_last_error_active=email_last_error_active,
            prepaid_voucher_amount=prepaid_voucher_amount,
            onsite_amount=onsite_amount,
            prepaid_allowed=prepaid_allowed,
            invoice_allows_sepa=_invoice_allows_sepa(invoice),
            invoice_payment_state_code=_invoice_payment_state(invoice),
            invoice_payment_state_label=_invoice_payment_state_label(invoice),
            invoice_has_tandem_jump_positions=_invoice_has_tandem_jump_positions(invoice),
            invoice_has_video_jump_positions=_invoice_has_video_jump_positions(invoice),
            invoice_has_ku_jump_positions=_invoice_has_ku_jump_positions(invoice),
            invoice_ku_regular_vat_rates=invoice_ku_regular_vat_rates,
            invoice_dynamic_fixed_net=invoice_dynamic_fixed_net,
            invoice_dynamic_fixed_vat=invoice_dynamic_fixed_vat,
        )

        # Partial-Reload (AJAX)
        if request.headers.get("X-Partial"):
            return render_template("billing/invoice_detail.html", **_tpl_kwargs)

        return render_template("billing/invoice_detail.html", **_tpl_kwargs)

    except Exception as e:
        return f"INVOICE DETAIL ERROR:<br><pre>{e}</pre>", 500


# ---------------------------------------------------------
# Entwurf verwerfen (Draft-Rechnung löschen)
# ---------------------------------------------------------
@bp.route("/invoice/<int:invoice_id>/discard", methods=["POST"])
def invoice_discard(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)

    # ✅ Nur Entwürfe dürfen verworfen werden
    if invoice.stage != "draft":
        flash("Nur Entwurfs-Rechnungen können verworfen werden.", "warning")
        return redirect(
            url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice))
        )

    person_id = invoice.person_id

    # ✅ Entwurf + Items löschen (keine billed-Flags!)
    db.session.delete(invoice)
    db.session.commit()

    flash("Rechnungsentwurf wurde verworfen.", "success")
    return redirect(
        url_for("billing.person_billing", person_id=person_id)
    )

# ---------------------------------------------------------
# Transaktionskosten in drei Versionen
# ---------------------------------------------------------
@bp.route("/invoice/<int:invoice_id>/transaction_fee", methods=["POST"])
def invoice_set_transaction_fee(invoice_id):
    # Helper: AJAX-Erkennung (fetch/XHR)
    def _is_ajax(req) -> bool:
        return req.headers.get("X-Requested-With") == "XMLHttpRequest"

    # Helper: einheitlicher Abschluss (AJAX: 204 / normal: flash+redirect)
    def _finish(message: str, category: str = "success"):
        if _is_ajax(request):
            # Kein Redirect, kein Flash -> Template macht danach Partial Reload
            return ("", 204)
        flash(message, category)
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    deny = _full_admin_required("billing.invoice_detail", invoice_id=invoice_id)
    if deny:
        # Bei AJAX sauber als 403 zurückgeben
        if _is_ajax(request):
            return ("", 403)
        return deny

    invoice = Invoice.query.get_or_404(invoice_id)
    cfg = BillingConfig.query.first()

    mode = (request.form.get("transaction_fee_mode") or "none").strip()
    if mode not in {"none", "fixed", "percent"}:
        mode = "none"

    # vorhandene Transaktionskosten-Items entfernen (falls vorhanden)
    def is_fee_item(it):
        desc = (it.description or "")
        return desc.startswith("Transaktionskosten") or desc.startswith("Transaktionsgebühr")

    for it in list(invoice.items):
        if is_fee_item(it):
            db.session.delete(it)

    # Falls keine Config oder none -> fertig
    if not cfg or mode == "none":
        invoice.calculate_total()
        db.session.commit()
        return _finish("Transaktionskosten entfernt.", "success")

    # Basisbetrag für Prozentberechnung = Summe aller bestehenden Posten (ohne Fee)
    base_gross = Decimal("0.00")
    for it in invoice.items:
        base_gross += Decimal(str(it.amount or 0))

    fee_gross = Decimal("0.00")
    if mode == "fixed":
        fee_gross = Decimal(str(cfg.transaction_fee_fixed_eur or 0))
        mode_label = "Festbetrag"
    else:  # percent
        pct = Decimal(str(cfg.transaction_fee_percent or 0))
        fee_gross = (base_gross * pct / Decimal("100.00")).quantize(Decimal("0.01"))
        mode_label = "Prozent"

    # Wenn Fee 0 -> nichts anlegen
    if fee_gross == 0:
        invoice.calculate_total()
        db.session.commit()
        return _finish("Transaktionskosten sind 0,00 € – nichts hinzugefügt.", "info")

    # Transaktionskosten sind ohne MwSt
    vat_rate = Decimal("0.00")
    net, vat = BillingService.split_gross_into_net_and_vat(gross=fee_gross, vat_rate=vat_rate)

    # load_entry_id: nimm den ersten vorhandenen Posten (damit DB-Constraints erfüllt sind)
    rep_load_entry_id = None
    for it in invoice.items:
        if getattr(it, "load_entry_id", None):
            rep_load_entry_id = it.load_entry_id
            break

    # Falls die Rechnung ausnahmsweise keine Items hat: sauber abbrechen
    if not rep_load_entry_id:
        invoice.calculate_total()
        db.session.commit()
        return _finish(
            "Keine Rechnungspositionen vorhanden – Transaktionskosten konnten nicht gesetzt werden.",
            "warning"
        )

    # Beschreibung (ohne cfg.transaction_fee_label, da dieses Feld im Model nicht existiert)
    desc = f"Transaktionskosten ({mode_label})"

    db.session.add(
        InvoiceItem(
            invoice_id=invoice.id,
            load_entry_id=rep_load_entry_id,
            amount=fee_gross,
            vat_rate=vat_rate,
            net_amount=net,
            vat_amount=vat,
            description=desc,
        )
    )

    invoice.calculate_total()
    db.session.commit()
    return _finish("Transaktionskosten aktualisiert.", "success")


# ---------------------------------------------------------
# Rechnung speichern
# - lässt eine Voransicht zu
# ---------------------------------------------------------

@bp.route("/invoice/<int:invoice_id>/save", methods=["POST"])
def invoice_save(invoice_id):
    invoice = (
        Invoice.query
        .options(
            selectinload(Invoice.items)
            .joinedload(InvoiceItem.load_entry)
            .joinedload(LoadEntry.load)
        )
        .get_or_404(invoice_id)
    )

    if invoice.stage != "draft":
        flash("Rechnung ist bereits gespeichert.", "info")
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice)))

    payment_method = (request.form.get("payment_method") or "").strip().lower()
    if payment_method not in {"", "cash", "card", "transfer", "wero", "sepa"}:
        flash("Ungültige Zahlungsart.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number_for_detail(invoice)))

    if payment_method == "sepa" and not _invoice_allows_sepa(invoice):
        flash("SEPA-Lastschrift ist für diese Person nicht zulässig.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number_for_detail(invoice)))

    prepaid_voucher_raw = request.form.get("prepaid_voucher_amount") or ""
    allow_prepaid = _invoice_allows_prepaid_voucher(invoice)
    parsed_prepaid, prepaid_error = _parse_prepaid_amount(
        prepaid_voucher_raw,
        total_amount=Decimal(str(invoice.total_amount or "0.00")),
        allow_prepaid=allow_prepaid,
    )
    if prepaid_error:
        flash(prepaid_error, "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number_for_detail(invoice)))

    invoice.payment_method = payment_method or None
    invoice.prepaid_voucher_amount = parsed_prepaid
    if payment_method == "sepa":
        _set_invoice_payment_state(invoice, INVOICE_PAYMENT_STATE_SEPA_PENDING)
    else:
        _set_invoice_payment_state(invoice, INVOICE_PAYMENT_STATE_OPEN)

    if _invoice_has_ku_jump_positions(invoice):
        invoice.is_tandem_kleinunternehmer = _parse_form_bool(
            request.form.get("invoice_is_tandem_kleinunternehmer"),
            default=bool(getattr(invoice, "is_tandem_kleinunternehmer", False)),
        )
        invoice.is_video_kleinunternehmer = _parse_form_bool(
            request.form.get("invoice_is_video_kleinunternehmer"),
            default=bool(getattr(invoice, "is_video_kleinunternehmer", False)),
        )
        BillingService.recalculate_invoice_ku_tax(invoice)

    if invoice.seq_number is None:
        # Bei Entwürfen die gleiche Anzeige-/Vorschaunummer als finale Nummer übernehmen.
        invoice.seq_number = _invoice_display_number_for_detail(invoice)

    invoice.stage = "final"

    # ✅ JETZT erst wirklich abrechnen
    for item in invoice.items:
        le = getattr(item, "load_entry", None)
        if le:
            le.billed = True
            le.billed_at = now_berlin().replace(tzinfo=None)

    db.session.commit()

    flash("Rechnung wurde gespeichert.", "success")
    return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice)))


@bp.route("/invoice/<int:invoice_id>/set_tandem_kleinunternehmer", methods=["POST"])
def invoice_set_tandem_kleinunternehmer(invoice_id):
    invoice = (
        Invoice.query
        .options(selectinload(Invoice.items).joinedload(InvoiceItem.load_entry))
        .get_or_404(invoice_id)
    )

    if invoice.is_deleted:
        flash("Stornierte Rechnungen können nicht geändert werden.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number_for_detail(invoice)))

    if not _invoice_has_ku_jump_positions(invoice):
        flash("Diese Rechnung enthält keine KU-fähigen Sprungpositionen.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number_for_detail(invoice)))

    if getattr(invoice, "email_sent_ok", False):
        flash("Nach E-Mail-Versand ist diese steuerliche Einstellung gesperrt.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number_for_detail(invoice)))

    invoice.is_tandem_kleinunternehmer = _parse_form_bool(
        request.form.get("invoice_is_tandem_kleinunternehmer"),
        default=bool(getattr(invoice, "is_tandem_kleinunternehmer", False)),
    )
    invoice.is_video_kleinunternehmer = _parse_form_bool(
        request.form.get("invoice_is_video_kleinunternehmer"),
        default=bool(getattr(invoice, "is_video_kleinunternehmer", False)),
    )
    BillingService.recalculate_invoice_ku_tax(invoice)
    db.session.commit()

    flash("Kleinunternehmer-Status wurde aktualisiert.", "success")
    return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number_for_detail(invoice)))
 
# ---------------------------------------------------------
# Rechnung per E-Mail versenden
# ---------------------------------------------------------
# ---------------------------------------------------------
# Progress-Heartbeat für lange Schritte (SCHRITT 3)
# ---------------------------------------------------------
def _start_progress_heartbeat(
    invoice_id: int,
    step_text: str,
    start_percent: int,
    max_percent: int,
    stop_event: threading.Event,
    interval_sec: float = 0.5,
):
    """
    Erhöht den Fortschritt in kleinen Schritten,
    solange stop_event NICHT gesetzt ist.
    """

    current = start_percent

    while not stop_event.is_set() and current < max_percent:
        time.sleep(interval_sec)
        current += 1
        set_progress(
            invoice_id,
            step_text,
            current,
            status="processing"
        )


def _send_invoice_email_async(invoice_id: int, is_admin: bool = False):
    """
    Asynchrone Funktion zur E-Mail-Versendung
    """

    # ✅ SCHRITT 2: Sofort sichtbarer Fortschritt
    # Muss VOR create_app() stehen, damit der Nutzer direkt Feedback sieht
    set_progress(
        invoice_id,
        "E-Mail-Versand wird vorbereitet…",
        2,
        status="processing"
    )

    app = create_app()


    # ✅ WICHTIG: In einem Thread gibt es keinen Request-Context.
    # url_for() / Templates mit url_for brauchen aber einen Request-Context.
    # -> test_request_context erzeugt ihn künstlich.
    base_url = app.config.get("BASE_URL") or "http://127.0.0.1:5000"

    with app.app_context(), app.test_request_context(base_url=base_url):
        invoice = None
        try:
            # ✅ Neue DB-Session für Thread (SQLAlchemy 2.x konform)
            invoice = db.session.get(Invoice, invoice_id)
            if not invoice:
                set_progress(invoice_id, "Fehler: Rechnung nicht gefunden", 0, "error")
                return

            # Persistenter Audit: letzter Versandversuch startet jetzt.
            invoice.email_last_attempt_at = now_berlin().replace(tzinfo=None)
            invoice.email_last_error = None
            db.session.commit()

            set_progress(invoice_id, "Rechnung wird geladen...", 5)

            person = invoice.person
            billing_config = BillingConfig.query.first()
            invoice_number = _invoice_display_number_for_detail(invoice)
            invoice_purpose = _build_invoice_payment_purpose(
                invoice,
                invoice_number=invoice_number,
            )
  

            # -------------------------------------------------
            # EPC-QR
            # -------------------------------------------------
            set_progress(invoice_id, "QR-Code wird generiert...", 10)
            epc_qr_data_uri = None
            try:
                amount = Decimal(str(invoice.total_amount or 0))
                if amount > 0 and billing_config and getattr(billing_config, "iban", None):
                    payment_context = build_payment_context(
                        invoice=invoice,
                        billing_config=billing_config,
                        invoice_number=invoice_number,
                        amount_eur=amount,
                    )
                    epc_qr_data_uri = _make_qr_data_uri(payment_context["epc_payload"])
            except Exception:
                epc_qr_data_uri = None

            # -------------------------------------------------
            # PDF
            # -------------------------------------------------
            set_progress(invoice_id, "PDF wird generiert...", 20)

            pdf_heartbeat_stop = threading.Event()
            pdf_heartbeat_thread = threading.Thread(
                target=_start_progress_heartbeat,
                args=(
                    invoice_id,
                    "PDF wird generiert...",
                    20,     # Start %
                    65,     # Max % während PDF
                    pdf_heartbeat_stop,
                ),
                daemon=True,
            )
            pdf_heartbeat_thread.start()

            try:
                # Immer frisch rendern, damit der E-Mail-Anhang exakt dem aktuellen
                # Rechnungs-Template/Layout entspricht (kein veraltetes Cache-PDF).
                pdf_bytes = BillingService.render_invoice_pdf(
                    invoice,
                    billing_config=billing_config,
                    epc_qr_data_uri=epc_qr_data_uri,
                    invoice_purpose=invoice_purpose,
                )
            finally:
                pdf_heartbeat_stop.set()
                try:
                    pdf_heartbeat_thread.join(timeout=1)
                except Exception:
                    pass

            set_progress(invoice_id, "E-Mail wird vorbereitet...", 70)

            # -------------------------------------------------
            # E-Mail Text
            # -------------------------------------------------
            cfg = billing_config
            subject = (
                (cfg.mail_subject_template if cfg else None)
                or "Deine Rechnung vom Dessauer Fallschirmsportverein"
            )
            is_manual_invoice = _is_manual_invoice(invoice)
            body_template = (
                (getattr(cfg, "mail_body_template_manual", None) if is_manual_invoice else getattr(cfg, "mail_body_template", None))
                or (MANUAL_MAIL_BODY_TEMPLATE_DEFAULT if is_manual_invoice else "")
            )
            if not is_manual_invoice:
                body_template = body_template or ""

            # Flugplatz
            try:
                airfields = {
                    item.load_entry.load.airfield.name
                    for item in invoice.items
                    if getattr(item, "load_entry", None)
                    and getattr(item.load_entry, "load", None)
                    and getattr(item.load_entry.load, "airfield", None)
                }
                if len(airfields) == 1:
                    airfield_name = next(iter(airfields))
                elif len(airfields) > 1:
                    airfield_name = "mehreren Flugplätzen"
                else:
                    airfield_name = ""
            except Exception:
                airfield_name = ""

            # Daten
            try:
                invoice_load_dates = sorted(
                    {
                        (
                            item.load_entry.load.actual_time
                            or item.load_entry.load.scheduled_time
                            or item.load_entry.load.created_at
                        ).date()
                        for item in invoice.items
                        if getattr(item, "load_entry", None)
                        and getattr(item.load_entry, "load", None)
                        and (
                            item.load_entry.load.actual_time
                            or item.load_entry.load.scheduled_time
                            or item.load_entry.load.created_at
                        )
                    }
                )
            except Exception:
                invoice_load_dates = []

            if not invoice_load_dates:
                if getattr(invoice, "service_date", None):
                    invoice_date = f"vom {invoice.service_date.strftime('%d.%m.%Y')}"
                else:
                    invoice_date = f"vom {invoice.created_at.strftime('%d.%m.%Y')}" if invoice.created_at else ""
            elif len(invoice_load_dates) == 1:
                invoice_date = f"vom {invoice_load_dates[0].strftime('%d.%m.%Y')}"
            else:
                invoice_date = (
                    f"vom {invoice_load_dates[0].strftime('%d.%m.%Y')} "
                    f"bis {invoice_load_dates[-1].strftime('%d.%m.%Y')}"
                )

            recipient_name = (
                invoice.billing_address_name.strip()
                if getattr(invoice, "billing_address_name", None)
                else ((person.full_name or "") if person else "")
            )
            manual_title = (getattr(invoice, "manual_title", "") or "").strip() or "Manuelle Positionen"
            body = body_template.format(
                first_name=recipient_name,
                last_name="",
                invoice_date=invoice_date,
                airfield_name=airfield_name or "",
                manual_title=manual_title,
            )

            # Versandzeit EINMAL festlegen (fachliche Wahrheit)
            send_time = now_berlin().replace(tzinfo=None)

            if invoice.email_sent_ok and is_admin:
                body = (
                    f"[⚠️ ERNEUTER VERSAND durch Admin am "
                    f"{send_time.strftime('%d.%m.%Y um %H:%M')}]\n\n{body}"
                )


            year = invoice.created_at.strftime("%Y") if invoice.created_at else "2026"
            filename = (
                f"Rechnung_Spruenge_{year}_Nr_{invoice_number}_{person.first_name}_{person.last_name}.pdf"
                if person else
                f"Rechnung_Spruenge_{year}_Nr_{invoice_number}.pdf"
            )


            # Zieladresse: abweichende Rechnungsanschrift/E-Mail bevorzugen
            to_email = invoice.billing_address_email or (person.email if person else None)

            set_progress(invoice_id, "E-Mail wird versendet...", 85)

            mail_heartbeat_stop = threading.Event()
            mail_heartbeat_thread = threading.Thread(
                target=_start_progress_heartbeat,
                args=(
                    invoice_id,
                    "E-Mail wird versendet...",
                    85,     # Start %
                    94,     # Max % während SMTP
                    mail_heartbeat_stop,
                ),
                daemon=True,
            )
            mail_heartbeat_thread.start()

            # Harter Gesamttimeout für den SMTP-Versand (alle Fallbacks zusammen).
            # Der Timeout schützt davor, dass der Hintergrundthread unbegrenzt hängt.
            _SMTP_TOTAL_TIMEOUT = 60  # Sekunden

            _smtp_result: dict[str, str | None] | None = None
            _smtp_exc: Exception | None = None

            def _do_smtp():
                nonlocal _smtp_result, _smtp_exc
                try:
                    _smtp_result = MailerService.send_invoice_email(
                        to_email=to_email,
                        subject=subject,
                        body=body,
                        pdf_bytes=pdf_bytes,
                        filename=filename,
                        billing_config=billing_config,
                    )
                except Exception as _e:
                    _smtp_exc = _e

            smtp_thread = threading.Thread(target=_do_smtp, daemon=True)
            try:
                smtp_thread.start()
                smtp_thread.join(timeout=_SMTP_TOTAL_TIMEOUT)

                if smtp_thread.is_alive():
                    # Thread hängt noch – Timeout ausgelöst
                    timeout_msg = (
                        f"SMTP-Versand-Timeout nach {_SMTP_TOTAL_TIMEOUT}s "
                        f"(Thread hängt, Hintergrundprozess abgebrochen)"
                    )
                    try:
                        current_app.logger.error(
                            f"E-Mail Versand Timeout (Invoice {invoice_id}): {timeout_msg}"
                        )
                    except Exception:
                        pass
                    raise TimeoutError(timeout_msg)

                if _smtp_exc is not None:
                    try:
                        current_app.logger.exception(
                            f"E-Mail Versand fehlgeschlagen (Invoice {invoice_id}): {_smtp_exc}"
                        )
                    except Exception:
                        pass
                    raise _smtp_exc

                audit_meta = _smtp_result
            finally:
                mail_heartbeat_stop.set()
                try:
                    mail_heartbeat_thread.join(timeout=1)
                except Exception:
                    pass

            # DB Update
            set_progress(invoice_id, "Datenbank wird aktualisiert...", 95)

            # ✅ WICHTIG: dieselbe Zeit verwenden wie im Admin-Hinweis
            invoice.email_last_attempt_at = send_time
            invoice.email_last_error = None
            invoice.email_last_recipient = audit_meta.get("recipient") if audit_meta else (person.email if person else None)
            invoice.email_last_message_id = audit_meta.get("message_id") if audit_meta else None
            invoice.email_sent_at = send_time
            invoice.email_sent_ok = True
            invoice.email_delivery_confirmed_at = None
            invoice.email_delivery_confirmed_by = None

            db.session.commit()

            # completed bewusst im Progress-Store belassen, damit die UI den
            # Abschluss des AKTUELLEN Versandlaufs zuverlässig sehen kann.
            mark_complete(invoice_id)

        except Exception as e:
            set_progress(invoice_id, f"Fehler: {str(e)[:100]}", 0, "error")
            if invoice is not None:
                try:
                    invoice.email_last_attempt_at = now_berlin().replace(tzinfo=None)
                    invoice.email_last_recipient = person.email if person else None
                    invoice.email_last_error = str(e)[:500]
                    db.session.commit()
                except Exception:
                    db.session.rollback()
            try:
                current_app.logger.error(f"Email send failed for invoice {invoice_id}: {e}")
            except Exception:
                pass

        finally:
            # Session aufräumen (Thread-sicher)
            try:
                db.session.remove()
            except Exception:
                pass


@bp.route("/invoice/<int:invoice_id>/send_email", methods=["POST"])
def invoice_send_email(invoice_id):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    # Harte Sperre: aus Simulationslaeufen duerfen keine E-Mails versendet werden.
    if request.headers.get("X-Manifest-Simulation") == "1":
        msg = "E-Mail-Versand ist im Simulationsmodus deaktiviert."
        flash(msg, "warning")
        if is_ajax:
            return jsonify({'success': False, 'message': msg})
        return redirect(url_for("billing.invoice_list"))

    invoice = Invoice.query.get_or_404(invoice_id)

    if invoice.is_deleted and not session.get("is_admin"):
        msg = "Stornierte Rechnungen dürfen nur von Admin per E-Mail versendet werden."
        flash(msg, "danger")
        if is_ajax:
            return jsonify({'success': False, 'message': msg})
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice)))

    # ✅ Nur gespeicherte Rechnungen dürfen versendet werden
    if invoice.stage != "final":
        msg = "Rechnung muss zuerst gespeichert werden."
        flash(msg, "warning")
        if is_ajax:
            return jsonify({'success': False, 'message': msg})
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice)))

    person = invoice.person
    if not person or not person.email:
        msg = "Für diese Person ist keine E-Mail-Adresse hinterlegt."
        flash(msg, "warning")
        if is_ajax:
            return jsonify({'success': False, 'message': msg})
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice)))

    # ✅ Prüfe, ob Rechnung bereits versendet wurde
    is_admin = session.get("is_admin", False)
    if invoice.email_sent_ok and not is_admin:
        sent_at_str = ""
        if invoice.email_sent_at:
            sent_at_str = invoice.email_sent_at.strftime("%d.%m.%Y um %H:%M")
        msg = f"Diese Rechnung wurde bereits am {sent_at_str} per E-Mail versendet."
        flash(msg, "warning")
        if is_ajax:
            return jsonify({'success': False, 'message': msg})
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice)))

    # Starte asynchronen Versand
    set_progress(invoice_id, 'Versand wird gestartet...', 1)
    is_admin = session.get("is_admin", False)
    thread = threading.Thread(
        target=_send_invoice_email_async,
        args=(invoice_id, is_admin),
        daemon=True
    )
    thread.start()

    if is_ajax:
        return jsonify({'success': True, 'message': 'E-Mail-Versand gestartet...'})
    else:
        flash("E-Mail-Versand gestartet...", "info")
        return redirect(url_for("billing.invoice_detail", invoice_id=_invoice_display_number(invoice)))

# ---------------------------------------------------------
# E-Mail Versand Status abfragen (Best Practice: DB ist Wahrheit)
# ---------------------------------------------------------
@bp.route("/invoice/<int:invoice_id>/send_email_status", methods=["GET"])
def invoice_send_email_status(invoice_id):
    """
    Gibt den aktuellen Status des E-Mail-Versands zurück.

    Best Practice:
    - DB (invoice.email_sent_ok) ist die fachliche Wahrheit.
    - Progress-Store ist nur ein kurzlebiges UI-Signal.
    - Niemals "idle" zurückgeben, wenn der Versand gestartet wurde oder fertig ist.
    """
    try:
        # ✅ Stelle sicher, dass wir keinen "stale" Zustand aus einer Session sehen
        # (bei manchen Setups kann ein Objekt gecached sein).
        inv = (
            Invoice.query
            .execution_options(populate_existing=True)
            .get(invoice_id)
        )

        # Technischer Progress (kurzlebig) hat Prioritaet,
        # damit ein erneuter Versandlauf nicht sofort durch
        # einen alten DB-"sent"-Status als completed erscheint.
        progress = get_progress(invoice_id)

        if progress is not None:
            return jsonify(progress)

        if inv and getattr(inv, "email_last_error", None):
            last_attempt = getattr(inv, "email_last_attempt_at", None)
            last_success = getattr(inv, "email_sent_at", None)
            if last_attempt and (last_success is None or last_attempt >= last_success):
                return jsonify({
                    "status": "error",
                    "step": str(inv.email_last_error),
                    "percent": 0,
                })

        if inv and inv.email_sent_ok:
            return jsonify({
                "status": "completed",
                "step": "E-Mail versendet",
                "percent": 100
            })

        # Kein Progress-Store-Eintrag, nicht versendet, kein Fehler → idle.
        # (War früher "processing" als Catch-all; das führte zu falschen
        # Spinner-Anzeigen wenn kein Job läuft.)
        return jsonify({
            "status": "idle",
            "step": "Noch nicht versendet",
            "percent": 0
        })

    except Exception as e:
        # Fail-safe: UI soll nicht hängen
        return jsonify({
            "status": "error",
            "step": f"Status-Fehler: {str(e)[:120]}",
            "percent": 0
        })


@bp.route("/email_jobs_active", methods=["GET"])
def invoice_email_jobs_active():
    """Gibt die IDs aller Rechnungen zurück, für die gerade ein Versand läuft."""
    return jsonify({"active": get_active_job_ids()})


@bp.route("/invoice/<int:invoice_id>/confirm_email_delivery", methods=["POST"])
def invoice_confirm_email_delivery(invoice_id):
    deny = _full_admin_required("billing.invoice_detail", invoice_id=invoice_id)
    if deny:
        return deny

    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash("Rechnung nicht gefunden.", "danger")
        return redirect(url_for("billing.invoice_list"))

    if not invoice.email_sent_ok or not invoice.email_sent_at:
        flash("Es gibt noch keinen bestätigbaren SMTP-Versand für diese Rechnung.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    invoice.email_delivery_confirmed_at = now_berlin().replace(tzinfo=None)
    invoice.email_delivery_confirmed_by = "Admin"
    db.session.commit()

    flash("Eingang der E-Mail wurde manuell bestätigt.", "success")
    return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

# ---------------------------------------------------------
# Zahlungsart speichern (OHNE Rechnung als bezahlt zu markieren)
# - Offene Rechnung: jeder Benutzer darf Zahlungsart setzen/ändern
# - Bereits bezahlt: nur Admin darf Bezahlart ändern
# ---------------------------------------------------------
@bp.route("/invoice/<int:invoice_id>/set_payment_method", methods=["POST"])
def invoice_set_payment_method(invoice_id):
    # ✅ SQLAlchemy‑2.x‑konform
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash("Rechnung nicht gefunden.", "danger")
        return redirect(url_for("billing.invoice_list"))

    # ✅ Entwurf-Rechnungen dürfen keine Zahlungsart bekommen
    if getattr(invoice, "stage", "final") != "final":
        flash("Rechnung muss zuerst gespeichert werden.", "warning")
        return redirect(
            url_for("billing.invoice_detail", invoice_id=invoice_id)
        )

    # Bereits bezahlt: nur Admin darf Bezahlart ändern
    if invoice.is_paid and not session.get("is_admin"):
        flash(
            "Nur Admin darf die Bezahlart bei bereits bezahlten Rechnungen ändern.",
            "danger"
        )
        return redirect(
            url_for("billing.invoice_detail", invoice_id=invoice_id)
        )

    method = (request.form.get("payment_method") or "").strip().lower()
    allowed = {"", "cash", "card", "transfer", "wero", "sepa"}
    if method not in allowed:
        flash("Ungültige Zahlungsart.", "danger")
        return redirect(
            url_for("billing.invoice_detail", invoice_id=invoice_id)
        )

    if method == "sepa" and not _invoice_allows_sepa(invoice):
        flash("SEPA-Lastschrift ist für diese Person nicht zulässig.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    prepaid_raw = request.form.get("prepaid_voucher_amount") or ""
    parsed_prepaid, prepaid_error = _parse_prepaid_amount(
        prepaid_raw,
        total_amount=Decimal(str(invoice.total_amount or "0.00")),
        allow_prepaid=_invoice_allows_prepaid_voucher(invoice),
    )
    if prepaid_error:
        flash(prepaid_error, "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    invoice.payment_method = method or None
    invoice.prepaid_voucher_amount = parsed_prepaid
    if invoice.is_paid:
        _set_invoice_payment_state(invoice, INVOICE_PAYMENT_STATE_PAID)
    elif method == "sepa":
        _set_invoice_payment_state(invoice, INVOICE_PAYMENT_STATE_SEPA_PENDING)
    else:
        _set_invoice_payment_state(invoice, INVOICE_PAYMENT_STATE_OPEN)
    db.session.commit()

    if method == "sepa" and not invoice.is_paid:
        flash("Zahlungsart gespeichert (SEPA vorgemerkt).", "success")
    else:
        flash("Zahlungsart gespeichert.", "success")
    return redirect(
        url_for("billing.invoice_detail", invoice_id=invoice_id)
    )


@bp.route("/invoice/<int:invoice_id>/mark_sepa_returned", methods=["POST"])
def invoice_mark_sepa_returned(invoice_id):
    deny = _admin_or_db_admin_required("billing.invoice_detail", invoice_id=invoice_id)
    if deny:
        return deny

    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash("Rechnung nicht gefunden.", "danger")
        return redirect(url_for("billing.invoice_list"))

    if getattr(invoice, "stage", "final") != "final":
        flash("Nur gespeicherte Rechnungen können markiert werden.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    current_state = _invoice_payment_state(invoice)
    if invoice.payment_method != "sepa" or current_state not in {
        INVOICE_PAYMENT_STATE_PAID,
        INVOICE_PAYMENT_STATE_SEPA_EXPORTED,
    }:
        flash("Diese Aktion ist nur für SEPA-Rechnungen mit Zustand 'Bezahlt' oder 'SEPA exportiert' verfügbar.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    _set_invoice_payment_state(invoice, INVOICE_PAYMENT_STATE_SEPA_RETURNED)
    db.session.commit()

    flash("Rücklastschrift erfasst.", "success")
    return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))


@bp.route("/invoice/<int:invoice_id>/mark_sepa_pending", methods=["POST"])
def invoice_mark_sepa_pending(invoice_id):
    deny = _admin_or_db_admin_required("billing.invoice_detail", invoice_id=invoice_id)
    if deny:
        return deny

    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash("Rechnung nicht gefunden.", "danger")
        return redirect(url_for("billing.invoice_list"))

    if getattr(invoice, "stage", "final") != "final":
        flash("Nur gespeicherte Rechnungen können markiert werden.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    if _invoice_payment_state(invoice) != INVOICE_PAYMENT_STATE_SEPA_RETURNED:
        flash("Diese Aktion ist nur für Rechnungen mit Rücklastschrift verfügbar.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    _set_invoice_payment_state(invoice, INVOICE_PAYMENT_STATE_SEPA_PENDING)
    db.session.commit()

    flash("Rechnung wurde erneut für SEPA vorgemerkt.", "success")
    return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))


@bp.route("/invoice/<int:invoice_id>/mark_sepa_open", methods=["POST"])
def invoice_mark_sepa_open(invoice_id):
    deny = _admin_or_db_admin_required("billing.invoice_detail", invoice_id=invoice_id)
    if deny:
        return deny

    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash("Rechnung nicht gefunden.", "danger")
        return redirect(url_for("billing.invoice_list"))

    if getattr(invoice, "stage", "final") != "final":
        flash("Nur gespeicherte Rechnungen können markiert werden.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    if _invoice_payment_state(invoice) != INVOICE_PAYMENT_STATE_SEPA_RETURNED:
        flash("Diese Aktion ist nur für Rechnungen mit Rücklastschrift verfügbar.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    _set_invoice_payment_state(invoice, INVOICE_PAYMENT_STATE_OPEN)
    db.session.commit()

    flash("Rechnung wurde auf offene Rechnung zurückgesetzt.", "success")
    return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))


@bp.route("/invoice/<int:invoice_id>/set_payment_state", methods=["POST"])
def invoice_set_payment_state(invoice_id):
    deny = _full_admin_required("billing.invoice_detail", invoice_id=invoice_id)
    if deny:
        return deny

    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash("Rechnung nicht gefunden.", "danger")
        return redirect(url_for("billing.invoice_list"))

    if getattr(invoice, "stage", "final") != "final":
        flash("Nur gespeicherte Rechnungen können einen Zahlungsstatus erhalten.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    state = (request.form.get("payment_state") or "").strip().lower()
    if state not in INVOICE_PAYMENT_STATES:
        flash("Ungültiger Zahlungsstatus.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    current_state = _invoice_payment_state(invoice)
    if not _is_allowed_payment_state_transition(current_state, state):
        flash("Unzulässiger Zahlungsstatus-Übergang.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    if state in {INVOICE_PAYMENT_STATE_SEPA_PENDING, INVOICE_PAYMENT_STATE_SEPA_EXPORTED, INVOICE_PAYMENT_STATE_SEPA_RETURNED} and not _invoice_allows_sepa(invoice):
        flash("SEPA-Status ist für diese Person nicht zulässig.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    _set_invoice_payment_state(invoice, state)
    db.session.commit()

    flash(f"Zahlungsstatus aktualisiert: {PAYMENT_STATE_LABELS.get(state, state)}", "success")
    return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))


# ---------------------------------------------------------
# Rechnung als bezahlt markieren – mit Zahlungsart
# - Offene Rechnung: jeder Benutzer darf bezahlen
# - Bereits bezahlt: nur Admin darf Bezahlart ändern
# ---------------------------------------------------------
@bp.route("/invoice/<int:invoice_id>/pay/<string:method>", methods=["POST"])
def invoice_pay(invoice_id, method):
    # ✅ SQLAlchemy‑2.x‑konform
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash("Rechnung nicht gefunden.", "danger")
        return redirect(url_for("billing.invoice_list"))

    # ✅ Entwurf-Rechnungen dürfen nicht bezahlt werden
    if getattr(invoice, "stage", "final") != "final":
        flash("Rechnung muss zuerst gespeichert werden.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    method = (method or "").strip().lower()
    allowed = {"cash", "card", "transfer", "wero"}
    if method not in allowed:
        flash("Ungültige Zahlungsart.", "danger")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    prepaid_raw = request.form.get("prepaid_voucher_amount") or ""
    parsed_prepaid, prepaid_error = _parse_prepaid_amount(
        prepaid_raw,
        total_amount=Decimal(str(invoice.total_amount or "0.00")),
        allow_prepaid=_invoice_allows_prepaid_voucher(invoice),
    )
    if prepaid_error:
        flash(prepaid_error, "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    invoice.prepaid_voucher_amount = parsed_prepaid

    # ✅ Regel bleibt: Nur Admin darf bei bereits bezahlten Rechnungen die Bezahlart ändern
    if invoice.is_paid and not session.get("is_admin"):
        flash("Nur Admin darf die Bezahlart bei bereits bezahlten Rechnungen ändern.", "danger")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    # Offene Rechnung: jeder Benutzer darf bezahlen (mark_paid)
    # Bereits bezahlt: Admin darf Zahlungsmethode ändern (ohne paid_at zu überschreiben)
    was_paid = bool(invoice.is_paid)
    ok = BillingService.mark_invoice_paid(invoice_id, payment_method=method)
    if ok:
        db.session.commit()
        if was_paid:
            # Hinweistext sauber: wenn bereits bezahlt -> war Method change
            flash("Zahlungsart aktualisiert.", "success")
        else:
            flash("Rechnung als bezahlt markiert.", "success")
    else:
        flash("Rechnung konnte nicht als bezahlt markiert werden.", "warning")

    return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))


# ---------------------------------------------------------
# Rechnung als bezahlt markieren (Legacy-Route)
# Erwartet payment_method aus Formular
# - Offene Rechnung: jeder Benutzer darf bezahlen
# - Bereits bezahlt: nur Admin darf Bezahlart ändern
# ---------------------------------------------------------
@bp.route("/invoice/<int:invoice_id>/mark_paid", methods=["POST"])
def invoice_mark_paid(invoice_id):
    # ✅ SQLAlchemy‑2.x‑konform
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash("Rechnung nicht gefunden.", "danger")
        return redirect(url_for("billing.invoice_list"))

    # ✅ Entwurf-Rechnungen dürfen nicht bezahlt werden
    if getattr(invoice, "stage", "final") != "final":
        flash("Rechnung muss zuerst gespeichert werden.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    payment_method = (request.form.get("payment_method") or "").strip().lower()
    if payment_method not in {"cash", "card", "transfer", "wero"}:
        flash("Bitte Zahlungsart auswählen.", "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    prepaid_raw = request.form.get("prepaid_voucher_amount") or ""
    parsed_prepaid, prepaid_error = _parse_prepaid_amount(
        prepaid_raw,
        total_amount=Decimal(str(invoice.total_amount or "0.00")),
        allow_prepaid=_invoice_allows_prepaid_voucher(invoice),
    )
    if prepaid_error:
        flash(prepaid_error, "warning")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    invoice.prepaid_voucher_amount = parsed_prepaid

    # ✅ Regel bleibt: Nur Admin darf bei bereits bezahlten Rechnungen
    # die Bezahlart ändern
    if invoice.is_paid and not session.get("is_admin"):
        flash(
            "Nur Admin darf die Bezahlart bei bereits bezahlten Rechnungen ändern.",
            "danger"
        )
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    was_paid = bool(invoice.is_paid)
    ok = BillingService.mark_invoice_paid(
        invoice_id,
        payment_method=payment_method
    )

    if ok:
        db.session.commit()
        if was_paid:
            flash("Zahlungsart aktualisiert.", "success")
        else:
            flash("Rechnung als bezahlt markiert.", "success")
    else:
        flash("Rechnung konnte nicht als bezahlt markiert werden.", "warning")

    return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))



# --------------------------------------------------------------------------------------------
# Admin: Zahlung zurücksetzen (Rechnung wieder OFFEN)
# - Nur Admin
# - billed bleibt erhalten (Rechnung bleibt fakturiert, aber nicht bezahlt)
# - paid Flags in LoadEntries werden zurückgesetzt
# --------------------------------------------------------------------------------------------
def _unpay_invoice_entries(invoice: Invoice) -> None:
    """
    Setzt NUR die 'paid' Flags der LoadEntries zurück.
    billed bleibt unverändert (Rechnung bleibt abgerechnet, nur Zahlung wird zurückgenommen).
    """
    for item in invoice.items:
        le = getattr(item, "load_entry", None)
        if not le:
            continue
        if hasattr(le, "paid"):
            le.paid = False
        if hasattr(le, "paid_at"):
            le.paid_at = None


@bp.route("/invoice/<int:invoice_id>/unpay", methods=["POST"])
def invoice_unpay(invoice_id):
    deny = _full_admin_required("billing.invoice_detail", invoice_id=invoice_id)
    if deny:
        return deny

    invoice = Invoice.query.get_or_404(invoice_id)

    if not invoice.is_paid:
        flash("Rechnung ist bereits offen.", "info")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    # Zahlung aufheben (Invoice + Entries)
    _set_invoice_payment_state(invoice, INVOICE_PAYMENT_STATE_OPEN)
    _unpay_invoice_entries(invoice)
    db.session.commit()

    flash("Zahlung zurückgesetzt – Rechnung ist wieder OFFEN.", "warning")
    return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))


# --------------------------------------------------------------------------------------------
# Admin: Rechnung "freigeben" (gemeinsame Logik)
# (bestehende Funktion bleibt unverändert)
# --------------------------------------------------------------------------------------------
def _release_invoice_entries(invoice: Invoice) -> None:
    for item in invoice.items:
        le = getattr(item, "load_entry", None)
        if not le:
            continue
        if hasattr(le, "billed"):
            le.billed = False
        if hasattr(le, "paid"):
            le.paid = False
        if hasattr(le, "paid_at"):
            le.paid_at = None

# --------------------------------------------------------------------------------------------
# Admin: Soft-Delete (Storno)
# --------------------------------------------------------------------------------------------
@bp.route("/invoice/<int:invoice_id>/soft_delete", methods=["POST"])
def invoice_soft_delete(invoice_id):
    deny = _full_admin_required("billing.invoice_detail", invoice_id=invoice_id)
    if deny:
        return deny

    invoice = Invoice.query.get_or_404(invoice_id)
    invoice.is_deleted = True
    invoice.deleted_at = now_local().replace(tzinfo=None)
    invoice.deleted_reason = "Admin: Soft-Delete"
    invoice.deleted_by = "admin"

    _release_invoice_entries(invoice)
    db.session.commit()

    flash("Rechnung storniert (Soft-Delete). Abrechnung ist wieder offen.", "success")
    return redirect(url_for("billing.invoice_list"))


# --------------------------------------------------------------------------------------------
# Admin: Hard-Delete (physisch löschen) – DEV-only serverseitig gesperrt
# --------------------------------------------------------------------------------------------
@bp.route("/invoice/<int:invoice_id>/hard_delete", methods=["POST"])
def invoice_hard_delete(invoice_id):
    deny = _full_admin_required("billing.invoice_detail", invoice_id=invoice_id)
    if deny:
        return deny

    if not is_dev_mode():
        flash("Hard-Delete ist im Produktivbetrieb nicht erlaubt.", "danger")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    invoice = Invoice.query.get_or_404(invoice_id)
    _release_invoice_entries(invoice)

    db.session.delete(invoice)
    db.session.commit()

    flash("Rechnung wurde endgültig gelöscht (Hard-Delete, DEV).", "warning")
    return redirect(url_for("billing.invoice_list"))


# --------------------------------------------------------------------------------------------
# Admin (DEV only): Rechnung neu berechnen
# --------------------------------------------------------------------------------------------
@bp.route("/invoice/<int:invoice_id>/recalculate", methods=["POST"])
def invoice_recalculate(invoice_id):
    deny = _full_admin_required("billing.invoice_detail", invoice_id=invoice_id)
    if deny:
        return deny

    if not is_dev_mode():
        flash("Neu berechnen ist nur im DEV-Modus erlaubt.", "danger")
        return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))

    ok = BillingService.recalculate_invoice(invoice_id)
    if ok:
        flash("Rechnung wurde neu berechnet (DEV).", "success")
    else:
        flash("Rechnung konnte nicht neu berechnet werden.", "warning")

    return redirect(url_for("billing.invoice_detail", invoice_id=invoice_id))


# ---------------------------------------------------------
# Admin: Rechnungssteller-Konfiguration bearbeiten
# ---------------------------------------------------------
@bp.route("/admin/config", methods=["GET", "POST"])
def admin_config_edit():
    deny = _full_admin_required("billing.overview")
    if deny:
        return deny

    cfg = BillingConfig.query.first()
    if not cfg:
        cfg = BillingConfig(
            company_name="",
            street="",
            zip_code="",
            city="",
            country="Deutschland",
        )
        db.session.add(cfg)
        db.session.commit()

    defaults_changed = False
    if not cfg.waiver_text_skydiver:
        cfg.waiver_text_skydiver = WAIVER_TEXT_SKYDIVER_DEFAULT
        defaults_changed = True
    if not cfg.waiver_text_tandem:
        cfg.waiver_text_tandem = WAIVER_TEXT_TANDEM_DEFAULT
        defaults_changed = True
    if defaults_changed:
        db.session.commit()

    if request.method == "POST":
        _apply_billing_config_form(cfg, request.form)

        db.session.commit()
        flash("Rechnungssteller-Konfiguration gespeichert.", "success")
        return redirect(url_for("billing.admin_config_edit"))

    return render_template("billing/admin/config_edit.html", cfg=cfg, smtp_test_to_email="")


def _apply_billing_config_form(cfg: BillingConfig, form) -> None:
    cfg.company_name = form.get("company_name", "").strip()
    cfg.street = form.get("street", "").strip()
    cfg.zip_code = form.get("zip_code", "").strip()
    cfg.city = form.get("city", "").strip()
    cfg.country = form.get("country", "Deutschland").strip()

    cfg.email = form.get("email", "").strip() or None
    cfg.phone = form.get("phone", "").strip() or None
    cfg.website = form.get("website", "").strip() or None
    cfg.tax_number = form.get("tax_number")

    cfg.mail_sender_address = form.get("mail_sender_address") or None
    cfg.mail_sender_name = form.get("mail_sender_name") or None
    cfg.mail_subject_template = form.get("mail_subject_template") or None
    cfg.mail_body_template = form.get("mail_body_template") or None
    cfg.mail_body_template_manual = form.get("mail_body_template_manual") or None
    cfg.waiver_text_skydiver = form.get("waiver_text_skydiver") or WAIVER_TEXT_SKYDIVER_DEFAULT
    cfg.waiver_text_tandem = form.get("waiver_text_tandem") or WAIVER_TEXT_TANDEM_DEFAULT

    cfg.smtp_server = form.get("smtp_server") or None
    cfg.smtp_fallback_host = form.get("smtp_fallback_host") or None
    raw_port = form.get("smtp_port")
    cfg.smtp_port = int(raw_port) if raw_port else None
    cfg.smtp_username = form.get("smtp_username") or None

    smtp_pwd = form.get("smtp_password")
    if smtp_pwd:
        cfg.smtp_password = smtp_pwd

    cfg.smtp_use_tls = bool(form.get("smtp_use_tls"))
    cfg.smtp_use_ssl = bool(form.get("smtp_use_ssl"))

    cfg.logo_filename = form.get("logo_filename", "").strip() or None
    cfg.payment_methods_text = form.get("payment_methods_text", "").strip() or None
    cfg.bank_name = form.get("bank_name", "").strip() or None
    cfg.iban = form.get("iban", "").strip() or None
    cfg.bic = form.get("bic", "").strip() or None
    cfg.creditor_id = (form.get("creditor_id", "") or "").strip()
    cfg.pain_version = (form.get("pain_version", "") or "pain.008.001.02").strip() or "pain.008.001.02"
    cfg.payment_terms = form.get("payment_terms", "").strip() or None
    cfg.transaction_fee_mode = form.get("transaction_fee_mode", "none").strip() or "none"

    try:
        raw = form.get("transaction_fee_fixed_eur", "0").replace(",", ".").strip()
        cfg.transaction_fee_fixed_eur = Decimal(raw) if raw else Decimal("0")
    except Exception:
        cfg.transaction_fee_fixed_eur = Decimal("0")

    try:
        raw = form.get("transaction_fee_percent", "0").replace(",", ".").strip()
        cfg.transaction_fee_percent = Decimal(raw) if raw else Decimal("0")
    except Exception:
        cfg.transaction_fee_percent = Decimal("0")

    cfg.instagram_url = form.get("instagram_url", "").strip() or None
    cfg.facebook_url = form.get("facebook_url", "").strip() or None
    cfg.qr_instagram_filename = form.get("qr_instagram_filename", "").strip() or None
    cfg.qr_facebook_filename = form.get("qr_facebook_filename", "").strip() or None
    cfg.qr_website_filename = form.get("qr_website_filename", "").strip() or None


def _build_config_from_posted_form() -> BillingConfig:
    cfg = BillingConfig.query.first()
    if not cfg:
        cfg = BillingConfig(country="Deutschland")
    _apply_billing_config_form(cfg, request.form)
    return cfg


@bp.route("/admin/config/test_email", methods=["POST"])
def admin_config_test_email():
    deny = _full_admin_required("billing.overview")
    if deny:
        return deny

    cfg = _build_config_from_posted_form()
    test_to_email = (request.form.get("smtp_test_to_email") or "").strip()
    if not test_to_email:
        flash("Bitte eine Test-E-Mail-Adresse angeben.", "warning")
        return render_template("billing/admin/config_edit.html", cfg=cfg, smtp_test_to_email=test_to_email)

    try:
        meta = MailerService.send_invoice_email(
            to_email=test_to_email,
            subject="SMTP-Test aus Manifest",
            body=(
                "Dies ist eine Testmail aus der SMTP-Konfiguration von Manifest.\n\n"
                "Wenn diese Nachricht ankommt, funktionieren Server, Login, TLS und Absender."
            ),
            billing_config=cfg,
        )
        flash(
            f"Testmail an {meta.get('recipient') or test_to_email} akzeptiert. Message-ID: {meta.get('message_id') or 'n/a'}",
            "success",
        )
    except Exception as exc:
        flash(f"SMTP-Test fehlgeschlagen: {exc}", "danger")

    return render_template("billing/admin/config_edit.html", cfg=cfg, smtp_test_to_email=test_to_email)


def _load_recent_sepa_exports(limit: int = 20) -> list[SepaExport]:
    try:
        return (
            SepaExport.query
            .options(selectinload(SepaExport.invoices))
            .order_by(SepaExport.created_at.desc(), SepaExport.id.desc())
            .limit(limit)
            .all()
        )
    except Exception:
        return []


@bp.route("/invoices/sepa/export", methods=["POST"])
def invoice_sepa_export():
    deny = _admin_or_db_admin_required("billing.invoice_list")
    if deny:
        return deny

    raw_ids = request.form.getlist("invoice_ids")
    invoice_ids: list[int] = []
    seen: set[int] = set()
    for raw in raw_ids:
        try:
            val = int(str(raw).strip())
        except Exception:
            continue
        if val > 0 and val not in seen:
            seen.add(val)
            invoice_ids.append(val)

    if not invoice_ids:
        flash("Bitte mindestens eine SEPA-vorgemerkte Rechnung auswählen.", "warning")
        return redirect(request.referrer or url_for("billing.invoice_list"))

    invoices = (
        Invoice.query
        .options(
            joinedload(Invoice.person),
            selectinload(Invoice.items)
            .joinedload(InvoiceItem.load_entry)
            .joinedload(LoadEntry.load),
        )
        .filter(
            Invoice.id.in_(invoice_ids),
            Invoice.stage == "final",
            Invoice.is_deleted.is_(False),
        )
        .all()
    )

    valid_invoices: list[Invoice] = []
    for inv in invoices:
        if _invoice_payment_state(inv) == INVOICE_PAYMENT_STATE_SEPA_PENDING:
            valid_invoices.append(inv)

    skipped_count = max(0, len(invoice_ids) - len(valid_invoices))
    if not valid_invoices:
        flash("Keine gültigen SEPA-vorgemerkten Rechnungen für den Export gefunden.", "warning")
        return redirect(request.referrer or url_for("billing.invoice_list"))

    actor = _current_admin_actor_label()
    created_at = now_berlin().replace(tzinfo=None)
    storage_dir = _sepa_export_storage_dir()
    billing_config = BillingConfig.query.first()
    creditor_id = ((getattr(billing_config, "creditor_id", None) or "").strip() or "")
    if not creditor_id:
        flash("SEPA-Export abgebrochen: Bitte hinterlegen Sie unter Rechnungssteller → Konfiguration eine Gläubiger-ID (Creditor-ID).", "warning")
        return redirect(request.referrer or url_for("billing.invoice_list"))

    file_path_written = None
    for _attempt in range(1, 6):
        try:
            export_code, export_seq_no = _next_sepa_export_code()
            file_name = _build_export_file_name(created_at, export_seq_no)

            file_path = os.path.join(storage_dir, file_name)
            if os.path.exists(file_path):
                suffix = 1
                while os.path.exists(file_path):
                    file_name = _build_export_file_name(created_at, export_seq_no).replace(".xml", f"_{suffix}.xml")
                    file_path = os.path.join(storage_dir, file_name)
                    suffix += 1

            collection_date = (created_at + timedelta(days=3)).date()
            message_id = export_code
            payment_information_id = f"PmtInf-{export_code}"
            export = SepaExport(
                export_code=export_code,
                created_at=created_at,
                created_by=actor,
                file_name=file_name,
                file_path=file_path,
                status="created",
                xml_version="pain.008.001.02",
                selection_scope="manual",
                message_id=message_id,
                payment_information_id=payment_information_id,
                collection_date=collection_date,
                control_sum=Decimal("0.00"),
                transaction_count=0,
            )
            db.session.add(export)
            db.session.flush()

            snapshot_rows: list[dict] = []
            total_amount = Decimal("0.00")

            for inv in valid_invoices:
                person = getattr(inv, "person", None)
                load_date_from, load_date_to, load_dates_text = _invoice_load_snapshot(inv)
                invoice_number = str(_invoice_display_number_for_detail(inv))
                invoice_amount = Decimal(str(inv.total_amount or "0.00"))
                payment_state_code = _invoice_payment_state(inv)

                payment_context = build_payment_context(
                    invoice=inv,
                    billing_config=billing_config,
                    invoice_number=invoice_number,
                    amount_eur=invoice_amount,
                )
                remittance_info = payment_context["remittance_information"]
                end_to_end_id = f"ETO-{inv.id}"
                sequence_type = "FRST"
                if getattr(person, "sepa_first_collection_done", False):
                    sequence_type = "RCUR"

                sepa_link = SepaExportInvoice(
                    export_id=export.id,
                    invoice_id=inv.id,
                    invoice_number_snapshot=invoice_number,
                    invoice_total_snapshot=invoice_amount,
                    person_name_snapshot=((person.full_name if person else "") or "").strip(),
                    iban_snapshot=((getattr(person, "iban", "") or "").replace(" ", "").strip() or None),
                    mandate_reference_snapshot=((getattr(person, "sepa_mandate_reference", "") or "").strip() or None),
                    payment_method_snapshot=((inv.payment_method or "").strip() or None),
                    payment_state_snapshot=payment_state_code,
                    end_to_end_id_snapshot=end_to_end_id,
                    sequence_type_snapshot=sequence_type,
                    remittance_information_snapshot=remittance_info,
                    load_date_from=load_date_from,
                    load_date_to=load_date_to,
                    load_dates_text=load_dates_text,
                )
                db.session.add(sepa_link)

                snapshot_rows.append({
                    "invoice_id": inv.id,
                    "invoice_number": invoice_number,
                    "amount": invoice_amount,
                    "payment_method": inv.payment_method or "",
                    "payment_state": payment_state_code,
                    "person_name": ((person.full_name if person else "") or "").strip(),
                    "iban": (getattr(person, "iban", "") or "").replace(" ", "").strip(),
                    "bic": (getattr(person, "bic", "") or "").replace(" ", "").strip(),
                    "mandate_reference": (getattr(person, "sepa_mandate_reference", "") or "").strip(),
                    "mandate_date": getattr(person, "sepa_mandate_date", None),
                    "sequence_type": "FRST" if not getattr(person, "sepa_first_collection_done", False) else "RCUR",
                    "remittance_information": remittance_info,
                    "end_to_end_id": f"ETO-{inv.id}",
                    "debtor_name": ((person.full_name if person else "") or "").strip(),
                    "debtor_country": "DE",
                    "load_date_from": load_date_from,
                    "load_date_to": load_date_to,
                    "load_dates_text": load_dates_text,
                })
                total_amount += invoice_amount

            export.invoice_count = len(valid_invoices)
            export.total_amount = total_amount
            export.control_sum = total_amount
            export.transaction_count = len(valid_invoices)

            xml_bytes = _build_sepa_export_placeholder_xml(export_code, created_at, snapshot_rows)

            with open(file_path, "wb") as f:
                f.write(xml_bytes)
            file_path_written = file_path

            for inv in valid_invoices:
                _set_invoice_payment_state(inv, INVOICE_PAYMENT_STATE_SEPA_EXPORTED)

            db.session.commit()

            msg = f"SEPA-Export {export_code} erstellt ({len(valid_invoices)} Rechnung(en))."
            if skipped_count:
                msg += f" {skipped_count} Auswahl(en) waren nicht mehr exportierbar und wurden übersprungen."
            flash(msg, "success")

            response = _build_invoice_list_redirect_response(export_id=export.id)
            return response

        except IntegrityError:
            db.session.rollback()
            continue
        except Exception as exc:
            db.session.rollback()
            if file_path_written and os.path.exists(file_path_written):
                try:
                    os.remove(file_path_written)
                except Exception:
                    pass
            flash(f"SEPA-Export fehlgeschlagen: {exc}", "danger")
            return redirect(request.referrer or url_for("billing.invoice_list"))

    flash("SEPA-Export konnte wegen kollidierender Exportnummern nicht abgeschlossen werden.", "danger")
    return redirect(request.referrer or url_for("billing.invoice_list"))


@bp.route("/invoices/sepa/export/<int:export_id>/download", methods=["GET"])
def invoice_sepa_export_download(export_id):
    deny = _admin_or_db_admin_required("billing.invoice_list")
    if deny:
        return deny

    export = SepaExport.query.get_or_404(export_id)
    file_path = (export.file_path or "").strip()
    if not file_path or not os.path.exists(file_path):
        flash("Exportdatei wurde nicht gefunden.", "warning")
        return redirect(request.referrer or url_for("billing.invoice_list"))

    with open(file_path, "rb") as f:
        xml_bytes = f.read()

    response = make_response(xml_bytes)
    response.headers["Content-Type"] = "application/xml; charset=utf-8"
    response.headers["Content-Disposition"] = f'attachment; filename="{export.file_name}"'
    return response


@bp.route("/invoices/sepa/export/<int:export_id>/rollback", methods=["POST"])
def invoice_sepa_export_rollback(export_id):
    if not is_dev_mode():
        flash("SEPA-Testexports können nur im Dev-Modus zurückgerollt werden.", "warning")
        return redirect(url_for("billing.invoice_list"))

    deny = _admin_or_db_admin_required("billing.invoice_list")
    if deny:
        return deny

    export = SepaExport.query.get_or_404(export_id)

    latest_export = (
        SepaExport.query
        .order_by(SepaExport.created_at.desc(), SepaExport.id.desc())
        .first()
    )
    if not latest_export or latest_export.id != export.id:
        flash("Nur der aktuell letzte SEPA-Export kann zurückgerollt werden.", "warning")
        return redirect(url_for("billing.invoice_list"))

    invoice_ids = [link.invoice_id for link in (getattr(export, "invoices", []) or []) if getattr(link, "invoice_id", None)]
    file_path = (export.file_path or "").strip()

    try:
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception:
                pass

        if invoice_ids:
            invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).all()
        else:
            invoices = []

        SepaExportInvoice.query.filter(SepaExportInvoice.export_id == export.id).delete(synchronize_session=False)
        db.session.delete(export)

        for invoice in invoices:
            _reset_invoice_after_sepa_rollback(invoice)

        db.session.commit()
        flash("SEPA-Testexport wurde vollständig zurückgerollt.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"SEPA-Testexport konnte nicht zurückgerollt werden: {exc}", "danger")

    return _build_invoice_list_redirect_response()


# ---------------------------------------------------------
# Liste aller Rechnungen inkl. Zeitfilter + Delta  (SQL-optimiert)
# ---------------------------------------------------------
@bp.route("/invoices")
def invoice_list():
    period = request.args.get("period", "all")  # all / today / week / month / year / range
    download_export_id = (request.args.get("sepa_download_export_id") or "").strip()
    from_str = (request.args.get("from") or "").strip()
    to_str = (request.args.get("to") or "").strip()
    invoice_from_str = (request.args.get("invoice_from") or "").strip()
    invoice_to_str = (request.args.get("invoice_to") or "").strip()
    filters = _parse_invoice_list_filters(request.args)
    delta_scope = _parse_invoice_delta_scope(request.args)

    ctx = _build_invoice_list_context(
        period=period,
        from_str=from_str,
        to_str=to_str,
        invoice_from_str=invoice_from_str,
        invoice_to_str=invoice_to_str,
        filters=filters,
        delta_scope=delta_scope,
    )

    if download_export_id:
        try:
            export_id = int(download_export_id)
        except Exception:
            export_id = None
        if export_id is not None:
            export = SepaExport.query.get(export_id)
            if export and export.file_path and os.path.exists(export.file_path):
                with open(export.file_path, "rb") as fh:
                    xml_bytes = fh.read()
                response = make_response(render_template("billing/invoice_list.html", **ctx))
                _set_no_store_headers(response)
                response.headers["X-SEPA-Download-Export-Id"] = str(export_id)
                response.headers["X-SEPA-Download-File-Name"] = export.file_name or "export.xml"
                response.headers["X-SEPA-Download-Content-Type"] = "application/xml; charset=utf-8"
                response.headers["X-SEPA-Download-Bytes"] = str(len(xml_bytes))
                return response

    response = make_response(render_template("billing/invoice_list.html", **ctx))
    _set_no_store_headers(response)
    return response


def _build_invoice_list_context(
    period: str,
    from_str: str = "",
    to_str: str = "",
    invoice_from_str: str = "",
    invoice_to_str: str = "",
    filters: dict | None = None,
    delta_scope: str = "visible",
):
    period, load_start_dt, load_end_dt = _resolve_load_date_range(period, from_str, to_str)
    _, invoice_start_dt, invoice_end_dt = _resolve_named_date_range(
        invoice_from_str,
        invoice_to_str,
    )
    load_dt_expr = func.coalesce(Load.actual_time, Load.scheduled_time, Load.created_at)

    q = Invoice.query.filter(
        Invoice.is_deleted.is_(False),
        Invoice.stage == "final"
    )

    if invoice_start_dt is not None:
        q = q.filter(Invoice.created_at.isnot(None), Invoice.created_at >= invoice_start_dt)
    if invoice_end_dt is not None:
        q = q.filter(Invoice.created_at.isnot(None), Invoice.created_at < invoice_end_dt)

    if load_start_dt is not None or load_end_dt is not None:
        load_items_in_range = (
            db.session.query(InvoiceItem.id)
            .join(LoadEntry, LoadEntry.id == InvoiceItem.load_entry_id)
            .join(Load, Load.id == LoadEntry.load_id)
            .filter(InvoiceItem.invoice_id == Invoice.id)
        )
        if load_start_dt is not None:
            load_items_in_range = load_items_in_range.filter(load_dt_expr >= load_start_dt)
        if load_end_dt is not None:
            load_items_in_range = load_items_in_range.filter(load_dt_expr < load_end_dt)

        manual_items_exist = (
            db.session.query(InvoiceItem.id)
            .filter(
                InvoiceItem.invoice_id == Invoice.id,
                func.lower(func.coalesce(InvoiceItem.item_source, "")) == "manual",
            )
            .exists()
        )

        service_date_matches = [
            Invoice.service_date.isnot(None),
            manual_items_exist,
        ]
        if load_start_dt is not None:
            service_date_matches.append(Invoice.service_date >= load_start_dt.date())
        if load_end_dt is not None:
            service_date_matches.append(Invoice.service_date < load_end_dt.date())

        q = q.filter(
            or_(
                load_items_in_range.exists(),
                and_(*service_date_matches),
            )
        )

    invoices_all = q.order_by(Invoice.payment_state.asc(), Invoice.created_at.desc()).all()

    invoices = invoices_all

    if filters:
        invoices = [inv for inv in invoices if _invoice_matches_filters(inv, filters)]
        invoices = _sort_invoices_for_list(invoices, filters.get("sort") or "date_desc")

    invoices_for_delta = invoices if delta_scope == "visible" else invoices_all

    # -------- Neu: Summen basierend auf final invoices (nicht LoadEntries) --------
    # sum_billable = Summe ALLER final invoices (unabhängig von paid/open)
    # sum_open_invoices = Summe der final invoices mit is_paid=False
    # Delta = sum_billable - sum_open_invoices (Differenz zu bezahlten Invoices)
    
    sum_billable = sum(
        (inv.total_amount or Decimal("0.00"))
        for inv in invoices_for_delta
    )
    sum_open_invoices = sum(
        _invoice_open_amount_for_kpi(inv)
        for inv in invoices_for_delta
    )
    filtered_invoice_subtotal = sum(
        (inv.total_amount or Decimal("0.00"))
        for inv in invoices
    )

    billable_rows = []
    le_q = (
        LoadEntry.query
        .options(
            joinedload(LoadEntry.load),
            joinedload(LoadEntry.person),
            joinedload(LoadEntry.status_definition),
        )
        .join(Load, LoadEntry.load_id == Load.id)
        .filter(LoadEntry.billed.is_(False))
        .filter(Load.status == "completed")
    )

    if load_start_dt is not None:
        le_q = le_q.filter(load_dt_expr >= load_start_dt)
    if load_end_dt is not None:
        le_q = le_q.filter(load_dt_expr < load_end_dt)

    open_entries_in_period = le_q.order_by(
        LoadEntry.person_id.asc(),
        LoadEntry.created_at.asc()
    ).all()

    by_person = defaultdict(list)
    for e in open_entries_in_period:
        if not e or not getattr(e, "person_id", None):
            continue
        by_person[e.person_id].append(e)

    for person_id, entries_in_period in by_person.items():
        if not entries_in_period:
            continue

        person = entries_in_period[0].person
        if not person:
            continue

        # --- Sprungpreise ---
        total_jump = Decimal("0.00")
        for e in entries_in_period:
            try:
                total_jump += Decimal(
                    str(BillingService.calculate_price_for_entry(e) or "0.00")
                )
            except Exception:
                pass

        extras_preview = BillingService.compute_extras_for_entries(
            entries_in_period,
            include_rental_items=False,
            include_orga_items=False,
        )
        rent_total = Decimal(str(extras_preview.get("rental_sum_gross") or "0.00"))
        orga_total = Decimal(str(extras_preview.get("orga_sum_gross") or "0.00"))

        amount = total_jump + rent_total + orga_total
        if amount != 0:
            content_status_codes = _billable_person_content_status_codes(person, entries_in_period)
            content_status_labels = [_invoice_content_status_label(code) for code in sorted(content_status_codes)]
            billable_rows.append({
                "person": person,
                "amount": amount,
                "count": len(entries_in_period),
                "content_status_codes_csv": ",".join(sorted(content_status_codes)),
                "content_status_sort": " | ".join(content_status_labels).casefold(),
                "content_status_labels": content_status_labels,
            })

    if delta_scope == "visible" and filters:
        billable_rows = [row for row in billable_rows if _billable_row_matches_filters(row, filters)]

    sum_billable_uninvoiced = sum(
        Decimal(str(row.get("amount") or "0.00"))
        for row in billable_rows
    )

    sum_paid_amount = sum(
        _invoice_paid_amount_for_kpi(inv)
        for inv in invoices_for_delta
    )

    delta = sum_billable - sum_open_invoices
    sum_total = sum_paid_amount + sum_open_invoices + sum_billable_uninvoiced
    active_filter_labels = _build_invoice_filter_labels(filters)
    delta_scope_label = (
        "aktuelle Filteransicht"
        if delta_scope == "visible"
        else "gesamter Zeitraum"
    )

    sepa_pending_count = sum(1 for inv in invoices if _invoice_payment_state(inv) == INVOICE_PAYMENT_STATE_SEPA_PENDING)
    sepa_exported_count = sum(1 for inv in invoices if _invoice_payment_state(inv) == INVOICE_PAYMENT_STATE_SEPA_EXPORTED)
    sepa_exportable_count = sepa_pending_count
    sepa_exports = _load_recent_sepa_exports(limit=30)

    return {
        "invoices": invoices,
        "invoice_display_number": _invoice_display_number_for_detail,
        "period": period,
        "filters": filters or {},
        "active_filter_labels": active_filter_labels,
        "delta_scope": delta_scope,
        "delta_scope_label": delta_scope_label,
        "sum_open_invoices": sum_open_invoices,
        "sum_billable": sum_billable,
        "sum_billable_uninvoiced": sum_billable_uninvoiced,
        "sum_paid_amount": sum_paid_amount,
        "delta": delta,
        "sum_paid": sum_paid_amount,
        "sum_total": sum_total,
        "filtered_invoice_subtotal": filtered_invoice_subtotal,
        "billable_rows": billable_rows,
        "from_date": from_str,
        "to_date": to_str,
        "invoice_from_date": invoice_from_str,
        "invoice_to_date": invoice_to_str,
        "sepa_pending_count": sepa_pending_count,
        "sepa_exported_count": sepa_exported_count,
        "sepa_exportable_count": sepa_exportable_count,
        "sepa_exports": sepa_exports,
        "can_execute_sepa_export": _can_manage_sepa_exports(),
        "is_dev_mode": is_dev_mode(),
    }


@bp.route("/invoices/pdf")
def invoice_list_pdf():
    period = request.args.get("period", "all")
    from_str = (request.args.get("from") or request.args.get("from_date") or "").strip()
    to_str = (request.args.get("to") or request.args.get("to_date") or "").strip()
    invoice_from_str = (request.args.get("invoice_from") or "").strip()
    invoice_to_str = (request.args.get("invoice_to") or "").strip()
    filters = _parse_invoice_list_filters(request.args)
    delta_scope = _parse_invoice_delta_scope(request.args)

    ctx = _build_invoice_list_context(
        period=period,
        from_str=from_str,
        to_str=to_str,
        invoice_from_str=invoice_from_str,
        invoice_to_str=invoice_to_str,
        filters=filters,
        delta_scope=delta_scope,
    )

    billing_config = BillingConfig.query.first()

    static_dir = os.path.join(current_app.root_path, "static")
    logo_data_uri = None
    if billing_config and billing_config.logo_filename:
        logo_path = os.path.join(static_dir, "img", billing_config.logo_filename)
        logo_data_uri = _image_to_data_uri(logo_path)

    generated_at_local = now_berlin().replace(tzinfo=None)

    html = render_template(
        "billing/invoice_list_pdf.html",
        billing_config=billing_config,
        logo_data_uri=logo_data_uri,
        generated_at_local=generated_at_local,
        **ctx,
    )

    pdf_bytes, pdf_error = generate_pdf_from_html(html, base_dir=current_app.root_path)
    if pdf_error:
        flash(pdf_error, "danger")
        return redirect(url_for("billing.invoice_list", **request.args.to_dict(flat=True)))

    period_slug_map = {
        "all": "alle",
        "today": "heute",
        "week": "woche",
        "month": "monat",
        "year": "jahr",
        "range": "zeitraum",
    }
    period_slug = period_slug_map.get(period, "alle")

    if period == "range" and (from_str or to_str):
        from_part = from_str.replace("-", "_") if from_str else "offen"
        to_part = to_str.replace("-", "_") if to_str else "offen"
        period_part = f"{period_slug}_{from_part}-{to_part}"
    else:
        period_part = period_slug

    filename = (
        f"Abrechnung_Rechnungsuebersicht_{period_part}_"
        f"{generated_at_local.strftime('%Y_%m_%d_%H_%M')}.pdf"
    )

    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------
# CSV-Export Rechnungsliste
# ---------------------------------------------------------
@bp.route("/invoices/export/csv")
def invoice_list_csv():
    period = request.args.get("period", "all")
    from_str = (request.args.get("from") or request.args.get("from_date") or "").strip()
    to_str = (request.args.get("to") or request.args.get("to_date") or "").strip()
    invoice_from_str = (request.args.get("invoice_from") or "").strip()
    invoice_to_str = (request.args.get("invoice_to") or "").strip()
    filters = _parse_invoice_list_filters(request.args)
    delta_scope = _parse_invoice_delta_scope(request.args)

    ctx = _build_invoice_list_context(
        period=period, from_str=from_str, to_str=to_str,
        invoice_from_str=invoice_from_str, invoice_to_str=invoice_to_str,
        filters=filters, delta_scope=delta_scope,
    )

    def _fmt(val):
        return "{:.2f}".format(val).replace(".", ",")

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)

    writer.writerow(["Nr.", "Datum", "Person", "Netto (EUR)", "MwSt (EUR)", "Brutto (EUR)", "Status", "Bezahlart", "Vor Ort (EUR)", "Vorkasse/Gutschein (EUR)", "Verwendungszweck"])
    for inv in ctx["invoices"]:
        prepaid = _invoice_prepaid_amount(inv)
        onsite = _invoice_onsite_amount(inv)
        net_total, vat_total, gross_total = _invoice_totals(inv)
        writer.writerow([
            inv.seq_number or inv.id,
            inv.created_at.strftime("%d.%m.%Y") if inv.created_at else "",
            inv.person.full_name if inv.person else "",
            _fmt(net_total),
            _fmt(vat_total),
            _fmt(gross_total),
            _invoice_payment_state_label(inv),
            _invoice_split_payment_label(inv),
            _fmt(onsite),
            _fmt(prepaid),
            _build_invoice_payment_purpose(inv),
        ])

    writer.writerow([])
    writer.writerow(["--- Zusammenfassung ---"])
    writer.writerow(["Bereits bezahlt", _fmt(ctx["sum_paid"])])
    paid_by_method = _paid_method_breakdown(ctx["invoices"])
    writer.writerow(["davon Bar", _fmt(paid_by_method["cash"])])
    writer.writerow(["davon Karte", _fmt(paid_by_method["card"])])
    writer.writerow(["davon Überweisung", _fmt(paid_by_method["transfer"])])
    writer.writerow(["davon WERO", _fmt(paid_by_method["wero"])])
    writer.writerow(["davon SEPA-Lastschrift", _fmt(paid_by_method["sepa"])])
    writer.writerow(["davon Vorkasse / Gutschein", _fmt(paid_by_method["voucher"])])
    writer.writerow(["Offen fakturiert", _fmt(ctx["sum_open_invoices"])])
    writer.writerow(["Noch nicht fakturiert", _fmt(ctx["sum_billable_uninvoiced"])])
    writer.writerow(["GESAMTSUMME abrechenbar", _fmt(ctx["sum_total"])])

    generated_at = now_berlin().strftime("%Y_%m_%d_%H_%M")
    filename = f"Rechnungsuebersicht_{period}_{generated_at}.csv"

    response = make_response(output.getvalue().encode("utf-8-sig"))
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------
# Excel-Export Rechnungsliste
# ---------------------------------------------------------
@bp.route("/invoices/export/xlsx")
def invoice_list_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("openpyxl nicht verfügbar – Excel-Export nicht möglich.", "danger")
        return redirect(url_for("billing.invoice_list", **request.args.to_dict(flat=True)))

    period = request.args.get("period", "all")
    from_str = (request.args.get("from") or request.args.get("from_date") or "").strip()
    to_str = (request.args.get("to") or request.args.get("to_date") or "").strip()
    invoice_from_str = (request.args.get("invoice_from") or "").strip()
    invoice_to_str = (request.args.get("invoice_to") or "").strip()
    filters = _parse_invoice_list_filters(request.args)
    delta_scope = _parse_invoice_delta_scope(request.args)

    ctx = _build_invoice_list_context(
        period=period, from_str=from_str, to_str=to_str,
        invoice_from_str=invoice_from_str, invoice_to_str=invoice_to_str,
        filters=filters, delta_scope=delta_scope,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rechnungen"

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    paid_fill = PatternFill("solid", fgColor="D1FAE5")
    open_fill = PatternFill("solid", fgColor="FEE2E2")
    total_fill = PatternFill("solid", fgColor="FEF3C7")
    total_font = Font(bold=True)
    thin = Side(style="thin")
    border = Border(bottom=thin)

    headers = ["Nr.", "Datum", "Person", "Netto (EUR)", "MwSt (EUR)", "Brutto (EUR)", "Status", "Bezahlart", "Vor Ort (EUR)", "Vorkasse/Gutschein (EUR)", "Verwendungszweck"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for inv in ctx["invoices"]:
        prepaid = float(_invoice_prepaid_amount(inv))
        onsite = float(_invoice_onsite_amount(inv))
        net_total, vat_total, gross_total = _invoice_totals(inv)
        row = [
            inv.seq_number or inv.id,
            inv.created_at.strftime("%d.%m.%Y") if inv.created_at else "",
            inv.person.full_name if inv.person else "",
            float(net_total),
            float(vat_total),
            float(gross_total),
            _invoice_payment_state_label(inv),
            _invoice_split_payment_label(inv),
            onsite,
            prepaid,
            _build_invoice_payment_purpose(inv),
        ]
        ws.append(row)
        row_idx = ws.max_row
        row_fill = paid_fill if _invoice_payment_state(inv) == INVOICE_PAYMENT_STATE_PAID else open_fill
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = row_fill
        # Beträge als Zahl formatieren
        ws.cell(row=row_idx, column=4).number_format = '#,##0.00 "€"'
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00 "€"'
        ws.cell(row=row_idx, column=6).number_format = '#,##0.00 "€"'
        ws.cell(row=row_idx, column=9).number_format = '#,##0.00 "€"'
        ws.cell(row=row_idx, column=10).number_format = '#,##0.00 "€"'

    # Leerzeile
    ws.append([])

    summary_rows = [
        ("Bereits bezahlt", float(ctx["sum_paid"])),
        ("Offen fakturiert", float(ctx["sum_open_invoices"])),
        ("Noch nicht fakturiert", float(ctx["sum_billable_uninvoiced"])),
        ("GESAMTSUMME abrechenbar", float(ctx["sum_total"])),
    ]
    paid_by_method = _paid_method_breakdown(ctx["invoices"])
    summary_rows = [
        ("Bereits bezahlt", float(ctx["sum_paid"])),
        ("davon Bar", float(paid_by_method["cash"])),
        ("davon Karte", float(paid_by_method["card"])),
        ("davon Überweisung", float(paid_by_method["transfer"])),
        ("davon WERO", float(paid_by_method["wero"])),
        ("davon SEPA-Lastschrift", float(paid_by_method["sepa"])),
        ("davon Vorkasse / Gutschein", float(paid_by_method["voucher"])),
        ("Offen fakturiert", float(ctx["sum_open_invoices"])),
        ("Noch nicht fakturiert", float(ctx["sum_billable_uninvoiced"])),
        ("GESAMTSUMME abrechenbar", float(ctx["sum_total"])),
    ]
    for label, value in summary_rows:
        ws.append([label, value])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=1).font = total_font
        ws.cell(row=row_idx, column=2).font = total_font
        ws.cell(row=row_idx, column=2).number_format = '#,##0.00 "€"'
        if label.startswith("GESAMT"):
            ws.cell(row=row_idx, column=1).fill = total_fill
            ws.cell(row=row_idx, column=2).fill = total_fill
            for col_idx in [1, 2]:
                ws.cell(row=row_idx, column=col_idx).border = border

    # Spaltenbreiten anpassen
    col_widths = [8, 12, 30, 14, 14, 14, 12, 24, 14, 20, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    generated_at = now_berlin().strftime("%Y_%m_%d_%H_%M")
    filename = f"Rechnungsuebersicht_{period}_{generated_at}.xlsx"

    response = make_response(excel_bytes.read())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------
# Liste stornierter Rechnungen (nur Admin)
# ---------------------------------------------------------
@bp.route("/invoices/cancelled")
def invoice_list_cancelled():
    deny = _full_admin_required("billing.invoice_list")
    if deny:
        return deny

    invoices = (
        Invoice.query
        .filter(
            Invoice.is_deleted.is_(True),
            Invoice.stage == "final"
        )
        .order_by(Invoice.deleted_at.desc().nullslast(), Invoice.id.desc())
        .all()
    )

    return render_template(
        "billing/invoice_list_cancelled.html",
        invoices=invoices,
        invoice_display_number=_invoice_display_number_for_detail,
    )
