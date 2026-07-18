# C:\manifest_fallschirm\app\routes\load.py
from __future__ import annotations

from datetime import datetime, date, time as dtime, timedelta
from io import BytesIO
from typing import Optional
import base64
import os
import socket
import qrcode
import csv
from decimal import Decimal
from collections import defaultdict
from urllib.parse import urlencode

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    jsonify,
    make_response,
    send_file,
    abort,
    current_app,
)

from sqlalchemy import or_

from app import db, now_local
now_berlin = now_local  # Alias für Abwärtskompatibilität
from app.models.load import Load
from app.models.load_entry import LoadEntry
from app.models.aircraft import Aircraft
from app.models.person import Person
from app.models.flugplatz import Flugplatz
from app.models.status_definition import StatusDefinition
from app.models.billing_config import BillingConfig, BillingPrice, BillingPricePeriod
from app.models.invoice import Invoice
from app.models.invoice_item import InvoiceItem
from app.services.billing_service import BillingService, _image_to_data_uri, _invoice_payment_label
from app.services.display_service import generate_qr_png_buffer, build_display_qr_url
from app.services.pdf_service import generate_pdf_from_html
from app.services.load_service import (
    operation_date_from_times,
    next_load_number_for_day,
    apply_follow_load_warnings,
    apply_follow_load_warning_single,
    ensure_pricing_model_available_for_load,
    archive_effective_datetime_expr as _archive_effective_datetime_expr,
    archive_effective_datetime as _archive_effective_datetime,
    _status_text_parts,
    _is_student_specific_status,
    _is_teacher_specific_status,
    _is_aff_teacher_status,
    _aff_student_level,
    _is_aff_student_status,
    _is_cost_status,
    _is_gear_rental_forbidden_status,
    _truthy,
    _parse_date_ymd,
    _dt_range,
    _money,
)
from app.helpers.network_utils import get_wifi_qr_string
from app.helpers.status_code import normalize_status_code
from app.utils import parse_int, parse_float, parse_bool
from app.constants import VALID_HEIGHTS, STUDENT_STATUSES, TANDEM_GUEST_STATUSES
from app.constants import (
    INVOICE_PAYMENT_STATE_OPEN,
    INVOICE_PAYMENT_STATE_SEPA_PENDING,
    INVOICE_PAYMENT_STATE_SEPA_EXPORTED,
    INVOICE_PAYMENT_STATE_PAID,
    INVOICE_PAYMENT_STATE_SEPA_RETURNED,
    INVOICE_PAYMENT_STATES,
)

bp_load = Blueprint("load", __name__, url_prefix="/loads")
MAX_EXTRA_SEATS_PER_LOAD = 4


# ============================================================
# Helpers (Admin / Locks)
# ============================================================
def is_admin() -> bool:
    return bool(session.get("is_admin"))


def lock_reason(load: Load) -> Optional[str]:
    """
    Nicht-Admin darf NICHT bearbeiten, wenn:
    - Load completed (durchgeführt)
    - oder billed (abgerechnet)
    - oder paid vorhanden
    Admin darf bearbeiten, außer:
    - paid-Entries sind immer gesperrt (teilweise Sperre).
    """
    if is_admin():
        return None
    if load.has_paid_entries:
        return "Load enthält bezahlte Einträge und kann nicht mehr bearbeitet werden."
    if load.has_billed_entries:
        return "Load wurde bereits abgerechnet. Änderungen sind nur im Admin-Bereich möglich."
    if load.status == "completed":
        return "Load schon durchgeführt. Änderungen sind nur im Admin-Bereich möglich."
    return None


# ------------------------------------------------------------
# ============================================================
# ✅ MINIMALE PREISMATRIX-ABSICHERUNG (Backend)
# Ziel:
# - Ein Load darf nur angelegt werden, wenn es für den Flugplatz am Load-Datum
#   mindestens eine gültige Preismatrix (Preisperiode) gibt, zu der Preise existieren.
# - Damit können keine Loads entstehen, die später "keine Preise finden".
# ============================================================
def ensure_pricematrix_available_for_load(airfield_id: int, on_date: date) -> None:
    """
    Prüft minimal:
      - Existiert mindestens eine BillingPricePeriod, die am on_date gültig ist
      - UND gibt es dafür mindestens einen BillingPrice für den airfield_id
    Wenn nicht: ValueError mit Nutzertext.
    """
    # gültige Perioden am Tag (valid_from <= date <= valid_to bzw. valid_to NULL)
    # und dazu mindestens ein Preis für diesen Flugplatz
    row = (
        db.session.query(BillingPricePeriod.id)
        .join(BillingPrice, BillingPrice.period_id == BillingPricePeriod.id)
        .filter(BillingPricePeriod.valid_from <= on_date)
        .filter(
            (BillingPricePeriod.valid_to.is_(None)) |
            (BillingPricePeriod.valid_to >= on_date)
        )
        .order_by(BillingPricePeriod.valid_from.desc())  # "neueste" gültige zuerst
        .limit(1)
        .first()
    )

    if not row:
        raise ValueError(
            "Für diesen Flugplatz gibt es am gewählten Datum keine gültige Preismatrix mit Preisen. "
            "Bitte unter „Preismatrix“ eine gültige Preismatrix auswählen oder eine neue aus Vorlage kopieren."
        )


# ============================================================
# ✅ PREISMODELL (Matrix) – AKTIVES MODELL & LOAD-BINDUNG
# Ziel:
# - Das Preismodell wird ausschließlich über /pricing/ gewählt.
# - Neue Loads übernehmen das aktuell aktive Preismodell.
# - Die Load-UI zeigt das Preismodell nur an (read-only).
# ============================================================

def get_active_pricing_model_id() -> Optional[int]:
    """
    Liefert das aktuell aktive Preismodell (BillingPricePeriod.id),
    das über /pricing/ gesetzt wurde.

    Quelle:
    - Session-Wert "active_pricing_model_id"
    """
    mid = session.get("active_pricing_model_id")
    try:
        return int(mid) if mid is not None else None
    except Exception:
        return None



def _parse_archive_period_args(args) -> dict:
    period = (args.get("period") or "all").strip().lower()
    if period not in {"all", "today", "week", "month", "year", "range"}:
        period = "all"

    from_str = (args.get("from") or "").strip()
    to_str = (args.get("to") or "").strip()

    today = date.today()
    start_date: Optional[date] = None
    end_date_exclusive: Optional[date] = None

    if period == "today":
        start_date = today
        end_date_exclusive = today + timedelta(days=1)
    elif period == "week":
        start_date = today - timedelta(days=today.weekday())
        end_date_exclusive = start_date + timedelta(days=7)
    elif period == "month":
        start_date = today.replace(day=1)
        if start_date.month == 12:
            end_date_exclusive = start_date.replace(year=start_date.year + 1, month=1)
        else:
            end_date_exclusive = start_date.replace(month=start_date.month + 1)
    elif period == "year":
        start_date = today.replace(month=1, day=1)
        end_date_exclusive = start_date.replace(year=start_date.year + 1)
    elif period == "range":
        try:
            if from_str:
                start_date = datetime.strptime(from_str, "%Y-%m-%d").date()
            if to_str:
                end_date_exclusive = datetime.strptime(to_str, "%Y-%m-%d").date() + timedelta(days=1)
            if start_date and end_date_exclusive and end_date_exclusive <= start_date:
                start_date = None
                end_date_exclusive = None
                period = "all"
        except Exception:
            start_date = None
            end_date_exclusive = None
            period = "all"

    start_dt = datetime.combine(start_date, dtime.min) if start_date else None
    end_dt = datetime.combine(end_date_exclusive, dtime.min) if end_date_exclusive else None

    period_labels = {
        "all": "Alle",
        "today": "Heute",
        "week": f"Woche (KW {today.isocalendar()[1]:02d})",
        "month": "Monat",
        "year": f"Jahr {today.year}",
        "range": "Zeitraum",
    }

    if period == "range" and (from_str or to_str):
        period_label = f"Zeitraum {from_str or '...'} bis {to_str or '...'}"
    else:
        period_label = period_labels.get(period, "Alle")

    return {
        "period": period,
        "from": from_str,
        "to": to_str,
        "start_dt": start_dt,
        "end_dt": end_dt,
        "label": period_label,
    }


def _apply_archive_period_filter(q, archive_period: dict):
    start_dt = archive_period.get("start_dt")
    end_dt = archive_period.get("end_dt")
    if start_dt is None and end_dt is None:
        return q

    eff = _archive_effective_datetime_expr()
    if start_dt is not None:
        q = q.filter(eff >= start_dt)
    if end_dt is not None:
        q = q.filter(eff < end_dt)
    return q


def _build_archive_load_query(args):
    archive_period = _parse_archive_period_args(args)
    q = (
        Load.query
        .filter(Load.status == "completed")
        .options(
            joinedload(Load.airfield),
            joinedload(Load.aircraft),
            selectinload(Load.entries).selectinload(LoadEntry.person),
        )
    )
    q = _apply_archive_period_filter(q, archive_period)
    q = q.order_by(Load.actual_time.desc().nullslast(), Load.created_at.desc())
    return q, archive_period


_TANDEM_EXPORT_COLORS = {
    1: ("#e8f1ff", "#6aa8ff"),
    2: ("#e9f9e9", "#7acb7a"),
    3: ("#fff7e0", "#e6c463"),
    4: ("#ffece5", "#e6a48c"),
    5: ("#f3e8ff", "#b79ce6"),
    6: ("#e7fbff", "#6fd3e6"),
    7: ("#f0ffe8", "#9ad36f"),
    8: ("#fff0f7", "#e6a1c4"),
    9: ("#f7f7ff", "#9aa8ff"),
    10: ("#fff6e8", "#e6b26f"),
}

_INSTRUCTION_EXPORT_COLORS = {
    1: ("#fff7e0", "#e6c463"),
    2: ("#fff2c7", "#e0b84e"),
    3: ("#fffbe8", "#d6c86a"),
    4: ("#fff3d9", "#d9a84e"),
    5: ("#fff8cc", "#d1b100"),
    6: ("#f0f6ff", "#6aa8ff"),
    7: ("#eef9f0", "#6fbf8a"),
    8: ("#fdf1f7", "#e6a1c4"),
    9: ("#f6f6f6", "#9aa0a6"),
    10: ("#fef6e8", "#e6b26f"),
}


def _archive_entry_css_class(load: Load, entry: LoadEntry) -> str:
    entry_id = getattr(entry, "id", None)
    if entry_id is None:
        return ""

    blocks = list(getattr(load, "blocks", None) or [])
    for b in blocks:
        if b.get("type") != "block":
            continue
        if entry_id in (b.get("entry_ids") or []):
            return (b.get("css_class") or "").strip()

    status = (getattr(entry, "status_code", "") or "").strip()
    if status in {
        "Schüler",
        "Schüler Ek 1",
        "Schüler Ek 2",
        "Schüler GK 6",
        "SCHUELER-AFF-1",
        "SCHUELER-AFF-2",
        "Schueler-Aff-1",
        "Schueler-Aff-2",
    }:
        return "instruction-seat instruction-seat-1"

    return ""


def _archive_colors_from_css_class(css_class: str) -> tuple[Optional[str], Optional[str]]:
    css = (css_class or "").strip()
    if not css:
        return (None, None)

    classes = set(css.split())
    tandem_idx = next(
        (
            int(c.replace("tandem-seat-", ""))
            for c in classes
            if c.startswith("tandem-seat-") and c.replace("tandem-seat-", "").isdigit()
        ),
        None,
    )
    instruction_idx = next(
        (
            int(c.replace("instruction-seat-", ""))
            for c in classes
            if c.startswith("instruction-seat-") and c.replace("instruction-seat-", "").isdigit()
        ),
        None,
    )

    if "tandem-seat" in classes:
        idx = tandem_idx or 1
        return _TANDEM_EXPORT_COLORS.get(idx, _TANDEM_EXPORT_COLORS[1])

    if "instruction-seat" in classes:
        idx = instruction_idx or 1
        return _INSTRUCTION_EXPORT_COLORS.get(idx, _INSTRUCTION_EXPORT_COLORS[1])

    return (None, None)


def _attach_archive_export_status_style(load: Load, entry: LoadEntry) -> None:
    css_class = _archive_entry_css_class(load, entry)
    bg, border = _archive_colors_from_css_class(css_class)
    setattr(entry, "archive_status_css_class", css_class)
    setattr(entry, "archive_status_bg", bg)
    setattr(entry, "archive_status_border", border)


def _build_archive_entry_rows(loads: list[Load]) -> list[dict]:
    rows: list[dict] = []
    for l in loads:
        eff_dt = _archive_effective_datetime(l)
        sorted_entries = sorted(
            list(getattr(l, "entries", None) or []),
            key=lambda e: ((getattr(e, "seat", None) is None), getattr(e, "seat", 0), getattr(e, "id", 0)),
        )
        for e in sorted_entries:
            person = getattr(e, "person", None)
            css_class = _archive_entry_css_class(l, e)
            status_bg, status_border = _archive_colors_from_css_class(css_class)
            rows.append({
                "load_id": getattr(l, "id", None),
                "load_number": getattr(l, "load_number", ""),
                "load_date": eff_dt.strftime("%d.%m.%Y") if eff_dt else "",
                "load_time": eff_dt.strftime("%H:%M") if eff_dt else "",
                "airfield": getattr(getattr(l, "airfield", None), "name", ""),
                "aircraft": f"{getattr(getattr(l, 'aircraft', None), 'type', '')} {getattr(getattr(l, 'aircraft', None), 'registration', '')}".strip(),
                "entry_id": getattr(e, "id", None),
                "seat": getattr(e, "seat", None),
                "person": getattr(person, "full_name", ""),
                "person_id": getattr(e, "person_id", None),
                "status": (getattr(e, "status_code", "") or "").strip(),
                "height_m": getattr(e, "height_m", None),
                "gear_rental": bool(getattr(e, "gear_rental", False)),
                "billed": bool(getattr(e, "billed", False)),
                "paid": bool(getattr(e, "paid", False)),
                "status_css_class": css_class,
                "status_bg": status_bg,
                "status_border": status_border,
            })
    return rows


# ============================================================
# BLOCK 1 — LOAD-LISTE MIT FILTERN + ARCHIV
# ============================================================
@bp_load.route("/")
def list_loads():
    # ✅ Wenn explizit gewünscht (Archiv / Filter / Admin-Navigation):
    if request.args:
        # normale Listenansicht
        q = Load.query
        show = request.args.get("show", "active")
        archive_period = _parse_archive_period_args(request.args)
        if show == "archive":
            q = q.filter(Load.status == "completed")
            q = _apply_archive_period_filter(q, archive_period)
        else:
            q = q.filter(Load.status != "completed")
        loads = q.order_by(Load.created_at.desc()).all()
        apply_follow_load_warnings(loads)

        # ✅ NUR aktive (nicht archivierte) Flugplätze für Dropdown/Filter anzeigen
        airfields = (
            Flugplatz.query
            .filter(Flugplatz.deleted_at.is_(None))
            .filter(Flugplatz.active.is_(True))
            .order_by(Flugplatz.name.asc())
            .all()
        )
        aircrafts = Aircraft.query.order_by(Aircraft.type.asc()).all()
        return render_template(
            "load/list.html",
            loads=loads,
            airfields=airfields,
            aircrafts=aircrafts,
            show=show,
            archive_period=archive_period["period"],
            archive_from=archive_period["from"],
            archive_to=archive_period["to"],
        )

    # ✅ Default-Einstieg: Split-View / active
    return redirect(url_for("load.split_view", show="active"))


# ============================================================
# BLOCK 1a split — View (ROBUST, mit korrekten Defaults)
# ============================================================
@bp_load.route("/split")
def split_view():
    show = request.args.get("show", "active")
    archive_period = _parse_archive_period_args(request.args)

    # -------------------------------
    # Rechter Editor (optional)
    # ✅ MINIMAL-FIX: edit_id früh lesen, damit wir new=1 entfernen können,
    # wenn edit gesetzt ist (sonst bleibt man im new-Formular hängen).
    # -------------------------------
    edit_id = request.args.get("edit")
    if edit_id and request.args.get("new") == "1":
        return redirect(url_for("load.split_view", edit=edit_id, show=show))

    # -------------------------------
    # Linke Liste (gefiltert)
    # -------------------------------
    q = Load.query
    if show == "archive":
        q = q.filter(Load.status == "completed")
        q = _apply_archive_period_filter(q, archive_period)
    else:
        q = q.filter(Load.status != "completed")
    loads = q.order_by(Load.created_at.desc()).all()
    apply_follow_load_warnings(loads)

    # -------------------------------
    # Rechter Editor (optional)
    # -------------------------------
    load = None
    if edit_id:
        try:
            edit_id_int = int(edit_id)
            candidate = Load.query.get(edit_id_int)
            # Editor-Load nur anzeigen, wenn er zum aktuellen Listenfilter passt.
            if candidate and (
                (show == "archive" and candidate.status == "completed")
                or (show != "archive" and candidate.status != "completed")
            ):
                load = candidate
        except Exception:
            load = None

    # -------------------------------
    # ✅ Aktive Flugplätze
    # -------------------------------
    airfields = (
        Flugplatz.query
        .filter(Flugplatz.deleted_at.is_(None))
        .filter(Flugplatz.active.is_(True))
        .order_by(Flugplatz.name.asc())
        .all()
    )

    # -------------------------------
    # ✅ Aktive Flugzeuge
    # -------------------------------
    aircrafts = (
        Aircraft.query
        .filter(Aircraft.active.is_(True))
        .order_by(Aircraft.type.asc(), Aircraft.registration.asc())
        .all()
    )

    # -------------------------------
    # ✅ Default‑Flugplatz bestimmen
    # Reihenfolge:
    # 1) Session
    # 2) Heimatflugplatz
    # 3) erster aktiver
    # -------------------------------
    session_airfield_id = session.get("current_airfield_id")
    home_airfield = next(
        (af for af in airfields if getattr(af, "is_home_airfield", False)),
        None
    )

    if session_airfield_id and any(af.id == session_airfield_id for af in airfields):
        default_airfield_id = session_airfield_id
    elif home_airfield:
        default_airfield_id = home_airfield.id
    else:
        default_airfield_id = airfields[0].id if airfields else None

    # -------------------------------
    # ✅ Default‑Flugzeug bestimmen
    # (erstes aktives)
    # -------------------------------
    default_aircraft_id = aircrafts[0].id if aircrafts else None

    return render_template(
        "load/split.html",
        loads=loads,
        load=load,
        airfields=airfields,
        aircrafts=aircrafts,
        default_airfield_id=default_airfield_id,
        default_aircraft_id=default_aircraft_id,
        show=show,
        archive_period=archive_period["period"],
        archive_from=archive_period["from"],
        archive_to=archive_period["to"],
        edit=edit_id,  # ✅ wichtig für Editor
        is_split_view=True,
    )


@bp_load.route("/archive/export.csv", endpoint="archive_export_csv")
def archive_export_csv():
    q, archive_period = _build_archive_load_query(request.args)
    loads = q.all()
    rows = _build_archive_entry_rows(loads)

    from io import StringIO

    text_buffer = StringIO()
    text_buffer.write("\ufeff")

    columns = [
        "load_id", "load_number", "load_date", "load_time",
        "airfield", "aircraft",
        "entry_id", "seat", "person_id", "person", "status", "height_m",
        "gear_rental", "billed", "paid",
    ]

    writer = csv.DictWriter(
        text_buffer,
        fieldnames=columns,
        delimiter=";",
        lineterminator="\n",
        extrasaction="ignore",
    )

    writer.writerow({
        "load_id": "Load-ID",
        "load_number": "Load",
        "load_date": "Datum",
        "load_time": "Zeit",
        "airfield": "Flugplatz",
        "aircraft": "Flugzeug",
        "entry_id": "Entry-ID",
        "seat": "Sitz",
        "person_id": "Person-ID",
        "person": "Person",
        "status": "Status",
        "height_m": "Sprunghöhe (m)",
        "gear_rental": "Schirmmiete",
        "billed": "Abgerechnet",
        "paid": "Bezahlt",
    })

    for row in rows:
        out = dict(row)
        out["gear_rental"] = "Ja" if row.get("gear_rental") else "Nein"
        out["billed"] = "Ja" if row.get("billed") else "Nein"
        out["paid"] = "Ja" if row.get("paid") else "Nein"
        writer.writerow(out)

    output = BytesIO(text_buffer.getvalue().encode("utf-8"))
    output.seek(0)
    filename = f"loads_archiv_status_{archive_period['period']}.csv"
    return send_file(
        output,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=filename,
    )


@bp_load.route("/archive/export.xlsx", endpoint="archive_export_xlsx")
def archive_export_xlsx():
    q, archive_period = _build_archive_load_query(request.args)
    loads = q.all()
    rows = _build_archive_entry_rows(loads)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return redirect(url_for("load.archive_export_csv", **request.args))

    wb = Workbook()
    ws = wb.active
    ws.title = "Loads Archiv"

    columns = [
        "load_id", "load_number", "load_date", "load_time",
        "airfield", "aircraft",
        "entry_id", "seat", "person_id", "person", "status", "height_m",
        "gear_rental", "billed", "paid",
    ]
    headers = {
        "load_id": "Load-ID",
        "load_number": "Load",
        "load_date": "Datum",
        "load_time": "Zeit",
        "airfield": "Flugplatz",
        "aircraft": "Flugzeug",
        "entry_id": "Entry-ID",
        "seat": "Sitz",
        "person_id": "Person-ID",
        "person": "Person",
        "status": "Status",
        "height_m": "Sprunghöhe (m)",
        "gear_rental": "Schirmmiete",
        "billed": "Abgerechnet",
        "paid": "Bezahlt",
    }

    ws.append([headers.get(c, c) for c in columns])
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row.get("load_id"),
            row.get("load_number"),
            row.get("load_date"),
            row.get("load_time"),
            row.get("airfield"),
            row.get("aircraft"),
            row.get("entry_id"),
            row.get("seat"),
            row.get("person_id"),
            row.get("person"),
            row.get("status"),
            row.get("height_m"),
            "Ja" if row.get("gear_rental") else "Nein",
            "Ja" if row.get("billed") else "Nein",
            "Ja" if row.get("paid") else "Nein",
        ])

        current_row = ws.max_row
        status_bg = row.get("status_bg")
        status_border = row.get("status_border")
        if status_bg:
            ws.cell(row=current_row, column=11).fill = PatternFill(
                "solid",
                fgColor=status_bg.replace("#", "").upper(),
            )
        if status_border:
            left_border = Border(
                left=Side(style="thick", color=status_border.replace("#", "").upper())
            )
            ws.cell(row=current_row, column=10).border = left_border
            ws.cell(row=current_row, column=11).border = left_border

    col_widths = {
        "A": 9, "B": 8, "C": 12, "D": 8,
        "E": 22, "F": 22,
        "G": 9, "H": 7, "I": 9, "J": 28, "K": 16, "L": 14,
        "M": 12, "N": 12, "O": 10,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row_cells, start=1):
            if idx in {1, 2, 7, 8, 9, 12}:
                cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"loads_archiv_status_{archive_period['period']}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@bp_load.route("/archive/export.pdf", endpoint="archive_export_pdf")
def archive_export_pdf():
    q, archive_period = _build_archive_load_query(request.args)
    loads = q.all()

    for l in loads:
        entries = sorted(
            list(getattr(l, "entries", None) or []),
            key=lambda e: ((getattr(e, "seat", None) is None), getattr(e, "seat", 0), getattr(e, "id", 0)),
        )
        for e in entries:
            _attach_archive_export_status_style(l, e)
        setattr(l, "entries_sorted", entries)

    total_entries = sum(len(getattr(l, "entries", None) or []) for l in loads)
    generated_at = now_berlin().replace(tzinfo=None)

    html = render_template(
        "load/archive_export_pdf.html",
        loads=loads,
        archive_period_label=archive_period["label"],
        archive_from=archive_period["from"],
        archive_to=archive_period["to"],
        total_loads=len(loads),
        total_entries=total_entries,
        generated_at=generated_at,
        archive_effective_datetime=_archive_effective_datetime,
    )

    pdf_bytes, pdf_error = generate_pdf_from_html(html)
    if pdf_error:
        flash(pdf_error, "danger")
        return redirect(request.referrer or url_for("load.list_loads", show="archive"))

    filename = f"loads_archiv_status_{generated_at.strftime('%Y_%m_%d_%H_%M')}.pdf"
    return send_file(
        BytesIO(pdf_bytes),
        mimetype="application/pdf",
        download_name=filename,
        as_attachment=False,
    )


# ============================================================
# QR PNG (OFFLINE)
# ============================================================
@bp_load.route("/qr.png")
def qr_png():
    """
    Offline-QR-Code für Display / Mobile
    Nutzt display_service.generate_qr_png_buffer() für PNG-Generierung.
    """
    data = request.args.get("data", "", type=str).strip()
    if not data:
        abort(400)

    size = request.args.get("size", 150, type=int)
    
    buf = generate_qr_png_buffer(data, size)
    resp = make_response(send_file(buf, mimetype="image/png"))
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    return resp


# ============================================================
# DISPLAY-HELPER: Startzeit eines Loads ermitteln
# ============================================================
def _load_start_dt(load: Load) -> Optional[datetime]:
    """
    Bestimmt die Startzeit eines Loads.
    Primär: actual_time; Fallback: scheduled_time (Legacy).
    Wandelt time → datetime via created_at.
    """
    t = getattr(load, "actual_time", None)
    if t is None:
        t = getattr(load, "scheduled_time", None)
    if t is None:
        return None
    if isinstance(t, datetime):
        return t
    if isinstance(t, dtime):
        created = getattr(load, "created_at", None)
        if created is None:
            return None
        return datetime.combine(created.date(), t)
    return None


# ============================================================
# DISPLAY VIEW (READ-ONLY)
# ============================================================
@bp_load.route("/display")
def display_view():
    loads = (
        Load.query
        .filter(Load.status != "completed")
        .order_by(Load.created_at.desc())
        .all()
    )
    apply_follow_load_warnings(loads)
    now = now_local().replace(tzinfo=None)

    timed: list[tuple[datetime, Load]] = []
    unknown: list[Load] = []

    for l in loads:
        sd = _load_start_dt(l)
        if sd is None:
            unknown.append(l)
        else:
            timed.append((sd, l))

    # ============================================================
    # QR-URL (LOCK): kommt NUR aus veröffentlichter URL
    # ============================================================
    qr_available, qr_url = build_display_qr_url()
    wifi_qr_available, wifi_qr_data = get_wifi_qr_string()

    if not timed:
        return render_template(
            "load/display.html",
            loads=unknown,
            qr_url=qr_url,
            qr_available=qr_available,
            wifi_qr_available=wifi_qr_available,
            wifi_qr_data=wifi_qr_data,
        )

    past = [(sd, l) for (sd, l) in timed if sd <= now]
    future = [(sd, l) for (sd, l) in timed if sd > now]

    # -----------------------------
    # AKTUELL: nur wenn Startzeit <= 35 Minuten zurückliegt
    # -----------------------------
    current_pair = max(past, key=lambda x: x[0]) if past else None
    current_sd, current_load = current_pair if current_pair else (None, None)

    if current_sd is not None and (now - current_sd) > timedelta(minutes=35):
        current_sd, current_load = (None, None)

    # NÄCHSTER LOAD: nächster Start in der Zukunft
    next_pair = min(future, key=lambda x: x[0]) if future else None
    next_sd, next_load = next_pair if next_pair else (None, None)

    # Reihenfolge wie bisher: unknown, future(desc), current, past(desc)
    future_sorted_desc = sorted(future, key=lambda x: x[0], reverse=True)
    past_rest = [(sd, l) for (sd, l) in past if (current_load is None or l is not current_load)]
    past_sorted_desc = sorted(past_rest, key=lambda x: x[0], reverse=True)

    ordered: list[Load] = []
    ordered.extend(unknown)
    ordered.extend([l for _, l in future_sorted_desc])
    if current_load is not None:
        ordered.append(current_load)
    ordered.extend([l for _, l in past_sorted_desc])

    return render_template(
        "load/display.html",
        loads=ordered,
        current_load_id=current_load.id if current_load else None,
        next_load_id=next_load.id if next_load else None,
        next_start_iso=next_sd.isoformat() if next_sd else None,
        server_now_iso=now.isoformat(),
        qr_url=qr_url,
        qr_available=qr_available,
        wifi_qr_available=wifi_qr_available,
        wifi_qr_data=wifi_qr_data,
    )
            
# ============================================================
# DISPLAY-QUER VIEW (READ-ONLY)
# ============================================================
@bp_load.route("/display-quer")
def display_quer_view():
    loads = (
        Load.query
        .filter(Load.status != "completed")
        .order_by(Load.created_at.desc())
        .all()
    )
    apply_follow_load_warnings(loads)
    now = now_local().replace(tzinfo=None)

    timed: list[tuple[datetime, Load]] = []
    unknown: list[Load] = []

    for l in loads:
        sd = _load_start_dt(l)
        if sd is None:
            unknown.append(l)
        else:
            timed.append((sd, l))

    qr_available, qr_url = build_display_qr_url()
    wifi_qr_available, wifi_qr_data = get_wifi_qr_string()

    # Wenn keine gültigen Startzeiten vorhanden sind, zeigen wir unknown (stabil)
    if not timed:
        return render_template(
            "load/display-quer.html",
            loads=unknown,
            qr_url=qr_url,
            qr_available=qr_available,
            wifi_qr_available=wifi_qr_available,
            wifi_qr_data=wifi_qr_data,
        )

    # Marker (aktueller / nächster) – wie in display_view
    past = [(sd, l) for (sd, l) in timed if sd <= now]
    future = [(sd, l) for (sd, l) in timed if sd > now]

    current_pair = max(past, key=lambda x: x[0]) if past else None
    current_sd, current_load = current_pair if current_pair else (None, None)

    if current_sd is not None and (now - current_sd) > timedelta(minutes=35):
        current_sd, current_load = (None, None)

    next_pair = min(future, key=lambda x: x[0]) if future else None
    next_sd, next_load = next_pair if next_pair else (None, None)

    # ✅ DISPLAY-QUER SORTIERUNG:
    # Ältester (Vergangenheit/aktuell) links, Zukunft weiter rechts -> Startzeit AUFSTEIGEND
    timed_sorted_asc = sorted(timed, key=lambda x: x[0])  # ASC
    ordered: list[Load] = [l for _, l in timed_sorted_asc]

    # Loads ohne Startzeit ganz links davor (falls vorhanden)
    ordered = unknown + ordered

    return render_template(
        "load/display-quer.html",
        loads=ordered,
        current_load_id=current_load.id if current_load else None,
        next_load_id=next_load.id if next_load else None,
        next_start_iso=next_sd.isoformat() if next_sd else None,
        server_now_iso=now.isoformat(),
        qr_url=qr_url,
        qr_available=qr_available,
        wifi_qr_available=wifi_qr_available,
        wifi_qr_data=wifi_qr_data,
    )


    # ============================================================
    # QR-URL (LOCK): kommt NUR aus veröffentlichter URL
    # ============================================================
    qr_available, qr_url = build_display_qr_url()

    if not timed:
        return render_template(
            "load/display.html",
            loads=unknown,
            qr_url=qr_url,
            qr_available=qr_available,
        )

    past = [(sd, l) for (sd, l) in timed if sd <= now]
    future = [(sd, l) for (sd, l) in timed if sd > now]

    # -----------------------------
    # AKTUELL: nur wenn Startzeit <= 35 Minuten zurückliegt
    # -----------------------------
    current_pair = max(past, key=lambda x: x[0]) if past else None
    current_sd, current_load = current_pair if current_pair else (None, None)

    if current_sd is not None:
        if (now - current_sd) > timedelta(minutes=35):
            # zu alt -> kein aktueller Load mehr
            current_sd, current_load = (None, None)

    # NÄCHSTER LOAD: nächster Start in der Zukunft
    next_pair = min(future, key=lambda x: x[0]) if future else None
    next_sd, next_load = next_pair if next_pair else (None, None)

    future_sorted_desc = sorted(future, key=lambda x: x[0], reverse=True)
    past_rest = [
        (sd, l)
        for (sd, l) in past
        if current_load is None or l is not current_load
    ]
    past_sorted_desc = sorted(past_rest, key=lambda x: x[0], reverse=True)

    ordered: list[Load] = []
    ordered.extend(unknown)
    ordered.extend([l for _, l in future_sorted_desc])
    if current_load is not None:
        ordered.append(current_load)
    ordered.extend([l for _, l in past_sorted_desc])

    return render_template(
        "load/display.html",
        loads=ordered,
        current_load_id=current_load.id if current_load else None,
        next_load_id=next_load.id if next_load else None,
        next_start_iso=next_sd.isoformat() if next_sd else None,
        server_now_iso=now.isoformat(),
        qr_url=qr_url,
        qr_available=qr_available,
    )


# ============================================================
# BLOCK 2 — API: PERSONENSUCHE (Autocomplete)
# ============================================================
@bp_load.route("/api/person/search")
def api_person_search():
    q = request.args.get("q", "").strip()
    base_query = Person.query.filter(Person.deleted_at.is_(None))

    if q:
        base_query = base_query.filter(or_(
            Person.first_name.ilike(f"%{q}%"),
            Person.last_name.ilike(f"%{q}%"),
            Person.phone.ilike(f"%{q}%"),
        ))

    persons = (
        base_query
        .order_by(Person.last_name.asc(), Person.first_name.asc())
        .limit(200)
        .all()
    )

    # ✅ NEU: teacher_license_* und liability_waiver_year für Editor/Filter
    return jsonify([
        {
            "id": p.id,
            "name": p.full_name,
            "is_member": p.is_member,
            "is_partner_verein": getattr(p, "is_partner_verein", False),
            "is_tandem_guest": p.is_tandem_guest,
            "is_tandemmaster": getattr(p, "is_tandemmaster", False),
            "is_student": getattr(p, "is_student", False),
            "is_video": getattr(p, "is_video", False),
            "is_aff_student": getattr(p, "is_aff_student", False),
            "is_teacher": p.is_teacher,
            "is_aff_teacher": getattr(p, "is_aff_teacher", False),
            "weight_kg": p.weight_kg,

            "liability_waiver_valid": getattr(p, "liability_waiver_valid", False),
            "liability_waiver_year": getattr(p, "liability_waiver_year", None),

            "teacher_license_valid": getattr(p, "teacher_license_valid", False),
            "teacher_license_status": getattr(p, "teacher_license_status", "none"),
            "teacher_license_expires": (
                p.teacher_license_expires.isoformat()
                if getattr(p, "teacher_license_expires", None)
                else None
            ),

            "deleted_at": p.deleted_at.isoformat() if getattr(p, "deleted_at", None) else None,
        }
        for p in persons
    ])


# ============================================================
# BLOCK 2b — API: PERSON BY ID
# ============================================================
@bp_load.route("/api/person/<int:person_id>")
def api_person_by_id(person_id: int):
    p = Person.query.get_or_404(person_id)
    if getattr(p, "deleted_at", None):
        return jsonify({"error": "Person deleted"}), 404

    # ✅ NEU: teacher_license_* und liability_waiver_year für Editor/Filter
    return jsonify({
        "id": p.id,
        "name": p.full_name,
        "is_member": p.is_member,
        "is_partner_verein": getattr(p, "is_partner_verein", False),
        "is_tandem_guest": p.is_tandem_guest,
        "is_tandemmaster": getattr(p, "is_tandemmaster", False),
        "is_student": getattr(p, "is_student", False),
        "is_video": getattr(p, "is_video", False),
        "is_aff_student": getattr(p, "is_aff_student", False),
        "is_teacher": p.is_teacher,
        "is_aff_teacher": getattr(p, "is_aff_teacher", False),
        "weight_kg": p.weight_kg,

        "liability_waiver_valid": getattr(p, "liability_waiver_valid", False),
        "liability_waiver_year": getattr(p, "liability_waiver_year", None),

        "teacher_license_valid": getattr(p, "teacher_license_valid", False),
        "teacher_license_status": getattr(p, "teacher_license_status", "none"),
        "teacher_license_expires": (
            p.teacher_license_expires.isoformat()
            if getattr(p, "teacher_license_expires", None)
            else None
        ),

        "deleted_at": None,
    })

def _slot_for_person(person: Optional[Person]) -> str:
    if not person:
        return ""
    if bool(getattr(person, "is_teacher", False)) or bool(getattr(person, "is_aff_teacher", False)):
        return "teacher"
    if bool(getattr(person, "is_student", False)) or bool(getattr(person, "is_aff_student", False)):
        return "student"
    return ""


def _status_allowed_for_person_and_slot(code: str, label: str, person: Optional[Person], slot: str) -> bool:
    slot_norm = (slot or "").strip().lower()

    is_aff_teacher = bool(person and getattr(person, "is_aff_teacher", False))
    is_aff_student = bool(person and getattr(person, "is_aff_student", False))
    is_student = bool(person and getattr(person, "is_student", False))
    is_teacher = bool(person and getattr(person, "is_teacher", False))
    student_allowed = is_student or is_aff_student
    teacher_allowed = is_teacher or is_aff_teacher
    is_student_specific = _is_student_specific_status(code, label)
    is_teacher_specific = _is_teacher_specific_status(code, label)
    is_general = not is_student_specific and not is_teacher_specific

    if _is_cost_status(code, label):
        return False

    if not person:
        if slot_norm == "student":
            return is_general or is_student_specific
        if slot_norm == "teacher":
            return is_general or is_teacher_specific
        return is_general

    if _is_aff_teacher_status(code) and not is_aff_teacher:
        return False

    if _is_aff_student_status(code) and not is_aff_student:
        return False

    if slot_norm == "teacher" and is_student_specific:
        return False
    if slot_norm == "student" and is_teacher_specific:
        return False

    if student_allowed and not teacher_allowed:
        return is_student_specific

    if teacher_allowed and not student_allowed:
        if is_student_specific:
            return False
        return is_general or is_teacher_specific

    if student_allowed and teacher_allowed:
        if slot_norm == "student":
            return is_student_specific
        if slot_norm == "teacher":
            return is_general or is_teacher_specific
        return is_general or is_teacher_specific or is_student_specific

    return is_general


# ============================================================
# BLOCK 3 — API: STATUS-LISTE
# ============================================================
@bp_load.route("/api/status/list")



def api_status_list():
    """
    Kontext- und personenabhängige Statusauswahl für das Load-Dropdown.
    Query-Parameter:
      - slot=teacher|student (Pflicht für Block-Kontext)
      - person_id (optional, aber empfohlen)
    """
    slot = request.args.get("slot", "").strip().lower()
    person_id = request.args.get("person_id", type=int)

    # Auffüller-Status (werden ggf. ergänzt)
    AUFFUELLER = [
        {"code": "Auffüller Verein", "label": "Auffüller Verein", "sort_order": 999},
        {"code": "Auffüller Gast", "label": "Auffüller Gast", "sort_order": 999},
        {"code": "Auffüller Partner-Verein", "label": "Auffüller Partner-Verein", "sort_order": 999},
    ]

    # Query-Basis: nur aktive Status
    q = StatusDefinition.query.filter(StatusDefinition.is_active.is_(True))
    statuses = q.order_by(StatusDefinition.sort_order.asc(), StatusDefinition.code.asc()).all()

    # Person laden (wenn vorhanden)
    p = None
    if person_id:
        p = Person.query.get(person_id)

    filtered = []
    for s in statuses:
        code = (s.code or "").strip()
        label = (s.label or "").strip()
        if not _status_allowed_for_person_and_slot(code, label, p, slot):
            continue
        filtered.append(s)

    # "SCHUELER-AFF-1/2" ggf. ergänzen, wenn sie nicht in der DB sind, aber gebraucht werden
    if slot == "student" and p and getattr(p, "is_aff_student", False):
        for code in ("SCHUELER-AFF-1", "SCHUELER-AFF-2"):
            if not any((s.code or "").strip() == code for s in filtered):
                filtered.append(type("FakeStatus", (), {"code": code, "label": code, "sort_order": 200})())

    # Auffüller-Status am Ende ergänzen, falls nicht vorhanden
    existing_codes = set((s.code or "") for s in filtered)
    for auff in AUFFUELLER:
        if auff["code"] not in existing_codes:
            filtered.append(type("FakeStatus", (), auff)())

    def _display_label(code: str, label: str) -> str:
        c = (code or "").strip()
        l = (label or "").strip()
        ll = l.lower()
        # Legacy-Langtexte im Load-Editor kurz halten.
        if c == "Verein" or ll == "vereinsmitglied":
            return "Verein"
        if c == "Gast" or ll == "fallschirmspringer gast":
            return "Gast"
        return l or c

    return jsonify([
        {"code": s.code, "label": _display_label(s.code, s.label), "sort_order": getattr(s, "sort_order", 100)}
        for s in filtered
    ])


# ============================================================
# BLOCK 4 — NEUEN LOAD ANLEGEN (Auto + manuell)
# Fachregel: ES GIBT NUR EIN ZEITFELD -> actual_time
# - Zukunft  => geplante Startzeit
# - Jetzt    => aktuelle Startzeit
# - Vergangenheit => vergangene Startzeit
# Nummerierung/Betriebstag basiert NUR auf actual_time.date()
# ============================================================
@bp_load.route("/new", methods=["GET", "POST"])
def new_load():
    # ✅ NUR aktive (nicht archivierte) Flugplätze für die Neuanlage anbieten
    airfields = (
        Flugplatz.query
        .filter(Flugplatz.deleted_at.is_(None))
        .filter(Flugplatz.active.is_(True))
        .order_by(Flugplatz.name.asc())
        .all()
    )
    aircrafts = (
        Aircraft.query
        .filter(Aircraft.active.is_(True))
        .order_by(Aircraft.type.asc(), Aircraft.registration.asc())
        .all()
    )

    # Schutz: ohne Stammdaten keine Neuanlage
    if not airfields or not aircrafts:
        flash("Kein Flugplatz oder Flugzeug verfügbar – bitte zuerst anlegen.", "danger")
        return redirect(url_for("load.list_loads"))

    # ✅ Heimatflugplatz nur verwenden, wenn er ebenfalls aktiv & nicht archiviert ist
    home_airfield = (
        Flugplatz.query
        .filter(Flugplatz.is_home_airfield.is_(True))
        .filter(Flugplatz.deleted_at.is_(None))
        .filter(Flugplatz.active.is_(True))
        .first()
    )
    session_airfield_id = session.get("current_airfield_id")

    # ✅ Default-Flugplatz bestimmen (robust gegen inaktive/archivierte Session-Werte)
    if session_airfield_id and any(af.id == session_airfield_id for af in airfields):
        default_airfield_id = session_airfield_id
    elif home_airfield:
        default_airfield_id = home_airfield.id
    else:
        default_airfield_id = airfields[0].id

    default_aircraft_id = aircrafts[0].id

    # ✅ Default Max Payload vom letzten Load am Flugplatz übernehmen (falls vorhanden)
    default_max_payload = None
    last_load_for_airfield = (
        Load.query
        .filter_by(airfield_id=default_airfield_id)
        .order_by(Load.created_at.desc())
        .first()
    )
    if last_load_for_airfield and last_load_for_airfield.max_payload_kg:
        default_max_payload = last_load_for_airfield.max_payload_kg

    # ✅ UX: Default-Zeit = jetzt + 20 Minuten
    fallback_dt = now_local().replace(tzinfo=None) + timedelta(minutes=20)
    default_planned_date = fallback_dt.strftime("%Y-%m-%d")
    default_planned_start_time = fallback_dt.strftime("%H:%M")

    # ------------------------------------------------------------
    # AUTO-MODUS (GET): erzeugt sofort einen Load (wie vorher)
    # ------------------------------------------------------------
    if request.method == "GET" and request.args.get("manual") != "1":
        session["current_airfield_id"] = default_airfield_id
        ac = Aircraft.query.get(default_aircraft_id)
        default_height_m = int(getattr(ac, "default_height", 3000) or 3000)

        planned_date_raw = (request.args.get("planned_date") or "").strip()
        planned_start_time_raw = (request.args.get("planned_start_time") or "").strip()

        # Datum
        try:
            op_date = datetime.strptime(planned_date_raw, "%Y-%m-%d").date() if planned_date_raw else fallback_dt.date()
        except Exception:
            op_date = fallback_dt.date()

        # Uhrzeit
        time_raw = planned_start_time_raw or fallback_dt.strftime("%H:%M")
        try:
            hour, minute = map(int, time_raw.split(":"))
            planned_dt = datetime(op_date.year, op_date.month, op_date.day, hour, minute)
        except Exception:
            planned_dt = datetime(op_date.year, op_date.month, op_date.day, fallback_dt.hour, fallback_dt.minute)

        # ------------------------------------------------------------
        # ✅ PREISMATRIX-ABSICHERUNG (AUTO) – bleibt erhalten
        # ------------------------------------------------------------
        try:
            ensure_pricematrix_available_for_load(default_airfield_id, planned_dt.date())
        except ValueError as e:
            flash(str(e), "danger")
            # Direkt zur Preismatrix-Seite für diesen Flugplatz springen
            return redirect(url_for("pricing.pricing_matrix"))

        # ------------------------------------------------------------
        # ✅ D.1: PREISMODELL-FREEZE pro Load
        # - Preismodell wird über /pricing/ gesetzt: session["active_pricing_model_id"]
        # - Fallback (Alt/Robust): wenn nicht gesetzt oder nicht passend, nehmen wir das
        #   "beste" gültige Modell am Tag für den Flugplatz (wie bisherige Logik).
        # ------------------------------------------------------------
        pricing_model_id = None

        # 1) Primär: explizit gesetztes aktives Preismodell aus /pricing/
        try:
            mid = session.get("active_pricing_model_id")
            pricing_model_id = int(mid) if mid is not None else None
        except Exception:
            pricing_model_id = None

        # 2) Validierung: existieren Preise für dieses Modell?
        if pricing_model_id is not None:
            exists_price = (
                db.session.query(BillingPrice.id)
                .filter(BillingPrice.period_id == pricing_model_id)
                .limit(1)
                .first()
            )
            if not exists_price:
                # Fallback: Session-Modell hat keine Preise -> wir wählen automatisch ein gültiges
                pricing_model_id = None

        # 3) Fallback: bestes gültiges Modell am Datum ermitteln
        if pricing_model_id is None:
            row = (
                db.session.query(BillingPricePeriod.id)
                .join(BillingPrice, BillingPrice.period_id == BillingPricePeriod.id)
                .filter(BillingPricePeriod.valid_from <= planned_dt.date())
                .filter(
                    (BillingPricePeriod.valid_to.is_(None)) |
                    (BillingPricePeriod.valid_to >= planned_dt.date())
                )
                .order_by(BillingPricePeriod.valid_from.desc())
                .limit(1)
                .first()
            )
            if row:
                pricing_model_id = int(row[0])
                session["active_pricing_model_id"] = pricing_model_id

        # Wenn wir trotz Absicherung nichts gefunden haben: sicher abbrechen
        if pricing_model_id is None:
            flash(
                "Kein Preismodell gefunden/gesetzt. Bitte unter „Preismatrix“ ein Preismodell aktiv setzen.",
                "danger"
            )
            return redirect(url_for("pricing.pricing_matrix"))

        # ✅ Nummerierung NUR nach Flugplatz + actual_time.date()
        next_number = next_load_number_for_day(
            default_airfield_id,
            planned_dt.date()
        )

        load = Load(
            airfield_id=default_airfield_id,
            aircraft_id=default_aircraft_id,
            load_number=next_number,
            height_m=default_height_m,
            actual_time=planned_dt,  # ✅ EIN Zeitfeld
            max_payload_kg=default_max_payload,
            status="open",
            pricing_model_id=pricing_model_id,  # ✅ NEU: Freeze Preismodell
        )
        db.session.add(load)
        db.session.commit()

        flash(f"Neuer Load {load.load_number} angelegt.", "success")
        if request.args.get("return_to") == "split":
            requested_show = request.args.get("show", "active")
            target_show = "active" if requested_show == "archive" else requested_show
            return redirect(url_for("load.split_view", edit=load.id, show=target_show, period="all"))
        return redirect(url_for("load.edit_load", id=load.id))

    # ------------------------------------------------------------
    # MANUELL: POST
    # ------------------------------------------------------------
    if request.method == "POST":
        try:
            airfield_id = int(request.form["airfield_id"])
            aircraft_id = int(request.form["aircraft_id"])
        except Exception:
            flash("Flugplatz oder Flugzeug ungültig.", "danger")
            return redirect(url_for("load.new_load"))

        session["current_airfield_id"] = airfield_id

        ac = Aircraft.query.get(aircraft_id)

        # Höhe
        height_m_raw = (request.form.get("height_m") or "").strip()
        height_m = None
        if height_m_raw:
            try:
                height_m = int(height_m_raw)
            except Exception:
                height_m = None
        if height_m is None:
            height_m = int(getattr(ac, "default_height", 3000) or 3000)
        if height_m not in VALID_HEIGHTS:
            height_m = int(getattr(ac, "default_height", 3000) or 3000)

        # Max Payload
        max_payload_raw = request.form.get("max_payload_kg")
        max_payload_kg = parse_float(max_payload_raw)

        planned_date_raw = (request.form.get("planned_date") or "").strip()
        planned_start_time_raw = (request.form.get("planned_start_time") or "").strip()

        # Datum
        try:
            op_date = datetime.strptime(planned_date_raw, "%Y-%m-%d").date() if planned_date_raw else fallback_dt.date()
        except Exception:
            op_date = fallback_dt.date()

        # Uhrzeit
        time_raw = planned_start_time_raw or fallback_dt.strftime("%H:%M")
        try:
            hour, minute = map(int, time_raw.split(":"))
            planned_dt = datetime(op_date.year, op_date.month, op_date.day, hour, minute)
        except Exception:
            planned_dt = datetime(op_date.year, op_date.month, op_date.day, fallback_dt.hour, fallback_dt.minute)

        # ------------------------------------------------------------
        # ✅ PREISMATRIX-ABSICHERUNG (MANUELL)
        # ------------------------------------------------------------
        try:
            ensure_pricematrix_available_for_load(airfield_id, planned_dt.date())
        except ValueError as e:
            flash(str(e), "danger")
            return redirect(url_for("pricing.pricing_matrix"))

        # ✅ Nummerierung NUR nach Flugplatz + actual_time.date()
        next_number = next_load_number_for_day(
            airfield_id,
            planned_dt.date()
        )

        # ------------------------------------------------------------
        # ✅ D.2: PREISMODELL-FREEZE pro Load (MANUELL)
        # ------------------------------------------------------------
        pricing_model_id = None
        try:
            mid = session.get("active_pricing_model_id")
            pricing_model_id = int(mid) if mid is not None else None
        except Exception:
            pricing_model_id = None

        # Validierung: existieren Preise für dieses Session-Modell?
        if pricing_model_id is not None:
            exists_price = (
                db.session.query(BillingPrice.id)
                .filter(BillingPrice.period_id == pricing_model_id)
                .limit(1)
                .first()
            )
            if not exists_price:
                pricing_model_id = None

        # Fallback: bestes gültiges Modell am Datum ermitteln
        if pricing_model_id is None:
            row = (
                db.session.query(BillingPricePeriod.id)
                .join(BillingPrice, BillingPrice.period_id == BillingPricePeriod.id)
                .filter(BillingPricePeriod.valid_from <= planned_dt.date())
                .filter(
                    (BillingPricePeriod.valid_to.is_(None)) |
                    (BillingPricePeriod.valid_to >= planned_dt.date())
                )
                .order_by(BillingPricePeriod.valid_from.desc())
                .limit(1)
                .first()
            )
            if row:
                pricing_model_id = int(row[0])
                session["active_pricing_model_id"] = pricing_model_id

        if pricing_model_id is None:
            flash(
                "Kein aktives Preismodell gesetzt. Bitte zuerst unter „Preismatrix“ ein Preismodell aktiv setzen.",
                "danger"
            )
            return redirect(url_for("pricing.pricing_matrix"))

        load = Load(
            airfield_id=airfield_id,
            aircraft_id=aircraft_id,
            load_number=next_number,
            height_m=height_m,
            actual_time=planned_dt,          # ✅ EIN Zeitfeld
            max_payload_kg=max_payload_kg,
            status="open",
            pricing_model_id=pricing_model_id,   # ✅ NEU: Freeze Preismodell
        )
        db.session.add(load)
        db.session.commit()

        flash(f"Neuer Load {load.load_number} angelegt.", "success")

        if request.form.get("return_to") == "split":
            requested_show = request.form.get("show", "active")
            target_show = "active" if requested_show == "archive" else requested_show
            return redirect(url_for("load.split_view", edit=load.id, show=target_show, period="all"))
        return redirect(url_for("load.edit_load", id=load.id))

    # ------------------------------------------------------------
    # GET (manual=1): Formular anzeigen (KEIN Auto-Create!)
    # ------------------------------------------------------------
    # Vorschau für "nächste Nummer" (rein informativ)
    try:
        preview_next = next_load_number_for_day(default_airfield_id, fallback_dt.date())
    except Exception:
        preview_next = 1

    return render_template(
        "load/new.html",
        airfields=airfields,
        aircrafts=aircrafts,
        default_airfield_id=default_airfield_id,
        default_aircraft_id=default_aircraft_id,
        default_planned_date=default_planned_date,
        default_planned_start_time=default_planned_start_time,
        default_max_payload=default_max_payload,
        next_number=preview_next,
    )


# ============================================================
# REINDEX-HELPER — Load-Nummern für einen Betriebstag neu vergeben
# MUSS VOR den Route-Definitionen stehen
#
# Fachregel: EIN Zeitfeld -> actual_time
# - Tageszuordnung basiert auf actual_time (Fallback: created_at)
# - scheduled_time wird NICHT berücksichtigt
# ============================================================

def loads_query_for_operation_day(airfield_id: int, op_date: date):
    """
    Robust: Filter per Zeitbereich [day_start, day_end)
    Tageszuordnung:
    - actual_time bestimmt den Tag
    - Fallback nur, wenn actual_time None: created_at
    """
    day_start = datetime.combine(op_date, dtime.min)
    day_end = day_start + timedelta(days=1)

    return (
        Load.query
        .filter(Load.airfield_id == airfield_id)
        .filter(
            or_(
                # actual_time bestimmt den Tag
                (
                    Load.actual_time.isnot(None) &
                    (Load.actual_time >= day_start) &
                    (Load.actual_time < day_end)
                ),
                # Fallback: created_at nur wenn actual_time fehlt
                (
                    Load.actual_time.is_(None) &
                    (Load.created_at >= day_start) &
                    (Load.created_at < day_end)
                )
            )
        )
    )

def resequence_load_numbers_for_day(airfield_id: int, op_date: date) -> None:
    """
    Vergibt Load-Nummern für op_date neu: 1..N
    Sortierung:
    - actual_time (wenn vorhanden)
    - sonst created_at
    - None ans Ende
    """
    loads = loads_query_for_operation_day(airfield_id, op_date).all()
    if not loads:
        return

    def effective_time(l: Load) -> datetime:
        return l.actual_time or l.created_at or datetime.max

    loads_sorted = sorted(loads, key=lambda l: (effective_time(l), l.id or 0))

    # Zwei-Phasen-Update zur Kollisionsvermeidung
    tmp_base = 1000
    for i, l in enumerate(loads_sorted, start=1):
        l.load_number = tmp_base + i
    db.session.flush()

    for i, l in enumerate(loads_sorted, start=1):
        l.load_number = i
        
# ============================================================
# BLOCK 5 — LOAD EDITOR
# ============================================================
@bp_load.route("/<int:id>/edit", methods=["GET", "POST"])
def edit_load(id):
    load = Load.query.get_or_404(id)
    apply_follow_load_warning_single(load)

    reason = lock_reason(load)
    if reason:
        flash(reason, "danger")
        show = "archive" if load.status == "completed" else "active"
        return redirect(url_for("load.split_view", edit=load.id, show=show))

    if request.method == "POST":
        return redirect(url_for("load.save_load", id=id))

    airfields = (
        Flugplatz.query
        .filter(
            or_(
                (Flugplatz.deleted_at.is_(None) & Flugplatz.active.is_(True)),
                Flugplatz.id == load.airfield_id
            )
        )
        .order_by(Flugplatz.name.asc())
        .all()
    )

    aircrafts = (
        Aircraft.query
        .filter(or_(Aircraft.active.is_(True), Aircraft.id == load.aircraft_id))
        .order_by(Aircraft.type.asc(), Aircraft.registration.asc())
        .all()
    )

    return render_template(
        "load/editor.html",
        load=load,
        airfields=airfields,
        aircrafts=aircrafts,
    )

# ============================================================
# BLOCK 5a — STARTZEIT = JETZT (✈️ aus Liste / Split-View)
# ============================================================
@bp_load.route("/<int:id>/time_now", methods=["POST"], endpoint="time_now_action")
def time_now_action(id: int):
    load = Load.query.get_or_404(id)

    reason = lock_reason(load)
    if reason:
        flash(reason, "danger")
        return redirect(url_for("load.edit_load", id=id))

    now = now_local().replace(tzinfo=None)

    # Alter Tag (NUR actual_time, Fallback created_at)
    old_day = (load.actual_time or load.created_at).date()

    try:
        load.actual_time = now
        new_day = now.date()

        if new_day != old_day:
            if load.has_paid_entries or load.has_billed_entries or load.status == "completed":
                flash(
                    "Startzeit geändert, aber Load ist gesperrt "
                    "(bezahlt/abgerechnet/abgeschlossen). "
                    "Nummer bleibt unverändert.",
                    "warning"
                )
            else:
                # vor Reindex flushen, damit Queries den Load korrekt sehen
                db.session.flush()

                load.load_number = next_load_number_for_day(
                    load.airfield_id,
                    new_day,
                    exclude_load_id=load.id
                )

                resequence_load_numbers_for_day(load.airfield_id, old_day)
                resequence_load_numbers_for_day(load.airfield_id, new_day)

                flash(
                    f"Startzeit gesetzt. Load-Nummer angepasst: {load.load_number} "
                    f"(Betriebstag {new_day.strftime('%d.%m.%Y')}).",
                    "info"
                )
        else:
            flash(
                f"Startzeit für Load {load.load_number} auf {now.strftime('%H:%M')} gesetzt.",
                "success"
            )

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        flash(f"Fehler beim Setzen der Startzeit: {e}", "danger")
        return redirect(url_for("load.edit_load", id=id))

    if request.form.get("return_to") == "split":
        show = request.form.get("show", "active")
        period = request.form.get("period", "all")
        from_ = request.form.get("from", "")
        to_ = request.form.get("to", "")
        return redirect(url_for("load.split_view", edit=load.id, show=show, period=period, **({"from": from_} if from_ else {}), **({"to": to_} if to_ else {})))

    return redirect(url_for("load.edit_load", id=id))


# ============================================================
# BLOCK 5b — LOAD DURCHGEFÜHRT (Phase 2)
# ============================================================
@bp_load.route("/<int:id>/complete", methods=["POST"], endpoint="complete_load")
def complete_load(id: int):
    load = Load.query.get_or_404(id)

    show = request.form.get("show", "active")
    period = request.form.get("period", "all")
    from_ = request.form.get("from", "")
    to_ = request.form.get("to", "")

    def _range_params():
        p = {}
        if from_:
            p["from"] = from_
        if to_:
            p["to"] = to_
        return p

    reason = lock_reason(load)
    if reason:
        flash(reason, "danger")
        return redirect(url_for("load.split_view", edit=load.id, show=show, period=period, **_range_params()))

    if load.status == "completed":
        flash("Load ist bereits durchgeführt.", "info")
    else:
        validation_entries = [
            (
                int(getattr(e, "seat", 0) or 0),
                (getattr(e, "status_code", "") or "").strip(),
                int(getattr(e, "height_m", 0) or 0),
            )
            for e in (load.entries or [])
            if getattr(e, "person_id", None)
        ]
        ok_rules, rule_errors = validate_load_business_rules(load, validation_entries)
        if not ok_rules:
            deduped = []
            seen = set()
            for msg in rule_errors:
                key = (msg or "").strip()
                if not key or key in seen:
                    continue
                seen.add(key)
                deduped.append(key)
            flash(" | ".join(deduped), "danger")
            return redirect(url_for("load.split_view", edit=load.id, show=show, period=period, **_range_params()))

        load.status = "completed"
        try:
            db.session.commit()
            flash("Load als durchgeführt markiert.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Fehler beim Abschließen: {e}", "danger")
            return redirect(url_for("load.split_view", edit=load.id, show=show, period=period, **_range_params()))

    # Bei aktivem Filter ist der Load nach Abschluss im Archiv und soll rechts nicht offen bleiben.
    if show != "archive":
        return redirect(url_for("load.split_view", show=show, period=period, **_range_params()))

    return redirect(url_for("load.split_view", edit=load.id, show=show, period=period, **_range_params()))


# ============================================================
# BLOCK 5c — LOAD LÖSCHEN (Phase 3)
# ============================================================
@bp_load.route("/<int:id>/delete", methods=["POST"], endpoint="delete_load")
def delete_load(id: int):
    load = Load.query.get_or_404(id)
    show = request.args.get("show") or request.form.get("show") or "active"
    period = request.args.get("period") or request.form.get("period") or "all"
    from_ = request.args.get("from") or request.form.get("from") or ""
    to_ = request.args.get("to") or request.form.get("to") or ""

    def redirect_back():
        p = {}
        if from_:
            p["from"] = from_
        if to_:
            p["to"] = to_
        return redirect(url_for("load.split_view", show=show, period=period, **p))

    if load.has_paid_entries:
        flash("Dieser Load enthält bezahlte Einträge und darf nicht gelöscht werden.", "danger")
        return redirect_back()

    if load.has_linked_invoice_items:
        flash("Zu diesem Load existieren bereits Rechnungen. Bitte zuerst die Rechnungen löschen und danach den Load.", "danger")
        return redirect_back()

    if not is_admin():
        if load.status == "completed" or load.has_billed_entries:
            flash("Dieser Load darf nur vom Admin gelöscht werden.", "warning")
            return redirect_back()

    op_day = (load.actual_time or load.created_at).date()
    airfield_id = load.airfield_id

    try:
        for e in list(load.entries or []):
            db.session.delete(e)

        db.session.delete(load)
        db.session.commit()

        # ✅ Nach Löschen neu nummerieren (nur actual_time)
        resequence_load_numbers_for_day(airfield_id, op_day)

        flash("Load gelöscht.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Fehler beim Löschen: {e}", "danger")
        return redirect_back()

    return redirect_back()

# ============================================================
# BLOCK 6 — BUSINESS-REGELN
# ============================================================
def validate_load_business_rules(load, entries, capacity_override: Optional[int] = None):
    errors = []
    base_capacity = int(load.aircraft.seats or 0)
    max_supported_capacity = base_capacity + MAX_EXTRA_SEATS_PER_LOAD

    highest_used_seat = max(
        (
            int(seat)
            for seat, _, _ in entries
            if seat is not None
        ),
        default=0,
    )

    allowed_auffueller_statuses = {
        "Auffüller Verein",
        "Auffüller Gast",
        "Auffüller Partner-Verein",
    }

    invalid_auffueller_statuses = sorted({
        s for _, s, _ in entries
        if (s or "").strip().startswith("Auffüller") and s not in allowed_auffueller_statuses
    })
    if invalid_auffueller_statuses:
        errors.append(
            "Ungültiger Auffüller-Status: "
            + ", ".join(invalid_auffueller_statuses)
            + ". Erlaubt sind nur Auffüller Verein, Auffüller Gast, Auffüller Partner-Verein."
        )

    count_g_td = sum(1 for _, s, _ in entries if s == "G-TD")
    count_g_td_video = sum(1 for _, s, _ in entries if s == "G-TD-Video")
    count_td = sum(1 for _, s, _ in entries if s in ("TD", "TD-Vereins-Schirm"))
    count_video = sum(1 for _, s, _ in entries if s in ("Video", "Videomann"))
    count_teacher = sum(1 for _, s, _ in entries if s == "Lehrer")
    count_aff_teacher = sum(1 for _, s, _ in entries if _is_aff_teacher_status(s))
    count_aff_student_1 = sum(1 for _, s, _ in entries if _aff_student_level(s) == 1)
    count_aff_student_2 = sum(1 for _, s, _ in entries if _aff_student_level(s) == 2)
    count_students = sum(
        1 for _, s, _
        in entries
        if s in ("Schüler", "Schüler Ek 1", "Schüler Ek 2", "Schüler GK 6")
    )
    count_auffueller = sum(
        1 for _, s, _ in entries
        if s in allowed_auffueller_statuses
    )

    total_tandem_guests = count_g_td + count_g_td_video
    capacity = int(
        capacity_override
        if capacity_override is not None
        else max(base_capacity, min(max_supported_capacity, highest_used_seat))
    )

    invalid_seats = sorted({
        int(seat)
        for seat, _, _ in entries
        if seat is not None and int(seat) > capacity
    })
    if invalid_seats:
        errors.append(
            "Sitz außerhalb der erlaubten Kapazität: "
            + ", ".join(str(seat) for seat in invalid_seats)
            + f". Erlaubt sind maximal {capacity} Sitze für diesen Load."
        )

    # --------------------------------------------------------
    # Tandem-Regeln
    # --------------------------------------------------------
    if count_td > 0 and total_tandem_guests == 0:
        errors.append("TD/TD-Vereins-Schirm ohne Tandemgast ist nicht erlaubt.")

    if count_td != total_tandem_guests:
        errors.append(
            f"Anzahl TD ({count_td}) entspricht nicht Anzahl Tandemgäste ({total_tandem_guests})."
        )

    if count_g_td_video > 0 and count_video != count_g_td_video:
        errors.append(
            f"Für {count_g_td_video} G-TD-Video müssen genau "
            f"{count_g_td_video} Video-Springer vorhanden sein."
        )

    # Video darf nicht alleinstehend auftreten:
    # Für jeden Video-Springer muss es einen Tandemgast mit G-TD-Video
    # sowie einen Tandemmaster (TD/TD-Vereins-Schirm) im selben Load geben.
    if count_video > 0 and count_g_td_video == 0:
        errors.append(
            "Status Video ist nur zusammen mit mindestens einem G-TD-Video erlaubt."
        )

    if count_video > 0 and count_td == 0:
        errors.append(
            "Status Video ist nur zusammen mit mindestens einem TD/TD-Vereins-Schirm erlaubt."
        )

    # --------------------------------------------------------
    # Lehrer / Schüler (Grundregel)
    # --------------------------------------------------------
    if count_teacher > 0 and count_students == 0:
        errors.append("Ein Lehrer erfordert mindestens einen Schüler.")

    # --------------------------------------------------------
    # AFF-Lehrer / AFF-Schueler
    # - SCHUELER-AFF-1 braucht 1 AFF-Lehrer
    # - SCHUELER-AFF-2 braucht 2 AFF-Lehrer
    # --------------------------------------------------------
    required_aff_teachers = count_aff_student_1 + (2 * count_aff_student_2)

    if count_aff_teacher > 0 and required_aff_teachers == 0:
        errors.append("AFF-Lehrer erfordert mindestens einen AFF-Schüler (Schüler-AFF-1/2).")

    aff2_missing = count_aff_student_2 > 0 and count_aff_teacher < (2 * count_aff_student_2)
    if aff2_missing:
        errors.append("Schüler-AFF-2-Lehrer erfordert zwei AFF-Lehrer.")
    elif required_aff_teachers > 0 and count_aff_teacher < required_aff_teachers:
        errors.append(
            "Zu wenige AFF-Lehrer: "
            f"benötigt {required_aff_teachers}, vorhanden {count_aff_teacher}."
        )

    # AFF-1 darf nicht mit mehr AFF-Lehrern als erforderlich gespeichert werden.
    # (gewünschter Spezialfall: 1x Schüler-AFF-1 + 2x AFF-Lehrer => Fehler)
    if required_aff_teachers > 0 and count_aff_teacher > required_aff_teachers:
        if count_aff_student_2 == 0 and count_aff_student_1 > 0:
            errors.append("Schüler-AFF-1-Lehrer erfordert genau einen AFF-Lehrer.")
        else:
            errors.append(
                "Zu viele AFF-Lehrer: "
                f"benötigt {required_aff_teachers}, vorhanden {count_aff_teacher}."
            )

    # --------------------------------------------------------
    # Auffüller-Regeln
    # --------------------------------------------------------
    if count_auffueller > 1:
        errors.append("Es darf maximal ein Auffüller im Load sein.")

    if count_auffueller == 1:
        auff_seat = [seat for seat, s, _ in entries if s in allowed_auffueller_statuses][0]
        if auff_seat > base_capacity:
            errors.append("Auffüller darf keinen Extrasitz belegen.")

        has_tandem_block = total_tandem_guests > 0 and count_td > 0
        has_instruction_block = (
            (count_teacher > 0 and count_students > 0)
            or (count_aff_teacher > 0 and (count_aff_student_1 + count_aff_student_2) > 0)
        )
        if not (has_tandem_block or has_instruction_block):
            errors.append(
                "Auffüller ist nur erlaubt, wenn bereits ein Tandemblock "
                "und/oder ein Schüler/Lehrer-Block im Load vorhanden ist."
            )

        occupied_regular_seats = {
            seat for seat, _, _ in entries
            if 1 <= int(seat) <= base_capacity
        }
        if len(occupied_regular_seats) != base_capacity:
            errors.append(
                "Auffüller ist nur als letzter freier Platz erlaubt "
                "(mit Auffüller muss der Load voll sein)."
            )

        allowed_block_statuses = {
            "G-TD",
            "G-TD-Video",
            "TD",
            "TD-Vereins-Schirm",
            "Video",
            "Videomann",
            "Lehrer",
            "Schüler",
            "Schüler Ek 1",
            "Schüler Ek 2",
            "Schüler GK 6",
        }
        invalid_with_auffueller = sorted({
            s for _, s, _ in entries
            if s not in allowed_auffueller_statuses
            and s not in allowed_block_statuses
            and not _is_aff_teacher_status(s)
            and not _is_aff_student_status(s)
        })
        if invalid_with_auffueller:
            errors.append(
                "Mit Auffüller sind nur Tandem- und/oder Schüler/Lehrer-Blöcke erlaubt. "
                "Nicht erlaubt mit Auffüller: " + ", ".join(invalid_with_auffueller) + "."
            )

    # ============================================================
    # BLOCK 6.1 — Lehrer-Zuordnung per Sitznähe (max. 3 Sitze Abstand)
    #
    # Regel:
    # - Es reicht NICHT, dass irgendwo Schüler existieren.
    # - Jeder Lehrer muss mindestens einen Schüler haben,
    #   der höchstens 3 Sitze entfernt sitzt.
    # ============================================================
    MAX_INSTRUCTION_SEAT_DISTANCE = 3

    teacher_seats = sorted(
        seat for seat, s, _ in entries if s == "Lehrer"
    )
    student_seats = sorted(
        seat for seat, s, _
        in entries
        if s in ("Schüler", "Schüler Ek 1", "Schüler Ek 2", "Schüler GK 6")
    )

    # Nur prüfen, wenn Lehrer UND Schüler existieren
    # (teacher > 0 && students == 0 wird oben bereits abgefangen)
    if teacher_seats and student_seats:
        for t_seat in teacher_seats:
            nearest_ok = any(
                abs(t_seat - s_seat) <= MAX_INSTRUCTION_SEAT_DISTANCE
                for s_seat in student_seats
            )
            if not nearest_ok:
                errors.append(
                    f"Lehrer auf Sitz {t_seat} hat keinen Schüler in Reichweite "
                    f"(max. {MAX_INSTRUCTION_SEAT_DISTANCE} Sitze Abstand)."
                )

    return len(errors) == 0, errors

# ============================================================
# BLOCK 7 — LOAD SPEICHERN (ROBUST, KEIN DATENVERLUST, KEIN begin())
# Fachregel: ES GIBT NUR EIN ZEITFELD -> actual_time
# Nummerierung/Betriebstag basiert NUR auf actual_time.date()
# (Fallback: created_at.date() nur wenn actual_time None)
# ============================================================
@bp_load.route("/<int:id>/save", methods=["POST"])
def save_load(id):
    load = Load.query.get_or_404(id)

    is_split_view = request.form.get("return_to") == "split"
    show = request.form.get("show", "active")
    period = request.form.get("period", "all")
    from_ = request.form.get("from", "")
    to_ = request.form.get("to", "")

    def redirect_back():
        """Zur passenden Ansicht zurück (Split rechts oder Vollbild)."""
        if is_split_view:
            p = {}
            if from_:
                p["from"] = from_
            if to_:
                p["to"] = to_
            return redirect(url_for("load.split_view", edit=load.id, show=show, period=period, **p))
        return redirect(url_for("load.edit_load", id=load.id))

    reason = lock_reason(load)
    if reason:
        flash(reason, "danger")
        return redirect_back()

    def fail(msg: str, category: str = "danger"):
        db.session.rollback()
        flash(msg, category)
        return redirect_back()

    # --------------------------------------------------------
    # Prospektive Grunddaten (noch nicht final committen)
    # --------------------------------------------------------
    new_airfield_id = load.airfield_id
    new_aircraft_id = load.aircraft_id
    new_max_payload = load.max_payload_kg

    requested_load_height = parse_int(request.form.get("height_m"))
    if requested_load_height is None:
        requested_load_height = int(load.height_m or 3000)
    if requested_load_height not in VALID_HEIGHTS:
        return fail("Absprunghöhe (Load) ist ungültig.", "danger")


    # --------------------------------------------------------
    # ✅ Tankpause / Tanken erforderlich (reine Information)
    # --------------------------------------------------------
    # Formular sendet "0/1" (oder "on"); parse_bool ist robust.
    # Speichert nur den Zustand am Load. Keine weitere Logik.
    requested_fuel_required = parse_bool(request.form.get("fuel_required"))

    # ------------------------------------------------------------
    # STARTZEIT / BETRIEBSTAG (NUR actual_time)
    # ------------------------------------------------------------
    time_str = (request.form.get("actual_time_hm") or "").strip()
    date_str = (request.form.get("actual_date") or "").strip()

    old_day = (load.actual_time or load.created_at).date()
    old_airfield_id = int(load.airfield_id)

    new_actual_time = load.actual_time

    if date_str or time_str:
        try:
            base_dt = load.actual_time or load.created_at

            d = (
                datetime.strptime(date_str, "%Y-%m-%d").date()
                if date_str else base_dt.date()
            )
            t = (
                datetime.strptime(time_str, "%H:%M").time()
                if time_str else base_dt.time()
            )

            new_actual_time = datetime(d.year, d.month, d.day, t.hour, t.minute)
        except Exception:
            return fail("Datum/Uhrzeit im Editor ist ungültig.", "danger")

    new_day = (new_actual_time or load.created_at).date()
    day_or_airfield_changed = (
        new_day != old_day or int(new_airfield_id) != old_airfield_id
    )

    # ------------------------------------------------------------
    # Nummerierung bei Tages-/Flugplatzwechsel
    # ------------------------------------------------------------
    if day_or_airfield_changed:
        if load.has_paid_entries or load.has_billed_entries or load.status == "completed":
            flash(
                "Betriebstag oder Flugplatz geändert, aber Load ist gesperrt "
                "(bezahlt/abgerechnet/abgeschlossen). Nummer bleibt unverändert.",
                "warning",
            )
        else:
            load.airfield_id = new_airfield_id
            load.actual_time = new_actual_time
            db.session.flush()

            load.load_number = next_load_number_for_day(
                int(new_airfield_id),
                new_day,
                exclude_load_id=load.id,
            )

            resequence_load_numbers_for_day(old_airfield_id, old_day)
            resequence_load_numbers_for_day(int(new_airfield_id), new_day)

            flash(
                f"Load-Nummer wurde angepasst: {load.load_number} "
                f"(Betriebstag {new_day.strftime('%d.%m.%Y')}).",
                "info",
            )
    else:
        load.actual_time = new_actual_time

    # --------------------------------------------------------
    # Sitze & Einträge
    # --------------------------------------------------------
    existing_by_seat = {e.seat: e for e in load.entries if e.seat is not None}
    paid_seats = {e.seat for e in load.entries if getattr(e, "paid", False)}

    status_label_by_code = {
        (s.code or "").strip(): (s.label or s.code or "").strip()
        for s in StatusDefinition.query.filter(StatusDefinition.is_active.is_(True)).all()
    }
    for fallback_code in ("Auffüller Verein", "Auffüller Gast", "Auffüller Partner-Verein"):
        status_label_by_code.setdefault(fallback_code, fallback_code)

    desired = {}

    for seat in range(1, int(load.aircraft.seats or 0) + MAX_EXTRA_SEATS_PER_LOAD + 1):
        person_id = request.form.get(f"seat_{seat}_person")
        status_code = (request.form.get(f"seat_{seat}_status_code") or "").strip()
        height_raw = (request.form.get(f"seat_{seat}_height_m") or "").strip()

        gear_rental = parse_bool(
            request.form.get(f"seat_{seat}_gear_rental")
            or request.form.get(f"seat_{seat}_schirmmiete")
            or request.form.get(f"seat_{seat}_rental")
        )

        if not person_id:
            desired[seat] = None
            continue

        person = Person.query.get(int(person_id))
        if not person:
            return fail(f"Sitz {seat}: Ungültige Person.", "danger")

        if not status_code:
            return fail(f"Sitz {seat}: Status fehlt.", "danger")

        # Serverseitig hart absichern: Für Schüler/AFF-Schüler/Tandemgast
        # darf niemals Schirmmiete gesetzt bleiben.
        if _is_gear_rental_forbidden_status(status_code):
            gear_rental = False

        status_label = status_label_by_code.get(status_code, status_code)
        slot_for_person = _slot_for_person(person)
        if not _status_allowed_for_person_and_slot(status_code, status_label, person, slot_for_person):
            return fail(f"Sitz {seat}: Status passt nicht zur Person.", "danger")

        if not height_raw:
            return fail(f"Sitz {seat}: Absprunghöhe fehlt.", "danger")
        try:
            height_m = int(height_raw)
        except Exception:
            return fail(f"Sitz {seat}: Absprunghöhe ungültig.", "danger")
        if height_m not in VALID_HEIGHTS:
            return fail(f"Sitz {seat}: Unbekannte Absprunghöhe {height_m}.", "danger")

        desired[seat] = {
            "person": person,
            "status": status_code,
            "height": height_m,
            "gear_rental": gear_rental,
        }

    # --------------------------------------------------------
    # Geschäftsregeln vor dem Speichern prüfen
    # --------------------------------------------------------
    requested_extra_seats = max(0, min(MAX_EXTRA_SEATS_PER_LOAD, parse_int(request.form.get("extra_seats_ui")) or 0))

    validation_entries = [
        (seat, data["status"], data["height"])
        for seat, data in desired.items()
        if data is not None
    ]
    base_capacity = int(load.aircraft.seats or 0)
    highest_desired_seat = max((seat for seat, _, _ in validation_entries), default=0)
    desired_extra_seats = max(0, min(MAX_EXTRA_SEATS_PER_LOAD, highest_desired_seat - base_capacity))
    effective_capacity = base_capacity + max(requested_extra_seats, desired_extra_seats)

    ok_rules, rule_errors = validate_load_business_rules(
        load,
        validation_entries,
        capacity_override=effective_capacity,
    )
    if not ok_rules:
        # Doppelte Meldungen vermeiden, Reihenfolge beibehalten.
        deduped = []
        seen = set()
        for msg in rule_errors:
            key = (msg or "").strip()
            if not key or key in seen:
                continue
            seen.add(key)
            deduped.append(key)
        return fail(" | ".join(deduped), "danger")

    # --------------------------------------------------------
    # Jetzt schreiben wir wirklich
    # --------------------------------------------------------
    try:
        load.height_m = requested_load_height
        
        # ✅ Tankpause speichern (reine Information)
        load.fuel_required = bool(requested_fuel_required)

        for seat, new in desired.items():
            old = existing_by_seat.get(seat)

            if new is None:
                if old and seat not in paid_seats:
                    db.session.delete(old)
                continue

            if old:
                if seat in paid_seats:
                    continue
                old.person_id = new["person"].id
                old.status_code = new["status"]
                old.height_m = new["height"]
                if hasattr(old, "gear_rental"):
                    old.gear_rental = new["gear_rental"]
            else:
                entry = LoadEntry(
                    load_id=load.id,
                    person_id=new["person"].id,
                    seat=seat,
                    status_code=new["status"],
                    height_m=new["height"],
                )
                if hasattr(entry, "gear_rental"):
                    entry.gear_rental = new["gear_rental"]
                db.session.add(entry)

        db.session.commit()
        flash("Load gespeichert.", "success")
        return redirect_back()

    except Exception as e:
        db.session.rollback()
        flash(f"Fehler beim Speichern des Loads: {e}", "danger")
        return redirect_back()


# ============================================================
# BLOCK 10 — DEBUG / DIAGNOSE
# ============================================================
@bp_load.route("/debug")
def debug_loads():
    persons = Person.query.order_by(Person.full_name.asc()).all()
    loads = Load.query.order_by(Load.id.desc()).all()
    aircrafts = Aircraft.query.order_by(Aircraft.registration.asc()).all()
    airfields = Flugplatz.query.order_by(Flugplatz.name.asc()).all()
    status_defs = StatusDefinition.query.order_by(StatusDefinition.sort_order.asc()).all()
    return render_template(
        "load/debug.html",
        persons=persons,
        loads=loads,
        aircrafts=aircrafts,
        airfields=airfields,
        status_defs=status_defs,
    )

# ============================================================
# BLOCK 11 — STATISTIK (nur durchgeführte Loads)
# - Filterbar: Zeitraum, Flugplatz, Flugzeug, Person, Status, Schirmmiete, Orga, Tanken
# - Datenbasis: durchgeführte Loads (Load.status == "completed")
# - Sprungpreise: BillingService (rechnungsunabhängig)
# - Orga/Schirmmiete: getrennte Blöcke, aus Loads/Entries abgeleitet
#   (mit Tages-Cap-Logik für Schirmmiete, wenn max_count gesetzt ist)
# - Export: CSV + Excel (Pivot-fähig, Flat-Table mit row_type/item_type)
# ============================================================

from datetime import datetime, date, time as dtime
from decimal import Decimal
from io import BytesIO
import csv
from typing import Optional

from sqlalchemy.orm import joinedload, selectinload


def _build_stats_query(args):
    """
    Query für completed Loads + optionale Filter.
    WICHTIG:
    - Entry-Filter werden per Join angewendet, aber Datenbasis bleibt Load.
    - Kein InvoiceItem Join in der Basis-Query (Statistik rechnungsunabhängig).
    """
    q = (
        Load.query
        .filter(Load.status == "completed")
        .options(
            joinedload(Load.airfield),
            joinedload(Load.aircraft),
            # entries + person in einem Rutsch (verhindert N+1)
            selectinload(Load.entries).selectinload(LoadEntry.person),
        )
    )

    # Zeitraum (effective time = actual_time else created_at)
    from_d = _parse_date_ymd(args.get("from"))
    to_d = _parse_date_ymd(args.get("to"))
    dt_from, dt_to = _dt_range(from_d, to_d)

    if dt_from or dt_to:
        eff = db.func.coalesce(Load.actual_time, Load.created_at)
        if dt_from:
            q = q.filter(eff >= dt_from)
        if dt_to:
            q = q.filter(eff <= dt_to)

    # Flugplatz / Flugzeug
    airfield_id = parse_int(args.get("airfield_id"))
    if airfield_id:
        q = q.filter(Load.airfield_id == airfield_id)

    aircraft_id = parse_int(args.get("aircraft_id"))
    if aircraft_id:
        q = q.filter(Load.aircraft_id == aircraft_id)

    # Tanken
    tank = (args.get("tanken") or "").strip().lower()
    if tank in ("1", "true", "on", "ja", "yes"):
        q = q.filter(Load.fuel_required.is_(True))
    elif tank in ("0", "false", "off", "nein", "no"):
        q = q.filter(Load.fuel_required.is_(False))

    # Entry-basierte Filter (join nur wenn gebraucht)
    person_id = parse_int(args.get("person_id"))

    if hasattr(args, "getlist"):
        status_codes = [(s or "").strip() for s in args.getlist("status_code") if (s or "").strip()]
    else:
        single = (args.get("status_code") or "").strip()
        status_codes = [single] if single else []

    gear_raw = (args.get("gear_rental") or "").strip().lower()
    gear_filter_active = bool(gear_raw)
    want_gear = gear_raw in ("1", "true", "on", "ja", "yes")

    if person_id or status_codes or gear_filter_active:
        q = q.join(LoadEntry, LoadEntry.load_id == Load.id)
        if person_id:
            q = q.filter(LoadEntry.person_id == person_id)
        if status_codes:
            q = q.filter(LoadEntry.status_code.in_(status_codes))
        if gear_filter_active:
            q = q.filter(LoadEntry.gear_rental.is_(True if want_gear else False))

    # Orga-Filter: als Load-Filter (berechnete Orga) wird später angewendet,
    # damit wir keine Rechnungs-Abhängigkeit in die Basis-Query bringen.

    q = q.distinct().order_by(Load.actual_time.desc().nullslast(), Load.created_at.desc())
    return q


def _make_entry_matcher(args):
    """Predicate: nur Entries, die den aktuellen Entry-Filtern entsprechen."""
    person_id = parse_int(args.get("person_id"))

    if hasattr(args, "getlist"):
        status_codes = [(s or "").strip() for s in args.getlist("status_code") if (s or "").strip()]
    else:
        single = (args.get("status_code") or "").strip()
        status_codes = [single] if single else []

    gear_raw = (args.get("gear_rental") or "").strip().lower()
    gear_filter_active = bool(gear_raw)
    want_gear = gear_raw in ("1", "true", "on", "ja", "yes")

    def entry_matches(e: LoadEntry) -> bool:
        if person_id and int(getattr(e, "person_id", 0) or 0) != int(person_id):
            return False
        if status_codes:
            sc = (getattr(e, "status_code", "") or "").strip()
            if sc not in status_codes:
                return False
        if gear_filter_active:
            if bool(getattr(e, "gear_rental", False)) != bool(want_gear):
                return False
        return True

    return entry_matches


def _compute_extras_from_completed_loads(loads: list[Load], args):
    """
    Berechnet Zusatzpositionen getrennt:
      - Schirmmiete (mit Tages-Cap, wenn max_count gesetzt) → PRO PERSON
      - Orga (optional pro Tag, wenn BillingService 'mode' dies so signalisiert) → PRO PERSON
    Datenquelle: Loads/Entries (Statistik bleibt rechnungsunabhängig).
    
    WICHTIG: Schirmmiete-Cap wird pro Person angewendet, nicht global über alle Personen.
    Dies matched die Abrechnung-Logik (billing.py Zeile 2176-2200).

    Rückgabe:
      (filtered_loads, rental_items, orga_items,
       rental_sum_net, rental_sum_vat, rental_sum_gross,
       orga_sum_net, orga_sum_vat, orga_sum_gross)
    """
    entry_matches = _make_entry_matcher(args)

    # Deterministische Reihenfolge für Loads (MUSS VOR DER ENTRY-SAMMLUNG SEIN)
    loads_sorted = sorted(
        loads,
        key=lambda l: (getattr(l, "actual_time", None) or getattr(l, "created_at", None) or datetime.min)
    )

    # Sammle alle Entries und gruppiere nach Person
    all_entries: list[LoadEntry] = []
    by_person = defaultdict(list)
    
    for l in loads_sorted:
        for e in (getattr(l, "entries", None) or []):
            all_entries.append(e)
            by_person[e.person_id].append(e)

    # Berechne Zusatzpositionen PRO PERSON (wie Abrechnung)
    all_rental_items = []
    all_orga_items = []
    total_rental_sum_net = Decimal("0.00")
    total_rental_sum_vat = Decimal("0.00")
    total_rental_sum_gross = Decimal("0.00")
    total_orga_sum_net = Decimal("0.00")
    total_orga_sum_vat = Decimal("0.00")
    total_orga_sum_gross = Decimal("0.00")
    
    for person_id, person_entries in by_person.items():
        if not person_entries:
            continue
        
        # Filtere person_entries durch entry_matches
        filtered_entries = [e for e in person_entries if entry_matches(e)]
        
        if not filtered_entries:
            continue
        
        # Rufe compute_extras_for_entries PRO PERSON auf (Schirmmiete-Cap wird pro Person angewendet)
        extras = BillingService.compute_extras_for_entries(
            filtered_entries,
            entry_matches=None,  # Bereits gefiltert oben
            include_rental_items=True,
            include_orga_items=True,
        )
        
        # Aggregiere die Results
        all_rental_items.extend(extras["rental_items"])
        all_orga_items.extend(extras["orga_items"])
        total_rental_sum_net += Decimal(str(extras.get("rental_sum_net") or "0.00"))
        total_rental_sum_vat += Decimal(str(extras.get("rental_sum_vat") or "0.00"))
        total_rental_sum_gross += Decimal(str(extras.get("rental_sum_gross") or "0.00"))
        total_orga_sum_net += Decimal(str(extras.get("orga_sum_net") or "0.00"))
        total_orga_sum_vat += Decimal(str(extras.get("orga_sum_vat") or "0.00"))
        total_orga_sum_gross += Decimal(str(extras.get("orga_sum_gross") or "0.00"))
    
    rental_items = all_rental_items
    orga_items = all_orga_items
    rental_sum_net = total_rental_sum_net
    rental_sum_vat = total_rental_sum_vat
    rental_sum_gross = total_rental_sum_gross
    orga_sum_net = total_orga_sum_net
    orga_sum_vat = total_orga_sum_vat
    orga_sum_gross = total_orga_sum_gross

    return (
        loads_sorted,
        rental_items, orga_items,
        rental_sum_net, rental_sum_vat, rental_sum_gross,
        orga_sum_net, orga_sum_vat, orga_sum_gross
    )


def _invoice_payment_state_code(
    is_paid: bool | None,
    payment_method: str | None,
    payment_state: str | None,
) -> str:
    if bool(is_paid):
        return INVOICE_PAYMENT_STATE_PAID

    raw_state = (payment_state or "").strip().lower()
    if raw_state in INVOICE_PAYMENT_STATES:
        if raw_state == INVOICE_PAYMENT_STATE_PAID:
            return INVOICE_PAYMENT_STATE_OPEN
        return raw_state

    if (payment_method or "").strip().lower() == "sepa":
        return INVOICE_PAYMENT_STATE_SEPA_PENDING
    return INVOICE_PAYMENT_STATE_OPEN


def _invoice_status_label(
    is_paid: bool | None,
    payment_method: str | None,
    payment_state: str | None,
) -> str:
    state = _invoice_payment_state_code(is_paid, payment_method, payment_state)
    mapping = {
        INVOICE_PAYMENT_STATE_OPEN: "offen",
        INVOICE_PAYMENT_STATE_SEPA_PENDING: "sepa vorgemerkt",
        INVOICE_PAYMENT_STATE_SEPA_EXPORTED: "sepa exportiert",
        INVOICE_PAYMENT_STATE_PAID: "bezahlt",
        INVOICE_PAYMENT_STATE_SEPA_RETURNED: "ruecklastschrift",
    }
    return mapping.get(state, "offen")


def _invoice_number_missing_label() -> str:
    return "noch nicht abgerechnet"


def _invoice_number_label(created_at, invoice_id) -> str:
    try:
        if not invoice_id:
            return _invoice_number_missing_label()
        year = created_at.strftime("%Y") if created_at else ""
        return f"{year}-Spruenge #{int(invoice_id)}" if year else str(int(invoice_id))
    except Exception:
        return str(invoice_id or _invoice_number_missing_label())


def _invoice_payment_split_label(payment_method: str | None, prepaid_voucher_amount) -> str:
    onsite_label = _invoice_payment_label(payment_method) if payment_method else ""
    try:
        prepaid = Decimal(str(prepaid_voucher_amount or "0"))
    except Exception:
        prepaid = Decimal("0.00")
    if prepaid > 0:
        if onsite_label:
            return f"{onsite_label} + Vorkasse / Gutschein"
        return "Vorkasse / Gutschein"
    return onsite_label or "—"


def _manual_status_codes_for_person(person: Person | None) -> set[str]:
    codes: set[str] = set()
    if not person:
        return codes

    if bool(getattr(person, "is_member", False)):
        codes.add("Verein")
    elif bool(getattr(person, "is_partner_verein", False)):
        codes.add("Partner-Verein")
    elif bool(getattr(person, "is_tandem_guest", False)):
        codes.add("G-TD")
    else:
        codes.add("Gast")

    if bool(getattr(person, "is_student", False)):
        codes.add("Schüler")
    if bool(getattr(person, "is_aff_student", False)):
        codes.add("Schueler-Aff-1")
        codes.add("Schüler")
    if bool(getattr(person, "is_teacher", False)):
        codes.add("Lehrer")
    if bool(getattr(person, "is_aff_teacher", False)):
        codes.add("Aff-Lehrer")
    if bool(getattr(person, "is_video", False)):
        codes.add("Video")
    if bool(getattr(person, "is_tandemmaster", False)):
        codes.add("TD")

    return codes


def _manual_invoice_number_label(invoice: Invoice) -> str:
    year = (getattr(invoice, "created_at", None) or datetime.now()).strftime("%Y")
    number = getattr(invoice, "seq_number", None)
    if number is None:
        number = getattr(invoice, "id", None)
    try:
        return f"{year}-manuell #{int(number)}"
    except Exception:
        return _invoice_number_missing_label()


def _build_manual_invoice_item_rows(args) -> list[dict]:
    from_d = _parse_date_ymd(args.get("from"))
    to_d = _parse_date_ymd(args.get("to"))
    dt_from, dt_to = _dt_range(from_d, to_d)

    person_id = parse_int(args.get("person_id"))
    if hasattr(args, "getlist"):
        status_filters = {
            (s or "").strip()
            for s in args.getlist("status_code")
            if (s or "").strip()
        }
    else:
        single_status = (args.get("status_code") or "").strip()
        status_filters = {single_status} if single_status else set()

    # Statistik bleibt load-zentriert bei Flugplatz/Flugzeug/Tanken/Schirmmiete/Orga.
    # Manuelle Rechnungspositionen ignorieren diese filterfremden Parameter bewusst.
    q = (
        Invoice.query
        .options(joinedload(Invoice.person), selectinload(Invoice.items))
        .join(InvoiceItem, InvoiceItem.invoice_id == Invoice.id)
        .filter(
            Invoice.is_deleted.is_(False),
            Invoice.stage == "final",
            InvoiceItem.item_source == "manual",
        )
        .distinct()
    )

    rows: list[dict] = []
    for inv in q.all():
        person = getattr(inv, "person", None)
        if person_id and int(getattr(inv, "person_id", 0) or 0) != int(person_id):
            continue

        status_codes = _manual_status_codes_for_person(person)
        if status_filters and not any(code in status_filters for code in status_codes):
            continue

        eff_dt = getattr(inv, "created_at", None)
        if getattr(inv, "service_date", None):
            eff_dt = datetime.combine(inv.service_date, dtime.min)

        if dt_from and (not eff_dt or eff_dt < dt_from):
            continue
        if dt_to and (not eff_dt or eff_dt > dt_to):
            continue

        invoice_status = _invoice_status_label(
            getattr(inv, "is_paid", False),
            getattr(inv, "payment_method", None),
            getattr(inv, "payment_state", None),
        )
        payment_label = _invoice_payment_split_label(
            getattr(inv, "payment_method", None),
            getattr(inv, "prepaid_voucher_amount", None),
        )
        invoice_label = _manual_invoice_number_label(inv)
        person_name = getattr(person, "full_name", "") if person else ""
        status_label = ", ".join(sorted(status_codes))

        for item in list(getattr(inv, "items", []) or []):
            if (getattr(item, "item_source", "") or "").strip().lower() != "manual":
                continue

            gross = Decimal(str(getattr(item, "amount", 0) or "0.00"))
            net = Decimal(str(getattr(item, "net_amount", 0) or "0.00"))
            vat = Decimal(str(getattr(item, "vat_amount", 0) or "0.00"))
            vat_rate = Decimal(str(getattr(item, "vat_rate", 0) or "0.00"))
            quantity = Decimal(str(getattr(item, "quantity", 1) or "1.00"))
            unit_price_gross = Decimal(str(getattr(item, "unit_price_gross", 0) or "0.00"))
            manual_unit = (getattr(item, "manual_unit", "") or "").strip()

            rows.append({
                "row_type": "manual_invoice",
                "item_type": "Manuelle Rechnung",
                "item_desc": (getattr(item, "description", "") or getattr(inv, "manual_title", "") or "Manuelle Position").strip(),
                "invoice_title": (getattr(inv, "manual_title", "") or "Manuelle Positionen").strip(),
                "load_id": "",
                "load_number": invoice_label,
                "load_date": eff_dt.strftime("%d.%m.%Y") if eff_dt else "",
                "load_time": eff_dt.strftime("%H:%M") if eff_dt else "",
                "airfield": "",
                "aircraft": "",
                "height_m": "",
                "fuel_required": "",
                "entry_id": "",
                "person_id": getattr(inv, "person_id", "") or "",
                "person_name": person_name,
                "status_code": status_label,
                "status_codes": sorted(status_codes),
                "effective_dt": eff_dt,
                "entry_height_m": "",
                "gear_rental": "",
                "invoice_status": invoice_status,
                "payment_method": payment_label,
                "invoice_number": invoice_label,
                "gross": str(gross),
                "net": str(net),
                "vat": str(vat),
                "vat_rate": str(vat_rate),
                "quantity": str(quantity),
                "manual_unit": manual_unit,
                "unit_price_gross": str(unit_price_gross),
                "status_css_class": "",
                "status_bg": "",
                "status_border": "",
            })

    return rows


def _fmt_money_de(value, decimals: int = 2) -> str:
    """Formatiert Zahlenwerte als deutschen Geld-String (z.B. 1.234,56)."""
    try:
        d = Decimal(str(value or "0"))
    except Exception:
        d = Decimal("0.00")
    s = f"{d:.{decimals}f}"
    return s.replace(".", ",")


def _build_invoice_info_by_entry(entry_ids: list[int]) -> dict[int, dict[str, object]]:
    """Liefert je LoadEntry die zuletzt erstellte, nicht stornierte Rechnung inkl. Snapshot-Betraege."""
    invoice_info_by_entry: dict[int, dict[str, object]] = {}
    if not entry_ids:
        return invoice_info_by_entry

    invoice_rows = (
        db.session.query(
            InvoiceItem.load_entry_id,
            InvoiceItem.amount,
            InvoiceItem.net_amount,
            InvoiceItem.vat_amount,
            InvoiceItem.vat_rate,
            Invoice.is_paid,
            Invoice.payment_method,
            Invoice.payment_state,
            Invoice.prepaid_voucher_amount,
            Invoice.created_at,
            Invoice.id,
        )
        .join(Invoice, Invoice.id == InvoiceItem.invoice_id)
        .filter(
            InvoiceItem.load_entry_id.in_(entry_ids),
            Invoice.is_deleted.is_(False),
            Invoice.stage == "final",
        )
        .order_by(
            InvoiceItem.load_entry_id.asc(),
            Invoice.created_at.desc(),
            Invoice.id.desc(),
        )
        .all()
    )

    for (
        load_entry_id,
        item_gross,
        item_net,
        item_vat,
        item_vat_rate,
        is_paid,
        payment_method,
        payment_state,
        prepaid_voucher_amount,
        created_at,
        invoice_id,
    ) in invoice_rows:
        le_id = int(load_entry_id or 0)
        if le_id <= 0 or le_id in invoice_info_by_entry:
            continue
        invoice_info_by_entry[le_id] = {
            "invoice_status": _invoice_status_label(is_paid, payment_method, payment_state),
            "payment_method": _invoice_payment_split_label(payment_method, prepaid_voucher_amount),
            "invoice_number": _invoice_number_label(created_at, invoice_id),
            "gross": Decimal(str(item_gross or "0.00")),
            "net": Decimal(str(item_net or "0.00")),
            "vat": Decimal(str(item_vat or "0.00")),
            "vat_rate": Decimal(str(item_vat_rate or "0.00")),
        }

    return invoice_info_by_entry


def _build_invoice_info_by_person(person_ids: list[int]) -> dict[int, dict[str, str]]:
    """Liefert je Person die neueste abgeschlossene, nicht stornierte Rechnung."""
    invoice_info: dict[int, dict[str, str]] = {}
    if not person_ids:
        return invoice_info

    rows = (
        db.session.query(
            Invoice.person_id,
            Invoice.is_paid,
            Invoice.payment_method,
            Invoice.payment_state,
            Invoice.prepaid_voucher_amount,
            Invoice.created_at,
            Invoice.id,
        )
        .filter(
            Invoice.person_id.in_(person_ids),
            Invoice.is_deleted.is_(False),
            Invoice.seq_number.isnot(None),
        )
        .order_by(
            Invoice.person_id.asc(),
            Invoice.created_at.desc(),
            Invoice.id.desc(),
        )
        .all()
    )

    for person_id, is_paid, payment_method, payment_state, prepaid_voucher_amount, created_at, invoice_id in rows:
        pid = int(person_id or 0)
        if pid <= 0 or pid in invoice_info:
            continue
        invoice_info[pid] = {
            "invoice_status": _invoice_status_label(is_paid, payment_method, payment_state),
            "payment_method": _invoice_payment_split_label(payment_method, prepaid_voucher_amount),
            "invoice_number": _invoice_number_label(created_at, invoice_id),
        }

    return invoice_info


def _iter_export_rows(loads: list[Load], args):
    """
    Pivot-fähige Flat-Table:
    - 1 Zeile pro MATCHING Entry (item_type="Sprung", row_type="entry")
    - plus Extra-Zeilen:
        item_type="Schirmmiete" / "Orga", row_type="extra"
    """
    entry_matches = _make_entry_matcher(args)

    (
        loads2,
        rental_items, orga_items,
        rental_sum_net, rental_sum_vat, rental_sum_gross,
        orga_sum_net, orga_sum_vat, orga_sum_gross
    ) = _compute_extras_from_completed_loads(loads, args)

    entry_ids = [
        int(getattr(e, "id", 0) or 0)
        for l in loads2
        for e in (getattr(l, "entries", None) or [])
        if int(getattr(e, "id", 0) or 0) > 0
    ]
    invoice_info_by_entry = _build_invoice_info_by_entry(entry_ids)

    # Person-Status-Lookup für Extra-Zeilen (erster bekannter Status pro Person)
    person_status_code: dict[int, str] = {}
    for l in loads2:
        for e in (getattr(l, "entries", None) or []):
            pid = int(getattr(e, "person_id", 0) or 0)
            if pid > 0 and pid not in person_status_code:
                sc = (getattr(e, "status_code", "") or "").strip()
                if sc:
                    person_status_code[pid] = sc

    # Invoice-Lookup für Extra-Zeilen (neueste Rechnung pro Person)
    extra_person_ids = list({
        int(x.get("person_id") or 0)
        for x in (rental_items + orga_items)
        if int(x.get("person_id") or 0) > 0
    })
    invoice_info_by_person = _build_invoice_info_by_person(extra_person_ids)

    # 1) Sprung-Zeilen
    for l in loads2:
        eff_dt = getattr(l, "actual_time", None) or getattr(l, "created_at", None)
        load_date = eff_dt.strftime("%d.%m.%Y") if eff_dt else ""
        load_time = eff_dt.strftime("%H:%M") if eff_dt else ""

        for e in (getattr(l, "entries", None) or []):
            if not entry_matches(e):
                continue

            person = getattr(e, "person", None)
            person_name = getattr(person, "full_name", "") if person else ""
            entry_invoice_info = invoice_info_by_entry.get(int(getattr(e, "id", 0) or 0), {})

            gross = net = vat = vat_rate = Decimal("0.00")
            if entry_invoice_info:
                gross = _money(entry_invoice_info.get("gross") or "0.00")
                net = _money(entry_invoice_info.get("net") or "0.00")
                vat = _money(entry_invoice_info.get("vat") or "0.00")
                vat_rate = _money(entry_invoice_info.get("vat_rate") or "0.00")
            else:
                try:
                    gross = _money(BillingService.calculate_price_for_entry(e))
                    vat_rate = _money(BillingService.get_entry_vat_rate(e))
                    net, vat = BillingService.split_gross_into_net_and_vat(gross=gross, vat_rate=vat_rate)
                except Exception:
                    pass

            status_css_class = _archive_entry_css_class(l, e)
            status_bg, status_border = _archive_colors_from_css_class(status_css_class)

            yield {
                "row_type": "entry",
                "item_type": "Sprung",
                "item_desc": (
                    "Sprung - Kein Umsatzsteuerausweis gemäß § 19 UStG"
                    if (vat_rate == Decimal("0.00") and (getattr(e, "status_code", "") or "") in {"TD", "TD-Vereins-Schirm"})
                    else "Sprung"
                ),
                "load_id": getattr(l, "id", ""),
                "load_number": getattr(l, "load_number", ""),
                "load_date": load_date,
                "load_time": load_time,
                "airfield": getattr(getattr(l, "airfield", None), "name", ""),
                "aircraft": f"{getattr(getattr(l, 'aircraft', None), 'type', '')} {getattr(getattr(l, 'aircraft', None), 'registration', '')}".strip(),
                "height_m": getattr(l, "height_m", ""),
                "fuel_required": 1 if getattr(l, "fuel_required", False) else 0,
                "entry_id": getattr(e, "id", ""),
                "person_id": getattr(e, "person_id", ""),
                "person_name": person_name,
                "status_code": getattr(e, "status_code", ""),
                "entry_height_m": getattr(e, "height_m", ""),
                "gear_rental": 1 if getattr(e, "gear_rental", False) else 0,
                "invoice_status": entry_invoice_info.get("invoice_status", "offen"),
                "payment_method": entry_invoice_info.get("payment_method", ""),
                "invoice_number": entry_invoice_info.get("invoice_number", _invoice_number_missing_label()),
                "gross": str(gross),
                "net": str(_money(net)),
                "vat": str(_money(vat)),
                "vat_rate": str(vat_rate),
                "status_css_class": status_css_class,
                "status_bg": status_bg,
                "status_border": status_border,
            }

    # 2) Extra-Zeilen (Schirmmiete + Orga)
    for x in (rental_items + orga_items):
        dt = x.get("date")
        load_date = dt.strftime("%d.%m.%Y") if dt else ""
        load_time = x.get("time") or (dt.strftime("%H:%M") if dt else "")
        _xpid = int(x.get("person_id") or 0)
        _xinv = invoice_info_by_person.get(_xpid, {})

        yield {
            "row_type": "extra",
            "item_type": x.get("item_type", ""),
            "item_desc": x.get("desc", ""),
            "load_id": x.get("load_id", ""),
            "load_number": x.get("load_number", ""),
            "load_date": load_date,
            "load_time": load_time,
            "airfield": x.get("airfield", ""),
            "aircraft": "",
            "height_m": "",
            "fuel_required": "",
            "entry_id": "",
            "person_id": x.get("person_id", ""),
            "person_name": x.get("person_name", ""),
            "status_code": person_status_code.get(_xpid, ""),
            "entry_height_m": "",
            "gear_rental": "",
            "invoice_status": _xinv.get("invoice_status", ""),
            "payment_method": _xinv.get("payment_method", ""),
            "invoice_number": _xinv.get("invoice_number", ""),
            "gross": str(x.get("gross", Decimal("0.00"))),
            "net": str(x.get("net", Decimal("0.00"))),
            "vat": str(x.get("vat", Decimal("0.00"))),
            "vat_rate": str(x.get("vat_rate", Decimal("0.00"))),
            "status_css_class": "",
            "status_bg": "",
            "status_border": "",
        }

    # 3) Manuelle Rechnungspositionen (rechnungsbasiert)
    for row in _build_manual_invoice_item_rows(args):
        yield row

def _build_statistics_context(args):
    print_filter_summary: list[str] = []

    def _arg_checked(name: str, default: bool) -> bool:
        raw = args.get(name)
        if raw is None:
            return default
        return str(raw).strip().lower() in {"1", "true", "yes", "on", "ja"}

    show_col_net = _arg_checked("show_net", False)
    show_col_vat = _arg_checked("show_vat", False)
    show_col_vatrate = _arg_checked("show_vatrate", False)
    show_col_payment = _arg_checked("show_payment", False)
    show_col_invoice_number = _arg_checked("show_invoice_number", False)
    show_col_gross = _arg_checked("show_gross", True)

    def _fmt_filter_date(value: str) -> str:
        try:
            return datetime.strptime(value, "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            return value

    airfields = (
        Flugplatz.query
        .filter(Flugplatz.deleted_at.is_(None))
        .order_by(Flugplatz.name.asc())
        .all()
    )
    aircrafts = (
        Aircraft.query
        .order_by(Aircraft.type.asc(), Aircraft.registration.asc())
        .all()
    )
    persons = (
        Person.query
        .filter(Person.deleted_at.is_(None))
        .order_by(Person.last_name.asc(), Person.first_name.asc())
        .all()
    )
    statuses = (
        StatusDefinition.query
        .filter(StatusDefinition.is_active.is_(True))
        .order_by(StatusDefinition.sort_order.asc(), StatusDefinition.code.asc())
        .all()
    )

    date_from = (args.get("from") or "").strip()
    date_to = (args.get("to") or "").strip()
    if date_from or date_to:
        pretty_from = _fmt_filter_date(date_from) if date_from else ""
        pretty_to = _fmt_filter_date(date_to) if date_to else ""
        if date_from and date_to:
            print_filter_summary.append(f"Zeitraum: {pretty_from} bis {pretty_to}")
        elif date_from:
            print_filter_summary.append(f"Zeitraum ab {pretty_from}")
        else:
            print_filter_summary.append(f"Zeitraum bis {pretty_to}")

    selected_airfield_id = (args.get("airfield_id") or "").strip()
    if selected_airfield_id:
        selected_airfield = next((af.name for af in airfields if str(af.id) == selected_airfield_id), None)
        if selected_airfield:
            print_filter_summary.append(f"Flugplatz: {selected_airfield}")

    selected_aircraft_id = (args.get("aircraft_id") or "").strip()
    if selected_aircraft_id:
        selected_aircraft = next(
            (
                f"{ac.type} - {ac.registration}".strip(" -")
                for ac in aircrafts
                if str(ac.id) == selected_aircraft_id
            ),
            None,
        )
        if selected_aircraft:
            print_filter_summary.append(f"Flugzeug: {selected_aircraft}")

    selected_person_id = (args.get("person_id") or "").strip()
    if selected_person_id:
        selected_person = next((p.full_name for p in persons if str(p.id) == selected_person_id), None)
        if selected_person:
            print_filter_summary.append(f"Person: {selected_person}")

    selected_statuses = [s.code for s in statuses if s.code in args.getlist("status_code")]
    if selected_statuses:
        print_filter_summary.append(f"Status: {', '.join(selected_statuses)}")

    if args.get("tanken") == "1":
        print_filter_summary.append("Tanken: Ja")
    elif args.get("tanken") == "0":
        print_filter_summary.append("Tanken: Nein")

    if args.get("gear_rental") == "1":
        print_filter_summary.append("Schirmmiete: Ja")
    elif args.get("gear_rental") == "0":
        print_filter_summary.append("Schirmmiete: Nein")

    if args.get("orga") == "1":
        print_filter_summary.append("Orga: Ja")

    loads = _build_stats_query(args).all()

    (
        loads,
        extra_rental_items, extra_orga_items,
        extra_rental_sum_net, extra_rental_sum_vat, extra_rental_sum_gross,
        extra_orga_sum_net, extra_orga_sum_vat, extra_orga_sum_gross
    ) = _compute_extras_from_completed_loads(loads, args)

    manual_invoice_rows = _build_manual_invoice_item_rows(args)
    manual_invoice_count = len({r.get("invoice_number") for r in manual_invoice_rows if r.get("invoice_number")})
    manual_item_count = len(manual_invoice_rows)
    manual_sum_gross = sum(Decimal(str(r.get("gross") or "0.00")) for r in manual_invoice_rows)
    manual_sum_net = sum(Decimal(str(r.get("net") or "0.00")) for r in manual_invoice_rows)
    manual_sum_vat = sum(Decimal(str(r.get("vat") or "0.00")) for r in manual_invoice_rows)

    stats_entry_ids = [
        int(getattr(e, "id", 0) or 0)
        for l in loads
        for e in (getattr(l, "entries", None) or [])
        if int(getattr(e, "id", 0) or 0) > 0
    ]
    invoice_info_by_entry = _build_invoice_info_by_entry(stats_entry_ids)

    entry_matches = _make_entry_matcher(args)

    total_loads = len(loads)
    total_entries = 0
    tank_loads = 0
    gear_entries = 0

    gross_sum_est = Decimal("0.00")
    net_sum_est = Decimal("0.00")
    vat_sum_est = Decimal("0.00")

    # Sprung-Summen (nur Sprungpreise, Orga/Schirmmiete kommen später dazu)

    status_counts: dict[str, int] = {}
    day_counts: dict[str, int] = {}
    per_load = {}

    for l in loads:
        if getattr(l, "fuel_required", False):
            tank_loads += 1

        eff_dt = getattr(l, "actual_time", None) or getattr(l, "created_at", None)
        if eff_dt:
            key = eff_dt.strftime("%Y-%m-%d")
            day_counts[key] = day_counts.get(key, 0) + 1

        matched_count = 0
        matched_payload = 0.0

        for e in (getattr(l, "entries", None) or []):
            if not entry_matches(e):
                continue

            matched_count += 1
            total_entries += 1

            if getattr(e, "gear_rental", False):
                gear_entries += 1

            sc = (getattr(e, "status_code", "") or "").strip()
            if sc:
                status_counts[sc] = status_counts.get(sc, 0) + 1

            try:
                if getattr(e, "payload_kg", None) is not None:
                    matched_payload += float(getattr(e, "payload_kg", 0) or 0)
                else:
                    matched_payload += float(getattr(e, "calculated_payload", 0) or 0)
            except Exception:
                pass

            try:
                entry_invoice_info = invoice_info_by_entry.get(int(getattr(e, "id", 0) or 0), {})
                if entry_invoice_info:
                    gross = _money(entry_invoice_info.get("gross") or "0.00")
                    net = _money(entry_invoice_info.get("net") or "0.00")
                    vat = _money(entry_invoice_info.get("vat") or "0.00")
                    vat_rate = _money(entry_invoice_info.get("vat_rate") or "0.00")
                else:
                    gross = _money(BillingService.calculate_price_for_entry(e))
                    vat_rate = _money(BillingService.get_entry_vat_rate(e))
                    net, vat = BillingService.split_gross_into_net_and_vat(gross=gross, vat_rate=vat_rate)
                gross_sum_est += gross
                net_sum_est += _money(net)
                vat_sum_est += _money(vat)
            except Exception:
                pass

        per_load[getattr(l, "id", None)] = {
            "matched_count": matched_count,
            "matched_payload": matched_payload,
        }

    # Auch manuelle Rechnungspositionen als statistische Positionen zählen.
    total_entries += manual_item_count

    # Gesamttotals (Sprünge + Schirmmiete + Orga) — nach der For-Schleife, wenn gross_sum_est befüllt ist
    gross_sum_jumps = gross_sum_est
    net_sum_jumps = net_sum_est
    vat_sum_jumps = vat_sum_est
    gross_sum_total = gross_sum_jumps + extra_rental_sum_gross + extra_orga_sum_gross + manual_sum_gross
    net_sum_total = net_sum_jumps + extra_rental_sum_net + extra_orga_sum_net + manual_sum_net
    vat_sum_total = vat_sum_jumps + extra_rental_sum_vat + extra_orga_sum_vat + manual_sum_vat

    matched_entries = []
    for l in loads:
        eff_dt = getattr(l, "actual_time", None) or getattr(l, "created_at", None)
        for e in (getattr(l, "entries", None) or []):
            if not entry_matches(e):
                continue

            gross = net = vat = vat_rate = Decimal("0.00")
            try:
                entry_invoice_info = invoice_info_by_entry.get(int(getattr(e, "id", 0) or 0), {})
                if entry_invoice_info:
                    gross = _money(entry_invoice_info.get("gross") or "0.00")
                    net = _money(entry_invoice_info.get("net") or "0.00")
                    vat = _money(entry_invoice_info.get("vat") or "0.00")
                    vat_rate = _money(entry_invoice_info.get("vat_rate") or "0.00")
                else:
                    gross = _money(BillingService.calculate_price_for_entry(e))
                    vat_rate = _money(BillingService.get_entry_vat_rate(e))
                    net, vat = BillingService.split_gross_into_net_and_vat(gross=gross, vat_rate=vat_rate)
            except Exception:
                pass

            status_css_class = _archive_entry_css_class(l, e)
            status_bg, status_border = _archive_colors_from_css_class(status_css_class)

            matched_entries.append({
                "date": eff_dt,
                "time": eff_dt.strftime("%H:%M") if eff_dt else "",
                "entry_id": getattr(e, "id", None),
                "load_id": getattr(l, "id", None),
                "load_number": getattr(l, "load_number", ""),
                "load_label": f"Load {getattr(l, 'load_number', '')}",
                "airfield": getattr(getattr(l, "airfield", None), "name", ""),
                "aircraft": f"{getattr(getattr(l, 'aircraft', None), 'type', '')} {getattr(getattr(l, 'aircraft', None), 'registration', '')}".strip(),
                "person": getattr(getattr(e, "person", None), "full_name", ""),
                "person_id": getattr(e, "person_id", None),
                "status": (getattr(e, "status_code", "") or "").strip(),
                "status_codes": [((getattr(e, "status_code", "") or "").strip())] if (getattr(e, "status_code", "") or "").strip() else [],
                "payment_method": invoice_info_by_entry.get(int(getattr(e, "id", 0) or 0), {}).get("payment_method", ""),
                "invoice_number": invoice_info_by_entry.get(int(getattr(e, "id", 0) or 0), {}).get("invoice_number", _invoice_number_missing_label()),
                "height_m": getattr(e, "height_m", None),
                "price_gross": gross,
                "price_net": _money(net),
                "price_vat": _money(vat),
                "vat_rate": vat_rate,
                "status_css_class": status_css_class,
                "status_bg": status_bg,
                "status_border": status_border,
            })

    for row in manual_invoice_rows:
        matched_entries.append({
            "date": row.get("effective_dt"),
            "time": row.get("load_time", ""),
            "entry_id": None,
            "load_id": None,
            "load_number": row.get("load_number", ""),
            "load_label": row.get("load_number", ""),
            "airfield": "",
            "aircraft": "",
            "person": row.get("person_name", ""),
            "person_id": row.get("person_id", None),
            "status": row.get("status_code", ""),
            "status_codes": list(row.get("status_codes") or []),
            "payment_method": row.get("payment_method", ""),
            "invoice_number": row.get("invoice_number", _invoice_number_missing_label()),
            "height_m": None,
            "price_gross": Decimal(str(row.get("gross") or "0.00")),
            "price_net": Decimal(str(row.get("net") or "0.00")),
            "price_vat": Decimal(str(row.get("vat") or "0.00")),
            "vat_rate": Decimal(str(row.get("vat_rate") or "0.00")),
            "status_css_class": "",
            "status_bg": "",
            "status_border": "",
        })

    for row in manual_invoice_rows:
        for sc in list(row.get("status_codes") or []):
            sc_clean = (sc or "").strip()
            if not sc_clean:
                continue
            status_counts[sc_clean] = status_counts.get(sc_clean, 0) + 1

    day_series = sorted(day_counts.items(), key=lambda x: x[0])
    status_series = sorted(status_counts.items(), key=lambda x: (-x[1], x[0]))
    max_day = max([c for _, c in day_series], default=0)
    max_status = max([c for _, c in status_series], default=0)

    jumper_by_id: dict[int, dict] = {}
    for r in matched_entries:
        pid = r.get("person_id")
        pname = r.get("person") or ""
        if not pid:
            continue
        if pid not in jumper_by_id:
            jumper_by_id[pid] = {
                "person_id": pid,
                "person_name": pname,
                "total": 0,
                "by_status": {},
            }
        jumper_by_id[pid]["total"] += 1
        by_status = jumper_by_id[pid]["by_status"]
        row_status_codes = [
            (code or "").strip()
            for code in list(r.get("status_codes") or [])
            if (code or "").strip()
        ]
        if row_status_codes:
            for sc in row_status_codes:
                by_status[sc] = by_status.get(sc, 0) + 1
        else:
            sc = (r.get("status") or "").strip()
            if sc:
                by_status[sc] = by_status.get(sc, 0) + 1

    jumper_ranking = sorted(
        jumper_by_id.values(),
        key=lambda x: (-x["total"], x["person_name"].lower()),
    )

    generated_at = now_berlin().replace(tzinfo=None)
    period_value = "Alle Daten"
    if print_filter_summary and print_filter_summary[0].startswith("Zeitraum"):
        period_value = print_filter_summary[0].split(": ", 1)[-1]

    ranking_max = jumper_ranking[0]["total"] if jumper_ranking else 0
    report_period_is_filtered = period_value != "Alle Daten"

    return {
        "loads": loads,
        "airfields": airfields,
        "aircrafts": aircrafts,
        "persons": persons,
        "statuses": statuses,
        "total_loads": total_loads,
        "total_entries": total_entries,
        "tank_loads": tank_loads,
        "gear_entries": gear_entries,
        "gross_sum_est": gross_sum_est,
        "net_sum_est": net_sum_est,
        "vat_sum_est": vat_sum_est,
        "gross_sum_jumps": gross_sum_jumps,
        "net_sum_jumps": net_sum_jumps,
        "vat_sum_jumps": vat_sum_jumps,
        "gross_sum_total": gross_sum_total,
        "net_sum_total": net_sum_total,
        "vat_sum_total": vat_sum_total,
        "day_series": day_series,
        "status_series": status_series,
        "max_day": max_day,
        "max_status": max_status,
        "per_load": per_load,
        "matched_entries": matched_entries,
        "jumper_ranking": jumper_ranking,
        "extra_rental_items": extra_rental_items,
        "extra_rental_sum_net": extra_rental_sum_net,
        "extra_rental_sum_vat": extra_rental_sum_vat,
        "extra_rental_sum_gross": extra_rental_sum_gross,
        "extra_orga_items": extra_orga_items,
        "extra_orga_sum_net": extra_orga_sum_net,
        "extra_orga_sum_vat": extra_orga_sum_vat,
        "extra_orga_sum_gross": extra_orga_sum_gross,
        "manual_invoice_rows": manual_invoice_rows,
        "manual_invoice_count": manual_invoice_count,
        "manual_item_count": manual_item_count,
        "manual_sum_gross": manual_sum_gross,
        "manual_sum_net": manual_sum_net,
        "manual_sum_vat": manual_sum_vat,
        "print_generated_at": generated_at,
        "print_filter_summary": print_filter_summary,
        "query_string": request.query_string.decode("utf-8"),
        "pdf_title": "Statistikbericht",
        "pdf_subtitle": "Auswertung durchgeführter Loads",
        "report_period_label": period_value,
        "report_period_is_filtered": report_period_is_filtered,
        "ranking_preview": jumper_ranking[:12],
        "ranking_max": ranking_max,
        "status_preview": status_series[:10],
        "day_preview": day_series[:10],
        "show_col_net": show_col_net,
        "show_col_vat": show_col_vat,
        "show_col_vatrate": show_col_vatrate,
        "show_col_payment": show_col_payment,
        "show_col_invoice_number": show_col_invoice_number,
        "show_col_gross": show_col_gross,
        "report_signature_roles": [
            "Erstellt von",
            "Geprüft von",
            "Freigabe Verein",
        ],
        "summary_tiles": [
            {"label": "Durchgeführte Loads", "value": total_loads},
            {"label": "Positionen gesamt", "value": total_entries},
            {"label": "Tank-Stops", "value": tank_loads},
            {"label": "Schirmmiete-Positionen", "value": gear_entries},
            {"label": "Brutto Sprünge", "value": _fmt_money_de(gross_sum_jumps)},
            {"label": "Brutto Schirmmiete", "value": _fmt_money_de(extra_rental_sum_gross)},
            {"label": "Brutto Orga", "value": _fmt_money_de(extra_orga_sum_gross)},
            {"label": "Brutto manuelle Rechnungen", "value": _fmt_money_de(manual_sum_gross)},
            {"label": "BRUTTO GESAMT", "value": _fmt_money_de(gross_sum_total)},
            {"label": "Netto gesamt", "value": _fmt_money_de(net_sum_total)},
            {"label": "MwSt gesamt", "value": _fmt_money_de(vat_sum_total)},
            {"label": "Zeitraum", "value": period_value},
        ],
    }


@bp_load.route("/statistics", endpoint="statistics_view")
def statistics_view():
    return render_template("statistics/index.html", **_build_statistics_context(request.args))


@bp_load.route("/statistics/report.pdf", endpoint="statistics_report_pdf")
def statistics_report_pdf():
    context = _build_statistics_context(request.args)
    compact_raw = (request.args.get("compact") or "").strip().lower()
    compact_mode = compact_raw in {"1", "true", "yes", "on", "kompakt"}

    # Für finanzielle Kennzahlen exakt dieselbe Datenbasis wie Billing-Report verwenden.
    from app.routes.billing import _build_invoice_list_context

    billing_config = BillingConfig.query.first()
    static_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "static", "img"))

    logo_filename = getattr(billing_config, "logo_filename", None) or "Head_Logo.png"
    logo_data_uri = _image_to_data_uri(os.path.join(static_dir, logo_filename))
    secondary_logo_data_uri = _image_to_data_uri(os.path.join(static_dir, "Logo_DZ.png"))
    fallback_logo_data_uri = _image_to_data_uri(os.path.join(static_dir, "HU_Bleistift.png"))

    period = (request.args.get("period") or "all").strip().lower() or "all"
    from_str = (request.args.get("from") or "").strip()
    to_str = (request.args.get("to") or "").strip()

    billing_period = period
    if billing_period == "all" and (from_str or to_str):
        billing_period = "range"

    billing_ctx = _build_invoice_list_context(
        period=billing_period,
        from_str=from_str,
        to_str=to_str,
    )

    period_slug_map = {
        "all": "alle",
        "today": "heute",
        "week": "woche",
        "month": "monat",
        "year": "jahr",
        "range": "zeitraum",
    }
    period_slug = period_slug_map.get(period, "alle")

    if period == "range" and (from_str or to_str):
        from_part = from_str.replace("-", "_") if from_str else "offen"
        to_part = to_str.replace("-", "_") if to_str else "offen"
        period_part = f"{period_slug}_{from_part}-{to_part}"
    else:
        period_part = period_slug

    generated_at_local = context.get("print_generated_at") or now_local().replace(tzinfo=None)
    compact_slug = "_kompakt" if compact_mode else ""
    report_filename = (
        f"Abrechnung_Statistikbericht_{period_part}{compact_slug}_"
        f"{generated_at_local.strftime('%Y_%m_%d_%H_%M')}.pdf"
    )

    context.update(
        billing_config=billing_config,
        logo_data_uri=logo_data_uri or fallback_logo_data_uri,
        secondary_logo_data_uri=secondary_logo_data_uri,
        compact_mode=compact_mode,
        report_generated_label=context["print_generated_at"].strftime("%d.%m.%Y %H:%M"),
        report_filename=report_filename,
        billing_sum_billable=billing_ctx.get("sum_billable", Decimal("0.00")),
        billing_sum_open_invoices=billing_ctx.get("sum_open_invoices", Decimal("0.00")),
        billing_delta=billing_ctx.get("delta", Decimal("0.00")),
        billing_billable_rows=billing_ctx.get("billable_rows", []),
        billing_invoices=billing_ctx.get("invoices", []),
    )

    html = render_template("statistics/report_pdf.html", **context)
    pdf_bytes, pdf_error = generate_pdf_from_html(
        html,
        presentational_hints=True,
        optimize_size=("fonts", "images"),
    )
    if pdf_error:
        flash(pdf_error, "danger")
        return redirect(url_for("load.statistics_view", **request.args.to_dict(flat=True)))

    return send_file(
        BytesIO(pdf_bytes),
        mimetype="application/pdf",
        download_name=context["report_filename"],
        as_attachment=False,
    )

# ============================================================
# ✅ Deutsche Spaltenüberschriften für Statistik-Export
# ============================================================

STATISTICS_EXPORT_HEADERS_DE = {
    "row_type": "Zeilentyp",
    "item_type": "Positionstyp",
    "item_desc": "Beschreibung",

    "load_id": "Load-ID",
    "load_number": "Load",
    "load_date": "Datum",
    "load_time": "Uhrzeit",

    "airfield": "Flugplatz",
    "aircraft": "Flugzeug",
    "height_m": "Absetzhöhe (m)",
    "fuel_required": "Tanken",

    "entry_id": "Entry-ID",
    "person_id": "Person-ID",
    "person_name": "Person",
    "status_code": "Status",
    "entry_height_m": "Sprunghöhe (m)",
    "gear_rental": "Schirmmiete",
    "invoice_status": "Rechnungsstatus",
    "payment_method": "Zahlart",
    "invoice_number": "Rechnungsnummer",

    "gross": "Brutto",
    "net": "Netto",
    "vat": "MwSt",
    "vat_rate": "MwSt-Satz",
}


@bp_load.route("/statistics/export.csv", endpoint="statistics_export_csv")
def statistics_export_csv():
    args = request.args
    loads = _build_stats_query(args).all()

    columns = [
        "row_type", "item_type", "item_desc",
        "load_id", "load_number", "load_date", "load_time",
        "airfield", "aircraft", "height_m", "fuel_required",
        "entry_id", "person_id", "person_name", "status_code", "entry_height_m", "gear_rental",
        "invoice_status", "payment_method",
        "invoice_number",
        "gross", "net", "vat", "vat_rate",
    ]

    from io import StringIO, BytesIO

    # ✅ 1) CSV in Text-Puffer schreiben
    text_buffer = StringIO()

    # ✅ Excel BOM (als Text!)
    text_buffer.write("\ufeff")

    writer = csv.DictWriter(
        text_buffer,
        fieldnames=columns,
        delimiter=";",
        lineterminator="\n",
        extrasaction="ignore",
    )

    # ✅ Deutsche Header
    writer.writerow({
        col: STATISTICS_EXPORT_HEADERS_DE.get(col, col)
        for col in columns
    })

    def _fmt_de_csv(value, decimals=2):
        try:
            d = Decimal(str(value or "0"))
        except Exception:
            d = Decimal("0.00")
        return f"{d:.{decimals}f}".replace(".", ",")

    sum_gross = Decimal("0.00")
    sum_net = Decimal("0.00")
    sum_vat = Decimal("0.00")

    # ✅ Datenzeilen
    for row in _iter_export_rows(loads, args):
        out = dict(row)

        try:
            gross_d = Decimal(str(out.get("gross") or "0"))
        except Exception:
            gross_d = Decimal("0.00")
        try:
            net_d = Decimal(str(out.get("net") or "0"))
        except Exception:
            net_d = Decimal("0.00")
        try:
            vat_d = Decimal(str(out.get("vat") or "0"))
        except Exception:
            vat_d = Decimal("0.00")

        sum_gross += gross_d
        sum_net += net_d
        sum_vat += vat_d

        out["gross"] = _fmt_de_csv(gross_d)
        out["net"] = _fmt_de_csv(net_d)
        out["vat"] = _fmt_de_csv(vat_d)
        out["vat_rate"] = _fmt_de_csv(out.get("vat_rate") or "0")

        writer.writerow(out)

    writer.writerow({
        "row_type": "summe",
        "item_type": "Gesamtsumme",
        "item_desc": "Summenzeile Export",
        "gross": _fmt_de_csv(sum_gross),
        "net": _fmt_de_csv(sum_net),
        "vat": _fmt_de_csv(sum_vat),
    })

    # ✅ 2) Text → Bytes konvertieren
    output = BytesIO(text_buffer.getvalue().encode("utf-8"))
    output.seek(0)

    # ✅ 3) Senden
    return send_file(
        output,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name="statistik_loads_completed.csv",
    )


@bp_load.route("/statistics/export.xlsx", endpoint="statistics_export_xlsx")
def statistics_export_xlsx():
    args = request.args
    loads = _build_stats_query(args).all()

    try:
        from openpyxl import Workbook
    except Exception:
        return redirect(url_for("load.statistics_export_csv", **request.args))

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Statistik"

    columns = [
        "row_type", "item_type", "item_desc",
        "load_id", "load_number", "load_date", "load_time",
        "airfield", "aircraft", "height_m", "fuel_required",
        "person_id", "person_name", "status_code", "entry_height_m", "gear_rental",
        "invoice_status", "payment_method",
        "invoice_number",
        "gross", "net", "vat", "vat_rate",
    ]

    # Spaltenbreiten (in Excel-Einheiten ≈ Zeichenanzahl)
    COL_WIDTHS = {
        "row_type":       12,
        "item_type":      14,
        "item_desc":      45,
        "load_id":         9,
        "load_number":     8,
        "load_date":      13,
        "load_time":       9,
        "airfield":       22,
        "aircraft":       20,
        "height_m":       16,
        "fuel_required":  10,
        "person_id":       9,
        "person_name":    28,
        "status_code":    20,
        "entry_height_m": 16,
        "gear_rental":    13,
        "invoice_status": 16,
        "payment_method": 28,
        "invoice_number": 22,
        "gross":          13,
        "net":            13,
        "vat":            12,
        "vat_rate":       13,
    }

    # Spalten
    PRICE_COLS    = {"gross", "net", "vat"}
    VAT_RATE_COLS = {"vat_rate"}
    INT_COLS      = {"load_id", "load_number", "person_id"}
    PERSON_NAME_COL_IDX = columns.index("person_name") + 1
    STATUS_CODE_COL_IDX = columns.index("status_code") + 1

    # Null-sichere Konvertierung
    def _to_int(v):
        try:
            return int(str(v)) if v not in ("", None) else None
        except Exception:
            return None

    def _to_float(v):
        try:
            return float(str(v)) if v not in ("", None) else None
        except Exception:
            return None

    def _fmt_de(v, suffix="", decimals=2):
        """Formatiert einen Zahlenwert als deutschen Dezimal-String.
        Beispiel: 1234.5 → '1.234,50 €'
        """
        try:
            f = float(str(v)) if v not in ("", None) else None
            if f is None:
                return ""
            # Python-Standard: '1,234.50'  →  deutsch: '1.234,50'
            s = f"{f:,.{decimals}f}"
            s = s.replace(",", "THOU").replace(".", ",").replace("THOU", ".")
            return f"{s} {suffix}".strip() if suffix else s
        except Exception:
            return str(v) if v not in ("", None) else ""

    # Header-Zeile
    header_row = [STATISTICS_EXPORT_HEADERS_DE.get(col, col) for col in columns]
    ws.append(header_row)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=False)

    # Datenzeilen
    for row in _iter_export_rows(loads, args):
        excel_row = []
        for c in columns:
            v = row.get(c, "")
            if c in PRICE_COLS:
                excel_row.append(_to_float(v))
            elif c in VAT_RATE_COLS:
                excel_row.append(_to_float(v))
            elif c in INT_COLS:
                excel_row.append(_to_int(v))
            else:
                excel_row.append(v if v is not None else "")
        ws.append(excel_row)

        if row.get("row_type") == "entry":
            status_bg = row.get("status_bg")
            status_border = row.get("status_border")
            current_row = ws.max_row

            if status_bg:
                ws.cell(row=current_row, column=PERSON_NAME_COL_IDX).fill = PatternFill(
                    "solid",
                    fgColor=str(status_bg).replace("#", "").upper(),
                )
                ws.cell(row=current_row, column=STATUS_CODE_COL_IDX).fill = PatternFill(
                    "solid",
                    fgColor=str(status_bg).replace("#", "").upper(),
                )

            if status_border:
                left_border = Border(
                    left=Side(style="thick", color=str(status_border).replace("#", "").upper())
                )
                ws.cell(row=current_row, column=PERSON_NAME_COL_IDX).border = left_border
                ws.cell(row=current_row, column=STATUS_CODE_COL_IDX).border = left_border

    data_start_row = 2
    data_end_row = ws.max_row

    if data_end_row >= data_start_row:
        total_row_idx = data_end_row + 1
        ws.cell(row=total_row_idx, column=1, value="summe")
        ws.cell(row=total_row_idx, column=2, value="Gesamtsumme")

        gross_col_idx = columns.index("gross") + 1
        net_col_idx = columns.index("net") + 1
        vat_col_idx = columns.index("vat") + 1

        gross_col_letter = get_column_letter(gross_col_idx)
        net_col_letter = get_column_letter(net_col_idx)
        vat_col_letter = get_column_letter(vat_col_idx)

        ws.cell(row=total_row_idx, column=gross_col_idx, value=f"=SUM({gross_col_letter}{data_start_row}:{gross_col_letter}{data_end_row})")
        ws.cell(row=total_row_idx, column=net_col_idx, value=f"=SUM({net_col_letter}{data_start_row}:{net_col_letter}{data_end_row})")
        ws.cell(row=total_row_idx, column=vat_col_idx, value=f"=SUM({vat_col_letter}{data_start_row}:{vat_col_letter}{data_end_row})")

        total_fill = PatternFill("solid", fgColor="E2F0D9")
        for cell in ws[total_row_idx]:
            cell.font = Font(bold=True)
            cell.fill = total_fill

    # Spaltenbreiten setzen
    for col_idx, col_key in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COL_WIDTHS.get(col_key, 14)

    # Ausrichtung: Preis-/Zahlenspalten rechtsbündig
    PRICE_AND_VAT = PRICE_COLS | VAT_RATE_COLS
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx, cell in enumerate(row_cells, start=1):
            col_key = columns[col_idx - 1]
            if col_key in PRICE_AND_VAT:
                cell.alignment = Alignment(horizontal="right")
                if col_key in PRICE_COLS:
                    cell.number_format = '#,##0.00 [$€-407]'
                elif col_key in VAT_RATE_COLS:
                    cell.number_format = '0.00" %"'
            elif col_key in INT_COLS:
                cell.alignment = Alignment(horizontal="right")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="statistik_loads_completed.xlsx",
    )

# ============================================================
# BLOCK 11a — STATISTIK: LOAD-DETAIL (Drill-Down, read-only)
# - nur durchgeführte Loads
# - Preise werden aus BillingService berechnet (nicht aus Rechnungen)
# ============================================================

@bp_load.route("/statistics/load/<int:load_id>", endpoint="statistics_load_detail")
def statistics_load_detail(load_id: int):
    load = (
        Load.query
        .options(
            db.joinedload(Load.entries).joinedload(LoadEntry.person),
            db.joinedload(Load.airfield),
            db.joinedload(Load.aircraft),
        )
        .filter(
            Load.id == load_id,
            Load.status == "completed"
        )
        .first_or_404()
    )

    rows = []
    for e in (load.entries or []):
        gross = net = vat = vat_rate = Decimal("0.00")
        try:
            gross = _money(BillingService.calculate_price_for_entry(e))
            vat_rate = _money(BillingService.get_entry_vat_rate(e))
            net, vat = BillingService.split_gross_into_net_and_vat(gross=gross, vat_rate=vat_rate)
        except Exception:
            pass

        rows.append({
            "entry": e,
            "person": getattr(getattr(e, "person", None), "full_name", ""),
            "status": getattr(e, "status_code", ""),
            "height_m": getattr(e, "height_m", None),
            "gear_rental": bool(getattr(e, "gear_rental", False)),
            "gross_est": gross,
            "net_est": _money(net),
            "vat_est": _money(vat),
            "vat_rate": vat_rate,
        })

    return render_template(
        "statistics/load_detail.html",
        load=load,
        rows=rows,
    )

# EOF