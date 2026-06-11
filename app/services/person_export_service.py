from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


class PersonExportService:

    @staticmethod
    def export_vertical_import_template():
        output = BytesIO()
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Vorlage"

        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _style(cell, *, fill=None, bold=False, center=False, locked=True, number_format=None):
            cell.font = Font(size=11, bold=bold)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if center else "left",
                vertical="center",
                wrap_text=False,
            )
            cell.protection = Protection(locked=locked)
            if fill:
                cell.fill = PatternFill(fill_type="solid", fgColor=fill)
            if number_format:
                cell.number_format = number_format

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 28
        for c in range(3, 8):  # C-G
            ws.column_dimensions[get_column_letter(c)].width = 20
        for c in range(8, 28):  # H-AA
            col = get_column_letter(c)
            ws.column_dimensions[col].width = 20
            ws.column_dimensions[col].hidden = True

        # Zeile 1
        ws.row_dimensions[1].height = 22
        ws["A1"] = "Pflichtfeld in gelb / optional in grün Eingabe der Daten ab Spalte C (Person 1)"
        _style(ws["A1"], fill="D9D9D9", locked=True)
        _style(ws["B1"], fill="D9D9D9", locked=True)
        for c in range(3, 8):
            _style(ws.cell(1, c), fill="D9D9D9", locked=True)
        for c in range(8, 28):
            ws.cell(1, c).protection = Protection(locked=False)

        # Zeile 2
        ws.row_dimensions[2].height = 22
        ws["A2"] = "Beschreibung"
        ws["B2"] = "Feldname"
        _style(ws["A2"], fill="D9D9D9", bold=True, center=True, locked=True)
        _style(ws["B2"], fill="D9D9D9", bold=True, center=True, locked=True)

        for i in range(1, 6):
            cell = ws.cell(2, 2 + i)
            cell.value = f"Person {i}"
            _style(cell, fill="D9D9D9", bold=True, center=True, locked=True)

        for c in range(8, 28):
            ws.cell(2, c).protection = Protection(locked=False)

        fields = [
            ("Vorname", "first_name", True, "@"),
            ("Nachname", "last_name", True, "@"),
            ("Telefonnummer", "phone", True, "@"),
            ("E-Mail_Adresse", "email", True, "@"),
            ("Gewicht in ganzen kg", "weight_kg", True, "0"),
            ("Notfallkontakt (Vorname + Nachname)", "emergency_name", True, "@"),
            ("Beziehung zum Notfallkontakt", "emergency_relation", True, "@"),
            ("Telefonnummer des Notfallkontaktes", "emergency_phone", True, "@"),
            ("E-Mail des Notfallkontaktes", "emergency_email", True, "@"),
            ("Vereinsmitglied", "is_member", False, "@"),
            ("Tandemgast", "is_tandem_guest", False, "@"),
            ("Größe in cm (nicht in m)", "height_cm", False, "0"),
            ("Geburtstag im Format dd.mm.yyyy", "birthdate", False, "dd.mm.yyyy"),
            ("Straße und Hausnummer", "street_and_number", False, "@"),
            ("PLZ", "zip_code", False, "@"),
            ("Wohnort", "city", False, "@"),
            ("Lizenznummer", "license_number", False, "@"),
            ("Haftpflichtversicherung für Springer bei:", "insurance_company", False, "@"),
            ("Versicherungsnummer", "insurance_number", False, "@"),
            ("Kto.-Nr. als IBAN (genau prüfen)", "iban", False, "@"),
            ("BIC", "bic", False, "@"),
            ("Kontoinhaber", "account_holder", False, "@"),
            ("Kommentar", "comment", False, "@"),
            ("Notizen", "notes", False, "@"),
        ]

        dv_yes_no = DataValidation(type="list", formula1='"ja,nein"', allow_blank=True)
        ws.add_data_validation(dv_yes_no)

        row = 3
        for desc, field, required, num_fmt in fields:
            ws.cell(row, 1, desc)
            ws.cell(row, 2, field)
            _style(ws.cell(row, 1), fill="D9D9D9", locked=True)
            _style(ws.cell(row, 2), fill="D9D9D9", locked=True)

            for c in range(3, 8):
                cell = ws.cell(row, c)
                if field in ("is_member", "is_tandem_guest"):
                    _style(cell, fill="E2EFDA", locked=False, number_format="@")
                    dv_yes_no.add(cell)
                else:
                    fill = "FFF2CC" if required else "E2EFDA"
                    _style(cell, fill=fill, locked=False, number_format=num_fmt)

            for c in range(8, 28):
                ws.cell(row, c).protection = Protection(locked=False)

            row += 1

        # Schutz aktivieren
        ws.protection.sheet = True
        ws.protection.selectLockedCells = True
        ws.protection.selectUnlockedCells = True

        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="personen_import_vorlage.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @staticmethod
    def export_persons_excel(persons):
        """Exportiert eine Personenliste als XLSX-Datei."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Personen"

        headers = [
            "ID",
            "Vorname",
            "Nachname",
            "Ursprungsname",
            "Telefon",
            "E-Mail",
            "Gewicht (kg)",
            "Größe (cm)",
            "Geburtstag",
            "Vereinsmitglied",
            "Partnerverein",
            "Tandemgast",
            "Lehrer",
            "AFF-Lehrer",
            "Schüler-AFF",
            "Video",
            "Lehrerlizenz bis",
            "Enthaftung",
            "Straße/Hausnummer",
            "PLZ",
            "Ort",
            "Lizenznummer",
            "Versicherung",
            "Versicherungsnummer",
            "Notfallkontakt",
            "Notfall-Beziehung",
            "Notfall-Telefon",
            "Notfall-E-Mail",
            "IBAN",
            "BIC",
            "Kontoinhaber",
            "Kommentar",
            "Notizen",
        ]

        ws.append(headers)

        def _as_yes_no(value):
            return "ja" if bool(value) else "nein"

        def _as_date(value):
            return value.strftime("%d.%m.%Y") if value else ""

        for p in persons:
            ws.append([
                p.id,
                p.first_name or "",
                p.last_name or "",
                p.original_name or "",
                p.phone or "",
                p.email or "",
                p.weight_kg if p.weight_kg is not None else "",
                p.height_cm if p.height_cm is not None else "",
                _as_date(p.birthdate),
                _as_yes_no(p.is_member),
                _as_yes_no(getattr(p, "is_partner_verein", False)),
                _as_yes_no(p.is_tandem_guest),
                _as_yes_no(p.is_teacher),
                _as_yes_no(getattr(p, "is_aff_teacher", False)),
                _as_yes_no(getattr(p, "is_aff_student", False)),
                _as_yes_no(getattr(p, "is_video", False)),
                _as_date(p.teacher_license_expires),
                _as_date(p.liability_waiver_date),
                p.street_and_number or "",
                p.zip_code or "",
                p.city or "",
                p.license_number or "",
                p.insurance_provider or "",
                p.insurance_number or "",
                p.emergency_name or "",
                p.emergency_relation or "",
                p.emergency_phone or "",
                p.emergency_email or "",
                p.iban or "",
                p.bic or "",
                p.account_holder or "",
                p.comment or "",
                p.notes or "",
            ])

        # Lesbare Spaltenbreiten
        widths = {
            "A": 8,
            "B": 16,
            "C": 18,
            "D": 24,
            "E": 18,
            "F": 28,
            "G": 12,
            "H": 11,
            "I": 13,
            "J": 14,
            "K": 14,
            "L": 12,
            "M": 10,
            "N": 14,
            "O": 12,
            "P": 10,
            "Q": 12,
            "R": 14,
            "S": 12,
            "T": 24,
            "U": 10,
            "V": 16,
            "W": 16,
            "X": 20,
            "Y": 22,
            "Z": 24,
            "AA": 18,
            "AB": 18,
            "AC": 24,
            "AD": 24,
            "AE": 14,
            "AF": 24,
            "AG": 24,
            "AH": 28,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="personen_export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
