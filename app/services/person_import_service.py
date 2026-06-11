from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from app.models.person import Person
from app import db
import os
import csv
import re

# ---------------------------------------------------------
# Normalisierungs- und Validierungsfunktionen
# ---------------------------------------------------------
EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[a-zA-Z0-9]+$")


def normalize_phone(phone):
    if not phone:
        return None
    p = str(phone)
    p = p.replace(" ", "").replace("-", "").replace("/", "")
    p = p.replace("(", "").replace(")", "")

    if p.startswith("+49"):
        p = "0" + p[3:]
    if p.startswith("0049"):
        p = "0" + p[4:]

    return p if len(p) >= 3 else None


def normalize_email(email):
    if not email:
        return None
    e = str(email).strip().lower().replace(" ", "")
    return e if e else None


def is_valid_email(email):
    return bool(email and EMAIL_REGEX.match(email))


def normalize_name(name):
    return (str(name).strip().lower()) if name else ""


# ---------------------------------------------------------
# FIXED VERSION – verhindert deinen SQLite-Date-Fehler
# ---------------------------------------------------------
def parse_birthdate(value, notes, field_errors):
    if value in (None, ""):
        return None

    # Excel-Seriennummern
    if isinstance(value, (int, float)):
        try:
            excel_start = date(1899, 12, 30)
            d = excel_start + timedelta(days=int(value))
            if d.year < 1900 or d > date.today():
                field_errors["birthdate"] = True
                notes.append(f"Geburtstag unplausibel („{d}“) – Wert geleert")
                return None
            return d
        except Exception:
            field_errors["birthdate"] = True
            notes.append(f"Geburtstag ungültig („{value}“) – Wert geleert")
            return None

    # datetime / date
    if isinstance(value, (datetime, date)):
        d = value.date() if isinstance(value, datetime) else value
        if d.year < 1900 or d > date.today():
            field_errors["birthdate"] = True
            notes.append(f"Geburtstag unplausibel („{d}“) – Wert geleert")
            return None
        return d

    # Strings
    s = str(value).strip()

    # Sonderfall: "1900-06-17 00:00:00"
    if " " in s:
        s = s.split(" ")[0]

    # Versuch: YYYY-MM-DD
    try:
        d = datetime.strptime(s, "%Y-%m-%d").date()
        if d.year < 1900 or d > date.today():
            raise ValueError
        return d
    except Exception:
        pass

    # Versuch: DD.MM.YYYY / DD.MM.YY
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            d = datetime.strptime(s, fmt).date()
            if d.year < 1900 or d > date.today():
                raise ValueError
            return d
        except Exception:
            continue

    # Wenn alles fehlschlägt
    field_errors["birthdate"] = True
    notes.append(f"Geburtstag ungültig („{s}“) – Wert geleert")
    return None


def parse_height_cm(value, notes, field_errors):
    if value in (None, ""):
        return None

    s = str(value).lower().replace(" ", "")
    s = s.replace("cm", "").replace("m", "")
    s = s.replace(",", ".")

    try:
        num = float(s)
    except Exception:
        field_errors["height_cm"] = True
        notes.append(f"Größe ungültig („{value}“) – Wert geleert")
        return None

    if 1.40 <= num <= 2.20:
        cm = int(round(num * 100))
        if 140 <= cm <= 220:
            return cm

    if 140 <= num <= 220:
        return int(round(num))

    field_errors["height_cm"] = True
    notes.append(f"Größe unlogisch („{value}“) – Wert geleert")
    return None


def parse_weight_kg(value, notes, field_errors):
    if value in (None, ""):
        return None

    s = str(value).lower().replace("kg", "").strip()

    try:
        w = float(s)
    except Exception:
        field_errors["weight_kg"] = True
        notes.append(f"Gewicht ungültig („{value}“) – Wert geleert")
        return None

    w = int(round(w))

    if w < 20 or w > 200:
        field_errors["weight_kg"] = True
        notes.append(f"Gewicht unplausibel („{w}“) – Wert geleert")
        return None

    return w


def parse_iban(value, notes, field_errors, field_warnings):
    if value in (None, ""):
        return None

    s = str(value).strip().replace(" ", "").upper()
    if not s:
        return None

    plausible = (
        s.startswith("DE")
        and len(s) == 22
        and s.isalnum()
    )

    if plausible:
        field_warnings["iban_manual_check"] = True
        notes.append("IBAN plausibel, manuelle Prüfung empfohlen")
        return s

    # Unplausible IBAN wird verworfen, soll den gesamten Datensatz aber nicht blockieren.
    field_warnings["iban"] = True
    notes.append(f"IBAN unplausibel („{s}“) – Wert geleert")
    return None


# ---------------------------------------------------------
# Service
# ---------------------------------------------------------
class PersonImportService:

    HORIZONTAL_HEADER_MAP = {
        "id": "id",
        "vorname": "first_name",
        "nachname": "last_name",
        "ursprungsname": "original_name",
        "telefon": "phone",
        "e-mail": "email",
        "gewicht (kg)": "weight_kg",
        "groesse (cm)": "height_cm",
        "größe (cm)": "height_cm",
        "geburtstag": "birthdate",
        "vereinsmitglied": "is_member",
        "partnerverein": "is_partner_verein",
        "tandemgast": "is_tandem_guest",
        "lehrer": "is_teacher",
        "aff-lehrer": "is_aff_teacher",
        "schüler-aff": "is_aff_student",
        "schueler-aff": "is_aff_student",
        "video": "is_video",
        "lehrerlizenz bis": "teacher_license_expires",
        "enthaftung": "liability_waiver_date",
        "strasse/hausnummer": "street_and_number",
        "straße/hausnummer": "street_and_number",
        "plz": "zip_code",
        "ort": "city",
        "lizenznummer": "license_number",
        "versicherung": "insurance_provider",
        "versicherungsnummer": "insurance_number",
        "notfallkontakt": "emergency_name",
        "notfall-beziehung": "emergency_relation",
        "notfall-telefon": "emergency_phone",
        "notfall-e-mail": "emergency_email",
        "iban": "iban",
        "bic": "bic",
        "kontoinhaber": "account_holder",
        "kommentar": "comment",
        "notizen": "notes",
    }

    FIELD_LABELS = {
        "first_name": "Vorname",
        "last_name": "Nachname",
        "phone": "Telefon",
        "email": "E-Mail",
        "weight_kg": "Gewicht",
        "height_cm": "Größe",
        "birthdate": "Geburtstag",
        "is_member": "Vereinsmitglied",
        "is_partner_verein": "Partner-Verein",
        "is_tandem_guest": "Tandemgast",
        "is_teacher": "Lehrer",
        "is_aff_teacher": "AFF-Lehrer",
        "is_aff_student": "Schüler-AFF",
        "is_video": "Video",
        "teacher_license_expires": "Lehrerlizenz bis",
        "liability_waiver_date": "Enthaftung",
        "street_and_number": "Straße/Hausnummer",
        "zip_code": "PLZ",
        "city": "Ort",
        "license_number": "Lizenznummer",
        "insurance_provider": "Versicherung",
        "insurance_number": "Versicherungsnummer",
        "emergency_name": "Notfallkontakt",
        "emergency_relation": "Notfall-Beziehung",
        "emergency_phone": "Notfall-Telefon",
        "emergency_email": "Notfall-E-Mail",
        "iban": "IBAN",
        "bic": "BIC",
        "account_holder": "Kontoinhaber",
        "comment": "Kommentar",
        "notes": "Notizen",
    }

    MERGE_FIELD_ORDER = [
        "first_name",
        "last_name",
        "phone",
        "email",
        "weight_kg",
        "height_cm",
        "birthdate",
        "is_member",
        "is_partner_verein",
        "is_tandem_guest",
        "is_teacher",
        "is_aff_teacher",
        "is_aff_student",
        "is_video",
        "teacher_license_expires",
        "liability_waiver_date",
        "street_and_number",
        "zip_code",
        "city",
        "license_number",
        "insurance_provider",
        "insurance_number",
        "emergency_name",
        "emergency_relation",
        "emergency_phone",
        "emergency_email",
        "iban",
        "bic",
        "account_holder",
        "comment",
        "notes",
    ]

    @staticmethod
    def _parse_bool(value):
        if value is None:
            return False
        s = str(value).strip().lower()
        return s in {"ja", "true", "1", "x", "yes"}

    @staticmethod
    def _is_blank(value):
        return value is None or (isinstance(value, str) and value.strip() == "")

    @staticmethod
    def _normalize_date_value(value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()

        s = str(value).strip()
        if not s:
            return None
        if " " in s:
            s = s.split(" ")[0]

        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except Exception:
                continue
        return s

    @staticmethod
    def _normalize_compare_value(field, value):
        if field in {"phone", "emergency_phone"}:
            return normalize_phone(value)
        if field in {"email", "emergency_email"}:
            return normalize_email(value)
        if field in {"first_name", "last_name"}:
            return normalize_name(value)
        if field in {"birthdate", "teacher_license_expires", "liability_waiver_date"}:
            return PersonImportService._normalize_date_value(value)
        if field in {"is_member", "is_partner_verein", "is_tandem_guest", "is_teacher", "is_aff_teacher", "is_aff_student", "is_video"}:
            if isinstance(value, bool):
                return value
            if value in (None, ""):
                return False
            return PersonImportService._parse_bool(value)
        if isinstance(value, str):
            return value.strip()
        return value

    @staticmethod
    def values_equal(field, old_value, new_value):
        return (
            PersonImportService._normalize_compare_value(field, old_value)
            == PersonImportService._normalize_compare_value(field, new_value)
        )

    @staticmethod
    def format_merge_value(field, value):
        if PersonImportService._is_blank(value):
            return "–"
        if field in {"is_member", "is_partner_verein", "is_tandem_guest", "is_teacher", "is_aff_teacher", "is_aff_student", "is_video"}:
            normalized = PersonImportService._normalize_compare_value(field, value)
            return "Ja" if normalized else "Nein"
        if field in {"birthdate", "teacher_license_expires", "liability_waiver_date"}:
            normalized = PersonImportService._normalize_date_value(value)
            if not normalized:
                return "–"
            try:
                return datetime.strptime(normalized, "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                return normalized
        return str(value)

    @staticmethod
    def _default_merge_choice(field, old_value, new_value, import_mode):
        if PersonImportService.values_equal(field, old_value, new_value):
            return "existing"
        if PersonImportService._is_blank(old_value) and not PersonImportService._is_blank(new_value):
            return "import"
        if PersonImportService._is_blank(new_value):
            return "existing"
        if import_mode == "horizontal":
            return "import"
        return "existing"

    @staticmethod
    def build_merge_candidates(existing, cleaned_data, *, import_mode="vertical"):
        candidates = []

        for field in PersonImportService.MERGE_FIELD_ORDER:
            new_value = cleaned_data.get(field)
            old_value = getattr(existing, field, None)

            if PersonImportService.values_equal(field, old_value, new_value):
                continue
            if PersonImportService._is_blank(old_value) and PersonImportService._is_blank(new_value):
                continue

            candidates.append({
                "field": field,
                "label": PersonImportService.FIELD_LABELS.get(field, field),
                "old_value": old_value,
                "new_value": new_value,
                "old_display": PersonImportService.format_merge_value(field, old_value),
                "new_display": PersonImportService.format_merge_value(field, new_value),
                "default_choice": PersonImportService._default_merge_choice(
                    field,
                    old_value,
                    new_value,
                    import_mode,
                ),
                "import_allowed": not PersonImportService._is_blank(new_value) or isinstance(new_value, bool),
                "is_name_change": field in {"first_name", "last_name"},
            })

        return candidates

    @staticmethod
    def build_existing_person_preview(existing):
        return {
            "id": existing.id,
            "full_name": existing.full_name,
            "current_name": existing.current_name,
            "original_name": existing.original_name,
        }

    @staticmethod
    def build_match_indexes(persons):
        index_id = {}
        index_key = {}
        index_email = {}
        index_birth = {}
        index_phone_birth = {}

        for p in persons:
            n_first = normalize_name(p.first_name)
            n_last = normalize_name(p.last_name)
            n_phone = normalize_phone(p.phone)
            n_email = normalize_email(p.email)
            n_birth = p.birthdate.isoformat() if p.birthdate else None

            index_id[p.id] = p
            index_key[(n_first, n_last, n_phone)] = p
            if n_email:
                index_email[n_email] = p
            if n_birth:
                index_birth[(n_first, n_last, n_birth)] = p
            if n_phone and n_birth:
                index_phone_birth[(n_phone, n_birth)] = p

        return {
            "id": index_id,
            "key": index_key,
            "email": index_email,
            "birth": index_birth,
            "phone_birth": index_phone_birth,
        }

    @staticmethod
    def find_existing_person(cleaned_data, match_indexes, *, import_mode="vertical"):
        existing = None
        duplicate_reason = ""

        raw_id = cleaned_data.get("id")
        try:
            pid = int(raw_id) if raw_id not in (None, "") else None
        except Exception:
            pid = None

        if import_mode == "horizontal" and pid and pid in match_indexes["id"]:
            existing = match_indexes["id"][pid]
            duplicate_reason = "ID bereits vorhanden"

        n_email = normalize_email(cleaned_data.get("email"))
        if not existing and n_email and n_email in match_indexes["email"]:
            existing = match_indexes["email"][n_email]
            duplicate_reason = "E-Mail bereits vorhanden"

        n_first = normalize_name(cleaned_data.get("first_name"))
        n_last = normalize_name(cleaned_data.get("last_name"))
        n_phone = normalize_phone(cleaned_data.get("phone"))

        n_birth = cleaned_data.get("birthdate")
        if hasattr(n_birth, "isoformat"):
            n_birth = n_birth.isoformat()
        elif n_birth:
            n_birth = str(n_birth).strip()

        if (
            not existing
            and n_phone
            and n_birth
            and (n_phone, n_birth) in match_indexes["phone_birth"]
        ):
            existing = match_indexes["phone_birth"][(n_phone, n_birth)]
            duplicate_reason = "Telefon + Geburtstag bereits vorhanden"

        if not existing and (n_first, n_last, n_phone) in match_indexes["key"]:
            existing = match_indexes["key"][(n_first, n_last, n_phone)]
            duplicate_reason = "Name + Telefon bereits vorhanden"

        if (
            not existing
            and n_birth
            and (n_first, n_last, n_birth) in match_indexes["birth"]
        ):
            existing = match_indexes["birth"][(n_first, n_last, n_birth)]
            duplicate_reason = "Name + Geburtstag bereits vorhanden"

        return existing, duplicate_reason

    @staticmethod
    def register_person_in_match_indexes(person, match_indexes):
        n_first = normalize_name(person.first_name)
        n_last = normalize_name(person.last_name)
        n_phone = normalize_phone(person.phone)
        n_email = normalize_email(person.email)
        n_birth = person.birthdate.isoformat() if person.birthdate else None

        if person.id is not None:
            match_indexes["id"][person.id] = person
        match_indexes["key"][(n_first, n_last, n_phone)] = person
        if n_email:
            match_indexes["email"][n_email] = person
        if n_birth:
            match_indexes["birth"][(n_first, n_last, n_birth)] = person
        if n_phone and n_birth:
            match_indexes["phone_birth"][(n_phone, n_birth)] = person

    @staticmethod
    def log_error(person_column, reason, data):
        logfile = os.path.join("logs", "import_errors.csv")
        os.makedirs(os.path.dirname(logfile), exist_ok=True)
        file_exists = os.path.isfile(logfile)

        with open(logfile, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=";")
            if not file_exists:
                writer.writerow(["timestamp", "column", "reason", "data"])
            writer.writerow([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                person_column,
                reason,
                str(data)
            ])

    # ---------------------------------------------------------
    # PREVIEW – an vertikale Vorlage (ab C3) angepasst
    # ---------------------------------------------------------
    @staticmethod
    def load_preview(file):
        try:
            wb = load_workbook(file, data_only=True)
        except Exception:
            return None, "Datei konnte nicht gelesen werden."

        ws = wb.active

        # Feldnamen aus Spalte B, ab Zeile 3
        fieldnames = []
        for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
            _, fieldname = row
            fieldnames.append(fieldname if fieldname else None)

        # Personen-Spalten: C–AA (3–27)
        person_columns = list(range(3, 28))

        preview_rows = []

        # Indexe für Duplikate
        existing_persons = Person.query.all()
        match_indexes = PersonImportService.build_match_indexes(existing_persons)

        # ---------------------------------------------------------
        # Jede Person aus Excel prüfen
        # ---------------------------------------------------------
        for col in person_columns:
            raw_data = {}
            cleaned_data = {}
            field_errors = {}
            field_warnings = {}
            notes = []

            # Rohdaten: ab Zeile 3
            for row_idx, fieldname in enumerate(fieldnames, start=3):
                if not fieldname:
                    continue
                value = ws.cell(row=row_idx, column=col).value
                raw_data[fieldname] = value

            # komplett leere Spalte überspringen
            if all(v in (None, "", " ") for v in raw_data.values()):
                continue

            # Pflichtfelder
            required_fields = [
                "first_name",
                "last_name",
                "phone",
                "email",
                "weight_kg",
                "emergency_name",
                "emergency_relation",
                "emergency_phone",
            ]

            for f in required_fields:
                v = raw_data.get(f)
                if not v or str(v).strip() == "":
                    field_errors[f] = True
                    notes.append(f"Pflichtfeld fehlt: {f}")

            # Rohdaten übernehmen
            cleaned_data.update(raw_data)

            # Normalisierung
            cleaned_data["email"] = normalize_email(raw_data.get("email"))
            if cleaned_data["email"] and not is_valid_email(cleaned_data["email"]):
                field_errors["email"] = True
                notes.append(f"E-Mail ungültig („{raw_data.get('email')}“) – Wert geleert")
                cleaned_data["email"] = ""

            cleaned_data["phone"] = normalize_phone(raw_data.get("phone"))
            cleaned_data["weight_kg"] = parse_weight_kg(raw_data.get("weight_kg"), notes, field_errors)
            cleaned_data["height_cm"] = parse_height_cm(raw_data.get("height_cm"), notes, field_errors)
            cleaned_data["birthdate"] = parse_birthdate(raw_data.get("birthdate"), notes, field_errors)
            cleaned_data["iban"] = parse_iban(raw_data.get("iban"), notes, field_errors, field_warnings)
            # Lizenznummer (optional, keine Logik/Validierung)
            cleaned_data["license_number"] = (raw_data.get("license_number") or "").strip() or None
            # Versicherung: Excel-Feld "insurance_company" → Modell-Feld "insurance_provider"
            cleaned_data["insurance_provider"] = (
                raw_data.get("insurance_company") or ""
            ).strip() or None

            # Mitglied/Tandem immer manuell prüfen
            cleaned_data["is_member"] = False
            cleaned_data["is_tandem_guest"] = False
            field_warnings["verify_member"] = True
            field_warnings["verify_tandem"] = True
            notes.append("Vereinsmitgliedschaft manuell prüfen")
            notes.append("Tandemstatus manuell prüfen")

            # ---------------------------------------------------------
            # Duplikaterkennung
            # ---------------------------------------------------------
            n_first = normalize_name(cleaned_data.get("first_name"))
            n_last = normalize_name(cleaned_data.get("last_name"))
            n_phone = normalize_phone(cleaned_data.get("phone"))
            n_email = normalize_email(cleaned_data.get("email"))
            n_birth = cleaned_data["birthdate"].isoformat() if cleaned_data.get("birthdate") else None

            existing, duplicate_reason = PersonImportService.find_existing_person(
                cleaned_data,
                match_indexes,
                import_mode="vertical",
            )

            is_duplicate = existing is not None

            if is_duplicate:
                field_warnings["duplicate"] = True
                notes.append(f"Duplikat erkannt: {duplicate_reason}")

            merge_candidates = (
                PersonImportService.build_merge_candidates(
                    existing,
                    cleaned_data,
                    import_mode="vertical",
                )
                if existing else []
            )
            existing_person = (
                PersonImportService.build_existing_person_preview(existing)
                if existing else None
            )

            if existing_person:
                notes.append(
                    f"Treffer: Person-ID {existing_person['id']} ({existing_person['full_name']})"
                )
                if merge_candidates:
                    changed_labels = ", ".join(c["label"] for c in merge_candidates)
                    notes.append(f"Abweichende Felder: {changed_labels}")
                else:
                    notes.append("Keine abweichenden Felder zum bestehenden Datensatz")

            preview_rows.append({
                "entry_key": f"col_{col}",
                "column": col,
                "raw_data": raw_data,
                "cleaned_data": cleaned_data,
                "field_errors": field_errors,
                "field_warnings": field_warnings,
                "notes": notes,
                "is_valid": len(field_errors) == 0,
                "is_duplicate": is_duplicate,
                "duplicate_reason": duplicate_reason,
                "existing_person": existing_person,
                "merge_candidates": merge_candidates,
            })

        return preview_rows, None

    @staticmethod
    def load_preview_horizontal(file):
        """
        Liest einen horizontalen Personen-Export ein (eine Person pro Zeile).
        Erwartetes Format: die durch PersonExportService.export_persons_excel erzeugten Spalten.
        """
        try:
            wb = load_workbook(file, data_only=True)
        except Exception:
            return None, "Datei konnte nicht gelesen werden."

        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return None, "Datei enthält keine Kopfzeile."

        col_to_field = {}
        for idx, header in enumerate(header_row):
            if header is None:
                continue
            key = str(header).strip().lower()
            field = PersonImportService.HORIZONTAL_HEADER_MAP.get(key)
            if field:
                col_to_field[idx] = field

        if "first_name" not in col_to_field.values() or "last_name" not in col_to_field.values():
            return None, "Die Datei entspricht nicht dem horizontalen Personen-Export (Kopfzeile unvollständig)."

        existing_persons = Person.query.all()
        match_indexes = PersonImportService.build_match_indexes(existing_persons)

        preview_rows = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw_data = {}
            cleaned_data = {}
            field_errors = {}
            field_warnings = {}
            notes = []

            for col_idx, field in col_to_field.items():
                raw_data[field] = row[col_idx] if col_idx < len(row) else None

            if all(v in (None, "", " ") for v in raw_data.values()):
                continue

            cleaned_data.update(raw_data)

            # Pflichtfelder (horizontales Backup-Format)
            for req in ("first_name", "last_name"):
                v = cleaned_data.get(req)
                if not v or str(v).strip() == "":
                    field_errors[req] = True
                    notes.append(f"Pflichtfeld fehlt: {req}")

            cleaned_data["first_name"] = (cleaned_data.get("first_name") or "").strip()
            cleaned_data["last_name"] = (cleaned_data.get("last_name") or "").strip()
            cleaned_data["phone"] = normalize_phone(cleaned_data.get("phone"))
            cleaned_data["email"] = normalize_email(cleaned_data.get("email"))

            if cleaned_data.get("email") and not is_valid_email(cleaned_data.get("email")):
                field_errors["email"] = True
                notes.append(f"E-Mail ungültig („{raw_data.get('email')}“) – Wert geleert")
                cleaned_data["email"] = ""

            cleaned_data["weight_kg"] = parse_weight_kg(cleaned_data.get("weight_kg"), notes, field_errors)
            cleaned_data["height_cm"] = parse_height_cm(cleaned_data.get("height_cm"), notes, field_errors)
            cleaned_data["birthdate"] = parse_birthdate(cleaned_data.get("birthdate"), notes, field_errors)
            cleaned_data["teacher_license_expires"] = parse_birthdate(
                cleaned_data.get("teacher_license_expires"), notes, field_errors
            )
            cleaned_data["liability_waiver_date"] = parse_birthdate(
                cleaned_data.get("liability_waiver_date"), notes, field_errors
            )
            cleaned_data["iban"] = parse_iban(cleaned_data.get("iban"), notes, field_errors, field_warnings)

            cleaned_data["is_member"] = PersonImportService._parse_bool(cleaned_data.get("is_member"))
            cleaned_data["is_partner_verein"] = PersonImportService._parse_bool(cleaned_data.get("is_partner_verein"))
            cleaned_data["is_tandem_guest"] = PersonImportService._parse_bool(cleaned_data.get("is_tandem_guest"))
            cleaned_data["is_teacher"] = PersonImportService._parse_bool(cleaned_data.get("is_teacher"))
            cleaned_data["is_aff_teacher"] = PersonImportService._parse_bool(cleaned_data.get("is_aff_teacher"))
            cleaned_data["is_aff_student"] = PersonImportService._parse_bool(cleaned_data.get("is_aff_student"))
            cleaned_data["is_video"] = PersonImportService._parse_bool(cleaned_data.get("is_video"))

            for txt_field in (
                "original_name",
                "street_and_number",
                "zip_code",
                "city",
                "license_number",
                "insurance_provider",
                "insurance_number",
                "emergency_name",
                "emergency_relation",
                "emergency_phone",
                "emergency_email",
                "bic",
                "account_holder",
                "comment",
                "notes",
            ):
                v = cleaned_data.get(txt_field)
                cleaned_data[txt_field] = (str(v).strip() if v not in (None, "") else None)

            existing, duplicate_reason = PersonImportService.find_existing_person(
                cleaned_data,
                match_indexes,
                import_mode="horizontal",
            )

            is_duplicate = existing is not None
            if is_duplicate:
                field_warnings["duplicate"] = True
                notes.append(f"Bestehender Datensatz gefunden: {duplicate_reason}")

            merge_candidates = (
                PersonImportService.build_merge_candidates(
                    existing,
                    cleaned_data,
                    import_mode="horizontal",
                )
                if existing else []
            )
            existing_person = (
                PersonImportService.build_existing_person_preview(existing)
                if existing else None
            )

            if existing_person:
                notes.append(
                    f"Treffer: Person-ID {existing_person['id']} ({existing_person['full_name']})"
                )
                if merge_candidates:
                    changed_labels = ", ".join(c["label"] for c in merge_candidates)
                    notes.append(f"Abweichende Felder: {changed_labels}")
                else:
                    notes.append("Keine abweichenden Felder zum bestehenden Datensatz")

            preview_rows.append({
                "entry_key": f"row_{row_idx}",
                "column": f"Zeile {row_idx}",
                "raw_data": raw_data,
                "cleaned_data": cleaned_data,
                "field_errors": field_errors,
                "field_warnings": field_warnings,
                "notes": notes,
                "is_valid": len(field_errors) == 0,
                "is_duplicate": is_duplicate,
                "duplicate_reason": duplicate_reason,
                "existing_person": existing_person,
                "merge_candidates": merge_candidates,
            })

        return preview_rows, None
